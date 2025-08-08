from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, g, send_file, current_app, abort, session
from flask_login import current_user
from app.utils.decorators import admin_required, mitarbeiter_required, login_required
from app.models.mongodb_models import MongoDBUser
from werkzeug.utils import secure_filename
import os
from pathlib import Path
import json
import colorsys
import logging
from datetime import datetime, timedelta
from app.utils.backup_manager import backup_manager
import openpyxl
from io import BytesIO
import time
from PIL import Image
from app.config.config import Config
from app.models.mongodb_database import mongodb, get_feature_settings, set_feature_setting, is_feature_enabled
from app.models.mongodb_models import MongoDBTool, MongoDBWorker, MongoDBConsumable
from bson import ObjectId
from werkzeug.security import generate_password_hash, check_password_hash
from app.utils.database_helpers import get_categories_from_settings, get_ticket_categories_from_settings, get_departments_from_settings, get_locations_from_settings
from docxtpl import DocxTemplate
from urllib.parse import unquote
import pandas as pd
import tempfile
from typing import Union
import re

# Import der neuen Services
from app.services.admin_dashboard_service import AdminDashboardService
from app.services.admin_user_service import AdminUserService
from app.services.admin_backup_service import AdminBackupService
from app.services.admin_system_service import AdminSystemService
from app.services.admin_email_service import AdminEmailService
from app.services.admin_notification_service import AdminNotificationService
from app.services.admin_ticket_service import AdminTicketService
from app.services.admin_debug_service import AdminDebugService
from app.services.admin_system_settings_service import AdminSystemSettingsService
from app.services.excel_export_service import ExcelExportService
from app.services.custom_fields_service import CustomFieldsService
from app.utils.permissions import get_role_permissions, set_role_permissions

from app.utils.id_helpers import convert_id_for_query, find_document_by_id, find_user_by_id

# Logger einrichten
logger = logging.getLogger(__name__)

bp = Blueprint('admin', __name__, url_prefix='/admin')

# Stelle sicher, dass die Standard-Einstellungen beim Start der App vorhanden sind
def ensure_default_settings():
    """Stellt sicher, dass die Standard-Label-Einstellungen vorhanden sind"""
    default_settings = [
        {'key': 'label_tickets_name', 'value': 'Tickets'},
        {'key': 'label_tickets_icon', 'value': 'fas fa-ticket-alt'},
        {'key': 'label_tools_name', 'value': 'Werkzeuge'},
        {'key': 'label_tools_icon', 'value': 'fas fa-tools'},
        {'key': 'label_consumables_name', 'value': 'Verbrauchsmaterial'},
        {'key': 'label_consumables_icon', 'value': 'fas fa-box-open'}
    ]
    
    for setting in default_settings:
        mongodb.update_one('settings', 
                         {'key': setting['key']}, 
                         {'$setOnInsert': setting}, 
                         upsert=True)

def create_excel(data, columns):
    """Erstellt eine Excel-Datei aus Daten"""
    wb = openpyxl.Workbook()
    ws = wb.active
    
    # Header
    for col, header in enumerate(columns, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Daten
    for row, item in enumerate(data, 2):
        for col, key in enumerate(columns.keys(), 1):
            value = item.get(key, '')
            ws.cell(row=row, column=col, value=value)
    
    # Speichern
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

def create_multi_sheet_excel(data_dict):
    """Erstellt eine Excel-Datei mit mehreren Arbeitsblättern"""
    wb = openpyxl.Workbook()
    
    # Entferne das Standard-Arbeitsblatt
    wb.remove(wb.active)
    
    for sheet_name, data in data_dict.items():
        ws = wb.create_sheet(title=sheet_name)
        
        if data and len(data) > 0:
            # Spezielle Behandlung für Tools und Consumables
            if sheet_name == 'Werkzeuge':
                _create_enhanced_tools_sheet(ws, data)
            elif sheet_name == 'Verbrauchsmaterial':
                _create_enhanced_consumables_sheet(ws, data)
            else:
                # Standard-Behandlung für andere Sheets
                headers = list(data[0].keys())
                
                # Header schreiben
                for col, header in enumerate(headers, 1):
                    ws.cell(row=1, column=col, value=header)
                
                # Daten schreiben
                for row, item in enumerate(data, 2):
                    for col, key in enumerate(headers, 1):
                        value = item.get(key, '')
                        ws.cell(row=row, column=col, value=value)
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

def _resolve_user_group_names(group_ids):
    """Löst Nutzergruppen-IDs zu Namen auf"""
    try:
        if not group_ids:
            return ''
        
        group_names = []
        for group_id in group_ids:
            try:
                from bson import ObjectId
                # Versuche ObjectId Konvertierung
                query_id = group_id
                if isinstance(group_id, str) and len(group_id) == 24:
                    try:
                        query_id = ObjectId(group_id)
                    except:
                        query_id = group_id
                
                # Lade Nutzergruppe aus Datenbank
                group = mongodb.find_one('user_groups', {'_id': query_id})
                if group:
                    group_names.append(group.get('name', str(group_id)))
                else:
                    group_names.append(str(group_id))
            except Exception:
                group_names.append(str(group_id))
        
        return ', '.join(group_names)
    except Exception:
        return ', '.join([str(gid) for gid in group_ids]) if group_ids else ''

def _create_enhanced_tools_sheet(ws, tools_data):
    """Erstellt eine erweiterte Tools-Tabelle mit allen Feldern"""
    try:
        # Lade Custom Fields für dynamische Header
        try:
            from app.services.custom_fields_service import CustomFieldsService
            custom_fields = CustomFieldsService.get_custom_fields_for_target('tools')
        except:
            custom_fields = []
        
        # Header definieren
        headers = [
            'barcode', 'name', 'category', 'location', 'description', 'status',
            'serial_number', 'invoice_number', 'mac_address', 'mac_address_wlan',
            'user_groups', 'additional_software', 'created_at', 'updated_at'
        ]
        
        # Custom Fields Header hinzufügen - speichere Mapping
        custom_field_mapping = {}
        for custom_field in custom_fields:
            header_name = custom_field['name']
            headers.append(header_name)
            custom_field_mapping[header_name] = custom_field['field_key']
        
        # Header schreiben
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Daten schreiben
        for row, tool in enumerate(tools_data, 2):
            for col, key in enumerate(headers, 1):
                if key in custom_field_mapping:
                    # Custom Field Wert
                    field_key = custom_field_mapping[key]
                    tool_custom_fields = tool.get('custom_fields', {})
                    value = tool_custom_fields.get(field_key, '')
                    
                    # Formatierung für Custom Fields
                    if isinstance(value, bool):
                        value = 'Ja' if value else 'Nein'
                    elif value is None:
                        value = ''
                elif key == 'user_groups':
                    # Nutzergruppen formatieren - Namen statt IDs
                    groups = tool.get('user_groups', [])
                    value = _resolve_user_group_names(groups)
                elif key == 'additional_software':
                    # Software formatieren
                    software = tool.get('additional_software', [])
                    value = ', '.join(software) if software else ''
                else:
                    # Standard-Wert
                    value = tool.get(key, '')
                
                ws.cell(row=row, column=col, value=value)
    except Exception as e:
        logger.error(f"Fehler beim Erstellen der erweiterten Tools-Tabelle: {str(e)}")

def _create_enhanced_consumables_sheet(ws, consumables_data):
    """Erstellt eine erweiterte Consumables-Tabelle mit allen Feldern"""
    try:
        # Lade Custom Fields für dynamische Header
        try:
            from app.services.custom_fields_service import CustomFieldsService
            custom_fields = CustomFieldsService.get_custom_fields_for_target('consumables')
        except:
            custom_fields = []
        
        # Header definieren
        headers = [
            'barcode', 'name', 'category', 'location', 'description', 'quantity',
            'min_quantity', 'created_at', 'updated_at'
        ]
        
        # Custom Fields Header hinzufügen - speichere Mapping
        custom_field_mapping = {}
        for custom_field in custom_fields:
            header_name = custom_field['name']
            headers.append(header_name)
            custom_field_mapping[header_name] = custom_field['field_key']
        
        # Header schreiben
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Daten schreiben
        for row, consumable in enumerate(consumables_data, 2):
            for col, key in enumerate(headers, 1):
                if key in custom_field_mapping:
                    # Custom Field Wert
                    field_key = custom_field_mapping[key]
                    consumable_custom_fields = consumable.get('custom_fields', {})
                    value = consumable_custom_fields.get(field_key, '')
                    
                    # Formatierung für Custom Fields
                    if isinstance(value, bool):
                        value = 'Ja' if value else 'Nein'
                    elif value is None:
                        value = ''
                else:
                    # Standard-Wert
                    value = consumable.get(key, '')
                
                ws.cell(row=row, column=col, value=value)
    except Exception as e:
        logger.error(f"Fehler beim Erstellen der erweiterten Consumables-Tabelle: {str(e)}")

@bp.route('/')
@mitarbeiter_required
def index():
    """Admin-Startseite"""
    return redirect(url_for('admin.dashboard'))

@bp.route('/dashboard')
@mitarbeiter_required
def dashboard():
    """Admin-Dashboard"""
    try:
        # Umfassende automatische Reparatur beim Dashboard-Aufruf
        try:
            from app.services.admin_debug_service import AdminDebugService
            from app.services.admin_backup_service import AdminBackupService
            
            # 1. Standard Dashboard-Reparatur
            AdminDebugService.fix_missing_created_at_fields()
            logger.info("Standard Dashboard-Reparatur durchgeführt")
            
            # 2. Umfassende Datentyp-Reparatur
            dashboard_fixes = AdminBackupService.fix_dashboard_after_backup()
            if dashboard_fixes.get('total', 0) > 0:
                logger.info(f"Umfassende Dashboard-Reparatur durchgeführt: {dashboard_fixes}")
            
            # 3. Zusätzliche Dashboard-spezifische Reparatur
            try:
                from app.models.mongodb_database import mongodb
                
                # Repariere spezifische Dashboard-Probleme
                dashboard_fixes_count = 0
                
                # Stelle sicher, dass alle Tools gültige Felder haben
                all_tools = mongodb.find('tools', {})
                for tool in all_tools:
                    try:
                        updates = {}
                        
                        # Stelle sicher, dass name Feld existiert
                        if 'name' not in tool or not tool['name']:
                            updates['name'] = tool.get('description', 'Unbekanntes Tool')
                        
                        # Stelle sicher, dass barcode Feld existiert
                        if 'barcode' not in tool:
                            updates['barcode'] = str(tool.get('_id', ''))
                        
                        # Stelle sicher, dass status Feld existiert
                        if 'status' not in tool:
                            updates['status'] = 'verfügbar'
                        
                        # Stelle sicher, dass location Feld existiert
                        if 'location' not in tool:
                            updates['location'] = 'Unbekannt'
                        
                        if updates:
                            mongodb.update_one('tools', {'_id': tool['_id']}, {'$set': updates})
                            dashboard_fixes_count += 1
                            
                    except Exception as e:
                        logger.warning(f"Fehler bei Tool-Reparatur: {e}")
                        continue
                
                # Stelle sicher, dass alle Workers gültige Felder haben
                all_workers = mongodb.find('workers', {})
                for worker in all_workers:
                    try:
                        updates = {}
                        
                        # Stelle sicher, dass name Feld existiert
                        if 'name' not in worker or not worker['name']:
                            firstname = worker.get('firstname', '')
                            lastname = worker.get('lastname', '')
                            if firstname or lastname:
                                updates['name'] = f"{firstname} {lastname}".strip()
                            else:
                                updates['name'] = 'Unbekannter Worker'
                        
                        # Stelle sicher, dass barcode Feld existiert
                        if 'barcode' not in worker:
                            updates['barcode'] = str(worker.get('_id', ''))
                        
                        if updates:
                            mongodb.update_one('workers', {'_id': worker['_id']}, {'$set': updates})
                            dashboard_fixes_count += 1
                            
                    except Exception as e:
                        logger.warning(f"Fehler bei Worker-Reparatur: {e}")
                        continue
                
                if dashboard_fixes_count > 0:
                    logger.info(f"Dashboard-spezifische Reparatur: {dashboard_fixes_count} Korrekturen")
                    
            except Exception as e:
                logger.warning(f"Dashboard-spezifische Reparatur fehlgeschlagen: {e}")
                
        except Exception as e:
            logger.warning(f"Automatische Dashboard-Reparatur fehlgeschlagen: {e}")
        
        # Verwende den AdminDashboardService für alle Dashboard-Daten
        recent_activity = AdminDashboardService.get_recent_activity()
        material_usage = AdminDashboardService.get_material_usage()
        warnings = AdminDashboardService.get_warnings()
        consumables_forecast = AdminDashboardService.get_consumables_forecast()
        consumable_trend = AdminDashboardService.get_consumable_trend()
        
        # Hole zusätzliche Statistiken
        try:
            total_tools = mongodb.count_documents('tools', {'deleted': {'$ne': True}})
            total_consumables = mongodb.count_documents('consumables', {'deleted': {'$ne': True}})
            total_workers = mongodb.count_documents('workers', {'deleted': {'$ne': True}})
            total_tickets = mongodb.count_documents('tickets', {})
            
            # Tool-Statistiken - Berücksichtige tatsächliche Ausleihen
            all_tools = list(mongodb.find('tools', {'deleted': {'$ne': True}}))
            current_lendings = list(mongodb.find('lendings', {'returned_at': {'$exists': False}}))
            
            # Debug: Zeige Ausleihen an
            logger.info(f"Dashboard Debug: {len(current_lendings)} aktuelle Ausleihen gefunden")
            for lending in current_lendings:
                logger.info(f"  Ausleihe: Tool={lending.get('tool_barcode')}, Worker={lending.get('worker_barcode')}, ID={lending.get('_id')}")
            
            # Erstelle Set der ausgeliehenen Tool-Barcodes (entferne Duplikate)
            lent_barcodes = set()
            for lending in current_lendings:
                tool_barcode = lending.get('tool_barcode')
                if tool_barcode:
                    lent_barcodes.add(tool_barcode)
            
            logger.info(f"Dashboard Debug: {len(lent_barcodes)} eindeutige ausgeliehene Tools: {list(lent_barcodes)}")
            
            available_count = 0
            lent_count = 0
            defect_count = 0
            
            for tool in all_tools:
                tool_barcode = tool.get('barcode')
                status = tool.get('status', 'verfügbar').lower()
                
                # Prüfe ob Tool tatsächlich ausgeliehen ist
                is_lent = tool_barcode in lent_barcodes
                
                if 'defekt' in status or 'defect' in status or 'broken' in status:
                    defect_count += 1
                elif is_lent:
                    lent_count += 1
                elif 'verfügbar' in status or 'available' in status:
                    available_count += 1
                else:
                    # Unbekannter Status - als verfügbar zählen
                    available_count += 1
            
            tool_stats = {
                'total': total_tools,
                'available': available_count,
                'lent': lent_count,
                'defect': defect_count
            }
            
            # Consumable-Statistiken - Verbesserte Logik
            consumables = list(mongodb.find('consumables', {'deleted': {'$ne': True}}))
            sufficient = 0
            warning = 0
            critical = 0
            
            for consumable in consumables:
                # Verwende verschiedene mögliche Feldnamen für den Bestand
                stock = consumable.get('stock', consumable.get('quantity', 0))
                warning_threshold = consumable.get('warning_threshold', 10)
                critical_threshold = consumable.get('critical_threshold', 5)
                
                # Konvertiere zu int falls nötig
                try:
                    stock = int(stock) if stock is not None else 0
                    warning_threshold = int(warning_threshold) if warning_threshold is not None else 10
                    critical_threshold = int(critical_threshold) if critical_threshold is not None else 5
                except (ValueError, TypeError):
                    stock = 0
                    warning_threshold = 10
                    critical_threshold = 5
                
                if stock >= warning_threshold:
                    sufficient += 1
                elif stock >= critical_threshold:
                    warning += 1
                else:
                    critical += 1
            
            consumable_stats = {
                'total': total_consumables,
                'sufficient': sufficient,
                'warning': warning,
                'critical': critical
            }
            
            # Worker-Statistiken
            workers = list(mongodb.find('workers', {'deleted': {'$ne': True}}))
            worker_stats = {
                'total': total_workers,
                'by_department': []
            }
            
            # Gruppiere nach Abteilung
            dept_counts = {}
            for worker in workers:
                dept = worker.get('department', 'Ohne Abteilung')
                dept_counts[dept] = dept_counts.get(dept, 0) + 1
            
            for dept, count in dept_counts.items():
                worker_stats['by_department'].append({
                    'name': dept,
                    'count': count
                })
            
            # Tool-Warnungen - Erweiterte Logik
            tool_warnings = []
            
            # Defekte Tools
            defect_tools = list(mongodb.find('tools', {'status': 'defekt', 'deleted': {'$ne': True}}))
            for tool in defect_tools:
                tool_warnings.append({
                    'name': tool.get('name', 'Unbekanntes Tool'),
                    'status': 'Defekt',
                    'severity': 'error'
                })
            
            # Warnung bei doppelten Ausleihen
            lending_counts = {}
            for lending in current_lendings:
                tool_barcode = lending.get('tool_barcode')
                if tool_barcode:
                    lending_counts[tool_barcode] = lending_counts.get(tool_barcode, 0) + 1
            
            duplicate_lendings = {barcode: count for barcode, count in lending_counts.items() if count > 1}
            if duplicate_lendings:
                logger.warning(f"Dashboard Debug: Doppelte Ausleihen gefunden: {duplicate_lendings}")
                for barcode, count in duplicate_lendings.items():
                    tool = mongodb.find_one('tools', {'barcode': barcode})
                    if tool:
                        tool_warnings.append({
                            'name': f"{tool.get('name', 'Unbekanntes Tool')} (Barcode: {barcode})",
                            'status': f'Doppelte Ausleihen: {count}x',
                            'severity': 'warning'
                        })
            
            # Überfällige Ausleihen (mehr als 30 Tage)
            overdue_lendings = []
            for lending in current_lendings:
                try:
                    lent_at = lending.get('lent_at')
                    if isinstance(lent_at, str):
                        try:
                            lent_at = datetime.strptime(lent_at, '%Y-%m-%d %H:%M:%S.%f')
                        except ValueError:
                            try:
                                lent_at = datetime.strptime(lent_at, '%Y-%m-%d %H:%M:%S')
                            except ValueError:
                                continue
                    
                    if isinstance(lent_at, datetime):
                        days_lent = (datetime.now() - lent_at).days
                        if days_lent > 30:
                            tool = mongodb.find_one('tools', {'barcode': lending.get('tool_barcode')})
                            worker = mongodb.find_one('workers', {'barcode': lending.get('worker_barcode')})
                            
                            if tool and worker:
                                overdue_lendings.append({
                                    'name': f"{tool.get('name', 'Unbekanntes Tool')} - {worker.get('name', 'Unbekannter Worker')}",
                                    'status': f'Überfällig ({days_lent} Tage)',
                                    'severity': 'warning'
                                })
                except Exception as e:
                    continue
            
            tool_warnings.extend(overdue_lendings)
            
            # Consumable-Warnungen - Verbesserte Logik
            consumable_warnings = []
            all_consumables = list(mongodb.find('consumables', {'deleted': {'$ne': True}}))
            
            for consumable in all_consumables:
                stock = consumable.get('stock', consumable.get('quantity', 0))
                try:
                    stock = int(stock) if stock is not None else 0
                except (ValueError, TypeError):
                    stock = 0
                
                if stock <= 0:
                    consumable_warnings.append({
                        'message': f"{consumable.get('name', 'Unbekanntes Verbrauchsmaterial')} (Bestand: {stock})",
                        'type': 'error',
                        'icon': 'times'
                    })
                elif stock < 5:
                    consumable_warnings.append({
                        'message': f"{consumable.get('name', 'Unbekanntes Verbrauchsmaterial')} (Bestand: {stock})",
                        'type': 'warning',
                        'icon': 'exclamation-triangle'
                    })
            
            # Aktuelle Ausleihen
            current_lendings = list(mongodb.find('lendings', {'returned_at': {'$exists': False}}))
            
            # Verarbeite Ausleihen für Anzeige
            processed_lendings = []
            for lending in current_lendings:
                try:
                    tool = mongodb.find_one('tools', {'barcode': lending.get('tool_barcode', '')})
                    worker = mongodb.find_one('workers', {'barcode': lending.get('worker_barcode', '')})
                    
                    if tool and worker:
                        # Sichere Datumsbehandlung
                        lent_at = lending.get('lent_at')
                        if isinstance(lent_at, str):
                            try:
                                lent_at = datetime.strptime(lent_at, '%Y-%m-%d %H:%M:%S.%f')
                            except ValueError:
                                try:
                                    lent_at = datetime.strptime(lent_at, '%Y-%m-%d %H:%M:%S')
                                except ValueError:
                                    lent_at = datetime.now()
                        elif not isinstance(lent_at, datetime):
                            lent_at = datetime.now()
                        
                        processed_lendings.append({
                            'tool_name': tool.get('name', 'Unbekanntes Tool'),
                            'worker_name': worker.get('name', 'Unbekannter Worker'),
                            'lent_at': lent_at,
                            'days_lent': (datetime.now() - lent_at).days
                        })
                except Exception as e:
                    logger.warning(f"Fehler bei Verarbeitung einer Ausleihe: {e}")
                    continue
            
            # Sortiere nach Ausleihdatum (älteste zuerst)
            processed_lendings.sort(key=lambda x: x.get('lent_at', datetime.now()))
            
        except Exception as e:
            logger.error(f"Fehler beim Laden der Dashboard-Statistiken: {e}")
            # Fallback-Werte
            tool_stats = {'total': 0, 'available': 0, 'lent': 0, 'defect': 0}
            consumable_stats = {'total': 0, 'sufficient': 0, 'warning': 0, 'critical': 0}
            worker_stats = {'total': 0, 'by_department': []}
            tool_warnings = []
            consumable_warnings = []
            processed_lendings = []
            total_tools = 0
            total_consumables = 0
            total_workers = 0
            total_tickets = 0
        
        return render_template('admin/dashboard.html',
                             recent_activity=recent_activity,
                             material_usage=material_usage,
                             warnings=warnings,
                             consumables_forecast=consumables_forecast,
                             consumable_trend=consumable_trend,
                             total_tools=total_tools,
                             total_consumables=total_consumables,
                             total_workers=total_workers,
                             total_tickets=total_tickets,
                             current_lendings=processed_lendings,
                             tool_stats=tool_stats,
                             consumable_stats=consumable_stats,
                             worker_stats=worker_stats,
                             tool_warnings=tool_warnings,
                             consumable_warnings=consumable_warnings)
        
    except Exception as e:
        logger.error(f"Fehler beim Laden des Dashboards: {e}")
        flash('Fehler beim Laden des Dashboards', 'error')
        return render_template('admin/dashboard.html',
                             recent_activity=[],
                             material_usage={'usage_data': [], 'period_days': 30},
                             warnings={'defect_tools': [], 'overdue_lendings': [], 'low_stock_consumables': []},
                             consumables_forecast=[],
                             consumable_trend={},
                             total_tools=0,
                             total_consumables=0,
                             total_workers=0,
                             total_tickets=0,
                             current_lendings=[],
                             tool_stats={'total': 0, 'available': 0, 'lent': 0, 'defect': 0},
                             consumable_stats={'total': 0, 'sufficient': 0, 'warning': 0, 'critical': 0},
                             worker_stats={'total': 0, 'by_department': []},
                             tool_warnings=[],
                             consumable_warnings=[])

@bp.route('/manual-lending', methods=['GET', 'POST'])
@mitarbeiter_required
def manual_lending():
    """Manuelle Ausleihe/Rückgabe"""
    if request.method == 'POST':
        logger.info("POST-Anfrage für manuelle Ausleihe empfangen")
        
        try:
            # JSON-Daten für manuelle Ausleihe
            data = request.get_json()
            
            if not data:
                return jsonify({'success': False, 'message': 'Keine Daten empfangen'}), 400
            
            # Validiere erforderliche Felder
            required_fields = ['item_barcode', 'worker_barcode', 'action', 'item_type']
            for field in required_fields:
                if field not in data:
                    return jsonify({'success': False, 'message': f'Feld {field} ist erforderlich'}), 400
            
            # Hole Daten aus JSON
            item_barcode = data.get('item_barcode', '').strip()
            worker_barcode = data.get('worker_barcode', '').strip()
            action = data.get('action', '').strip()
            item_type = data.get('item_type', '').strip()
            quantity = data.get('quantity', 1)
            
            if not item_barcode or not worker_barcode or not action or not item_type:
                return jsonify({'success': False, 'message': 'Alle Felder sind erforderlich'}), 400
            
            if action == 'lend' and not worker_barcode:
                return jsonify({
                    'success': False, 
                    'message': 'Mitarbeiter muss ausgewählt sein'
                }), 400
            
            try:
                # Prüfe ob der Mitarbeiter existiert
                if worker_barcode:
                    worker = mongodb.find_one('workers', {'barcode': worker_barcode, 'deleted': {'$ne': True}})
                    if not worker:
                        return jsonify({
                            'success': False,
                            'message': 'Mitarbeiter nicht gefunden'
                        }), 404

                # Verwende den zentralen LendingService für konsistente Verarbeitung
                from app.services.lending_service import LendingService
                
                # Erstelle Request-Daten für den Service
                service_data = {
                    'item_barcode': item_barcode,
                    'worker_barcode': worker_barcode,
                    'action': action,
                    'item_type': item_type,
                    'quantity': quantity
                }
                
                # Verarbeite über den Service
                success, message, result_data = LendingService.process_lending_request(service_data)
                
                if success:
                    return jsonify({
                        'success': True,
                        'message': message,
                        'data': result_data
                    })
                else:
                    return jsonify({
                        'success': False,
                        'message': message
                    }), 400
            except Exception as e:
                logger.error(f"Fehler bei der Ausleihe: {str(e)}")
                return jsonify({
                    'success': False,
                    'message': f'Fehler: {str(e)}'
                }), 500
            
        except Exception as e:
            logger.error(f"Fehler beim Verarbeiten der Anfrage: {str(e)}")
            return jsonify({
                'success': False,
                'message': 'Fehler beim Verarbeiten der Anfrage'
            }), 500
            
    # GET request - zeige das Formular
    try:
        # Hole alle verfügbaren Werkzeuge
        tools_pipeline = [
            {'$match': {'deleted': {'$ne': True}}},
            {
                '$lookup': {
                    'from': 'lendings',
                    'localField': 'barcode',
                    'foreignField': 'tool_barcode',
                    'as': 'active_lendings'
                }
            },
            {
                '$addFields': {
                    'current_status': {
                        '$cond': [
                            {'$gt': [{'$size': {'$filter': {'input': '$active_lendings', 'cond': {'$eq': ['$$this.returned_at', None]}}}}, 0]},
                            'ausgeliehen',
                            '$status'
                        ]
                    }
                }
            },
            {'$sort': {'name': 1}}
        ]
        
        tools = list(mongodb.aggregate('tools', tools_pipeline))
        
        # Hole alle Mitarbeiter
        workers = mongodb.find('workers', {'deleted': {'$ne': True}}, sort=[('firstname', 1)])
        
        # Verbrauchsmaterialien laden
        consumables = mongodb.find('consumables', {'deleted': {'$ne': True}}, sort=[('name', 1)])

        # Hole aktuelle Ausleihen
        current_lendings = []
        
        # Aktuelle Werkzeug-Ausleihen
        active_tool_lendings = mongodb.find('lendings', {'returned_at': None})
        for lending in active_tool_lendings:
            tool = mongodb.find_one('tools', {'barcode': lending['tool_barcode']})
            worker = mongodb.find_one('workers', {'barcode': lending['worker_barcode']})
            
            if tool and worker:
                current_lendings.append({
                    'item_name': tool['name'],
                    'item_barcode': tool['barcode'],
                    'worker_name': f"{worker['firstname']} {worker['lastname']}",
                    'worker_barcode': worker['barcode'],
                    'action_date': lending['lent_at'],
                    'category': 'Werkzeug',
                    'amount': None
                })

        # Aktuelle Verbrauchsmaterial-Ausgaben (letzte 30 Tage)
        thirty_days_ago = datetime.now() - timedelta(days=30)
        recent_consumable_usages = mongodb.find('consumable_usages', {
            'used_at': {'$gte': thirty_days_ago},
            'quantity': {'$lt': 0}  # Nur Ausgaben (negative Werte), nicht Entnahmen
        })
        
        for usage in recent_consumable_usages:
            consumable = mongodb.find_one('consumables', {'barcode': usage['consumable_barcode']})
            worker = mongodb.find_one('workers', {'barcode': usage['worker_barcode']})
            
            if consumable and worker:
                current_lendings.append({
                    'item_name': consumable['name'],
                    'item_barcode': consumable['barcode'],
                    'worker_name': f"{worker['firstname']} {worker['lastname']}",
                    'worker_barcode': worker['barcode'],
                    'action_date': usage['used_at'],
                    'category': 'Verbrauchsmaterial',
                    'amount': usage['quantity']
                })

        # Sortiere nach Datum (neueste zuerst)
        def safe_date_key(lending):
            action_date = lending.get('action_date')
            if isinstance(action_date, str):
                try:
                    return datetime.strptime(action_date, '%Y-%m-%d %H:%M:%S')
                except (ValueError, TypeError):
                    return datetime.min
            elif isinstance(action_date, datetime):
                return action_date
            else:
                return datetime.min
        
        current_lendings.sort(key=safe_date_key, reverse=True)

        return render_template('admin/manual_lending.html', 
                              tools=tools,
                              workers=workers,
                              consumables=consumables,
                              current_lendings=current_lendings)
    except Exception as e:
        print(f"Fehler beim Laden der Daten: {e}")
        flash('Fehler beim Laden der Daten', 'error')
        return render_template('admin/manual_lending.html', 
                              tools=[], 
                              workers=[], 
                              consumables=[],
                              current_lendings=[])

@bp.route('/trash')
@mitarbeiter_required
def trash():
    """Zeigt den Papierkorb mit gelöschten Einträgen an"""
    try:
        # Gelöschte Werkzeuge
        tools = mongodb.find('tools', {'deleted': True}, sort=[('deleted_at', -1)])
        
        # Gelöschte Verbrauchsmaterialien
        consumables = mongodb.find('consumables', {'deleted': True}, sort=[('deleted_at', -1)])
        
        # Gelöschte Mitarbeiter
        workers = mongodb.find('workers', {'deleted': True}, sort=[('deleted_at', -1)])
        
        # Gelöschte Tickets
        tickets = mongodb.find('tickets', {'deleted': True}, sort=[('deleted_at', -1)])
        
        return render_template('admin/trash.html',
                           tools=tools,
                           consumables=consumables,
                           workers=workers,
                           tickets=tickets)
    except Exception as e:
        logger.error(f"Fehler beim Laden des Papierkorbs: {str(e)}", exc_info=True)
        flash('Fehler beim Laden des Papierkorbs', 'error')
        return redirect(url_for('admin.dashboard'))

@bp.route('/trash/restore/<type>/<barcode>', methods=['POST'])
@mitarbeiter_required
def restore_item(type, barcode):
    """Stellt einen gelöschten Eintrag wieder her"""
    try:
        # Barcode URL-dekodieren
        decoded_barcode = unquote(barcode)
        
        if type == 'tool':
            # Prüfe ob das Werkzeug existiert
            tool = mongodb.find_one('tools', {'barcode': decoded_barcode, 'deleted': True})
            
            if not tool:
                return jsonify({
                    'success': False,
                    'message': 'Werkzeug nicht gefunden'
                }), 404
                
            # Stelle das Werkzeug wieder her
            mongodb.update_one('tools', 
                             {'barcode': decoded_barcode}, 
                             {'$set': {'deleted': False, 'deleted_at': None}})
            
        elif type == 'consumable':
            # Prüfe ob das Verbrauchsmaterial existiert
            consumable = mongodb.find_one('consumables', {'barcode': decoded_barcode, 'deleted': True})
            
            if not consumable:
                return jsonify({
                    'success': False,
                    'message': 'Verbrauchsmaterial nicht gefunden'
                }), 404
                
            # Stelle das Verbrauchsmaterial wieder her
            mongodb.update_one('consumables', 
                             {'barcode': decoded_barcode}, 
                             {'$set': {'deleted': False, 'deleted_at': None}})
            
        elif type == 'worker':
            # Prüfe ob der Mitarbeiter existiert
            worker = mongodb.find_one('workers', {'barcode': decoded_barcode, 'deleted': True})
            
            if not worker:
                return jsonify({
                    'success': False,
                    'message': 'Mitarbeiter nicht gefunden'
                }), 404
                
            # Stelle den Mitarbeiter wieder her
            mongodb.update_one('workers', 
                             {'barcode': decoded_barcode}, 
                             {'$set': {'deleted': False, 'deleted_at': None}})
                             
        elif type == 'ticket':
            # Prüfe ob das Ticket existiert
            ticket = mongodb.find_one('tickets', {'_id': convert_id_for_query(decoded_barcode), 'deleted': True})
            
            if not ticket:
                return jsonify({
                    'success': False,
                    'message': 'Ticket nicht gefunden'
                }), 404
                
            # Stelle das Ticket wieder her
            mongodb.update_one('tickets', 
                             {'_id': convert_id_for_query(decoded_barcode)}, 
                             {'$set': {'deleted': False, 'deleted_at': None}})
        else:
            return jsonify({
                'success': False,
                'message': 'Ungültiger Typ'
            }), 400
            
        return jsonify({
            'success': True,
            'message': 'Eintrag wurde wiederhergestellt'
        })
        
    except Exception as e:
        logger.error(f"Fehler beim Wiederherstellen des Eintrags: {str(e)}", exc_info=True)
        return jsonify({
            'success': False,
            'message': f'Fehler beim Wiederherstellen: {str(e)}'
        }), 500

@bp.route('/tools/delete', methods=['DELETE'])
@mitarbeiter_required
def delete_tool_soft_json():
    """Werkzeug soft löschen (markieren als gelöscht) - Barcode aus JSON-Body"""
    try:
        data = request.get_json()
        if not data or 'barcode' not in data:
            return jsonify({'success': False, 'message': 'Kein Barcode angegeben'}), 400
            
        barcode = data['barcode'].strip()  # Barcode bereinigen
        if len(barcode) > 50:
            return jsonify({'success': False, 'message': 'Barcode zu lang (max. 50 Zeichen)'}), 400
        
        # Prüfe ob das Werkzeug existiert
        tool = mongodb.find_one('tools', {'barcode': barcode, 'deleted': {'$ne': True}})
        
        if not tool:
            return jsonify({
                'success': False,
                'message': 'Werkzeug nicht gefunden'
            }), 404
            
        # Prüfe ob das Werkzeug ausgeliehen ist
        active_lending = mongodb.find_one('lendings', {
            'tool_barcode': barcode, 
            'returned_at': None
        })
        
        if active_lending:
            return jsonify({
                'success': False,
                'message': 'Werkzeug ist noch ausgeliehen und kann nicht gelöscht werden'
            }), 400
            
        # Führe das Soft-Delete durch
        mongodb.update_one('tools', {'barcode': barcode}, {
            '$set': {
                'deleted': True,
                'deleted_at': datetime.now()
            }
        })
        
        return jsonify({
            'success': True,
            'message': 'Werkzeug wurde erfolgreich gelöscht'
        })

    except Exception as e:
        logger.error(f"Fehler beim Löschen des Werkzeugs: {str(e)}", exc_info=True)
        return jsonify({
            'success': False,
            'message': f'Fehler beim Löschen: {str(e)}'
        }), 500

@bp.route('/tools/<barcode>/delete', methods=['DELETE'])
@mitarbeiter_required
def delete_tool_soft(barcode):
    """Werkzeug soft löschen (markieren als gelöscht) - Barcode aus URL (Legacy)"""
    try:
        # Barcode URL-dekodieren
        decoded_barcode = unquote(barcode)
        
        # Prüfe ob das Werkzeug existiert
        tool = mongodb.find_one('tools', {'barcode': decoded_barcode, 'deleted': {'$ne': True}})
        
        if not tool:
            return jsonify({
                'success': False,
                'message': 'Werkzeug nicht gefunden'
            }), 404
            
        # Prüfe ob das Werkzeug ausgeliehen ist
        active_lending = mongodb.find_one('lendings', {
            'tool_barcode': decoded_barcode, 
            'returned_at': None
        })
        
        if active_lending:
            return jsonify({
                'success': False,
                'message': 'Werkzeug ist noch ausgeliehen und kann nicht gelöscht werden'
            }), 400
            
        # Führe das Soft-Delete durch
        mongodb.update_one('tools', {'barcode': decoded_barcode}, {
            '$set': {
                'deleted': True,
                'deleted_at': datetime.now()
            }
        })
        
        return jsonify({
            'success': True,
            'message': 'Werkzeug wurde erfolgreich gelöscht'
        })
        
    except Exception as e:
        logger.error(f"Fehler beim Löschen des Werkzeugs: {str(e)}", exc_info=True)
        return jsonify({
            'success': False,
            'message': f'Fehler beim Löschen: {str(e)}'
        }), 500

@bp.route('/tools/<barcode>/delete-permanent', methods=['DELETE'])
@mitarbeiter_required
def delete_tool_permanent(barcode):
    """Werkzeug endgültig löschen"""
    try:
        # Barcode URL-dekodieren
        decoded_barcode = unquote(barcode)
        
        # Prüfe ob das Werkzeug existiert und gelöscht ist
        tool = mongodb.find_one('tools', {'barcode': decoded_barcode, 'deleted': True})
        
        if not tool:
            return jsonify({
                'success': False,
                'message': 'Werkzeug nicht gefunden oder nicht gelöscht'
            }), 404
            
        # Lösche das Werkzeug endgültig
        mongodb.delete_one('tools', {'barcode': decoded_barcode})
        
        return jsonify({
            'success': True,
            'message': 'Werkzeug wurde endgültig gelöscht'
        })
        
    except Exception as e:
        logger.error(f"Fehler beim endgültigen Löschen des Werkzeugs: {str(e)}", exc_info=True)
        return jsonify({
            'success': False,
            'message': f'Fehler beim Löschen: {str(e)}'
        }), 500

@bp.route('/consumables/delete', methods=['DELETE'])
@mitarbeiter_required
def delete_consumable_soft():
    try:
        data = request.get_json()
        if not data or 'barcode' not in data:
            return jsonify({'success': False, 'message': 'Kein Barcode angegeben'}), 400
            
        barcode = data['barcode'].strip()  # Barcode bereinigen
        if len(barcode) > 50:
            return jsonify({'success': False, 'message': 'Barcode zu lang (max. 50 Zeichen)'}), 400
            
        # Prüfe ob das Verbrauchsmaterial existiert
        consumable = mongodb.find_one('consumables', {'barcode': barcode, 'deleted': {'$ne': True}})
        if not consumable:
            return jsonify({'success': False, 'message': 'Verbrauchsmaterial nicht gefunden'}), 404
            
        # Führe das Soft-Delete durch
        mongodb.update_one('consumables', {'barcode': barcode}, {
            '$set': {
                'deleted': True, 
                'deleted_at': datetime.now()
            }
        })
        return jsonify({'success': True, 'message': 'Verbrauchsmaterial erfolgreich gelöscht'})
        
    except Exception as e:
        logger.error(f"Fehler beim Löschen des Verbrauchsmaterials: {e}")
        return jsonify({'success': False, 'message': 'Interner Serverfehler'}), 500

@bp.route('/consumables/<barcode>/delete-permanent', methods=['DELETE'])
@mitarbeiter_required
def delete_consumable_permanent(barcode):
    try:
        # Barcode URL-dekodieren
        decoded_barcode = unquote(barcode)
        
        # Prüfe ob das Verbrauchsmaterial existiert und gelöscht ist
        consumable = mongodb.find_one('consumables', {'barcode': decoded_barcode, 'deleted': True})
        if not consumable:
            return jsonify({'success': False, 'message': 'Verbrauchsmaterial nicht gefunden oder nicht gelöscht'}), 404
            
        # Führe das permanente Löschen durch
        mongodb.delete_one('consumables', {'barcode': decoded_barcode})
        return jsonify({'success': True, 'message': 'Verbrauchsmaterial permanent gelöscht'})
        
    except Exception as e:
        logger.error(f"Fehler beim permanenten Löschen des Verbrauchsmaterials: {e}")
        return jsonify({'success': False, 'message': 'Interner Serverfehler'}), 500

@bp.route('/workers/delete', methods=['DELETE'])
@mitarbeiter_required
def delete_worker_soft_json():
    """Mitarbeiter soft löschen (markieren als gelöscht) - Barcode aus JSON-Body"""
    try:
        data = request.get_json()
        if not data or 'barcode' not in data:
            return jsonify({'success': False, 'message': 'Kein Barcode angegeben'}), 400
            
        barcode = data['barcode'].strip()  # Barcode bereinigen
        if len(barcode) > 50:
            return jsonify({'success': False, 'message': 'Barcode zu lang (max. 50 Zeichen)'}), 400
        
        # Prüfe ob der Mitarbeiter existiert
        worker = mongodb.find_one('workers', {'barcode': barcode, 'deleted': {'$ne': True}})
        
        if not worker:
            return jsonify({
                'success': False,
                'message': 'Mitarbeiter nicht gefunden'
            }), 404
            
        # Prüfe ob der Mitarbeiter aktive Ausleihen hat
        active_lendings_count = mongodb.count_documents('lendings', {
            'worker_barcode': barcode, 
            'returned_at': None
        })
        
        if active_lendings_count > 0:
            return jsonify({
                'success': False,
                'message': 'Mitarbeiter hat noch aktive Ausleihen'
            }), 400
            
        # Führe das Soft-Delete durch
        mongodb.update_one('workers', {'barcode': barcode}, {
            '$set': {
                'deleted': True,
                'deleted_at': datetime.now()
            }
        })
        
        return jsonify({
            'success': True,
            'message': 'Mitarbeiter wurde erfolgreich gelöscht'
        })
        
    except Exception as e:
        logger.error(f"Fehler beim Löschen des Mitarbeiters: {str(e)}", exc_info=True)
        return jsonify({
            'success': False,
            'message': f'Fehler beim Löschen: {str(e)}'
        }), 500

@bp.route('/workers/<barcode>/delete', methods=['DELETE'])
@mitarbeiter_required
def delete_worker_soft(barcode):
    """Mitarbeiter soft löschen (markieren als gelöscht) - Barcode aus URL (Legacy)"""
    try:
        # Barcode URL-dekodieren
        decoded_barcode = unquote(barcode)
        
        # Prüfe ob der Mitarbeiter existiert
        worker = mongodb.find_one('workers', {'barcode': decoded_barcode, 'deleted': {'$ne': True}})
        
        if not worker:
            return jsonify({
                'success': False,
                'message': 'Mitarbeiter nicht gefunden'
            }), 404
            
        # Prüfe ob der Mitarbeiter aktive Ausleihen hat
        active_lendings_count = mongodb.count_documents('lendings', {
            'worker_barcode': decoded_barcode, 
            'returned_at': None
        })
        
        if active_lendings_count > 0:
            return jsonify({
                'success': False,
                'message': 'Mitarbeiter hat noch aktive Ausleihen'
            }), 400
            
        # Führe das Soft-Delete durch
        mongodb.update_one('workers', {'barcode': decoded_barcode}, {
            '$set': {
                'deleted': True,
                'deleted_at': datetime.now()
            }
        })
        
        return jsonify({
            'success': True,
            'message': 'Mitarbeiter wurde erfolgreich gelöscht'
        })
        
    except Exception as e:
        logger.error(f"Fehler beim Löschen des Mitarbeiters: {str(e)}", exc_info=True)
        return jsonify({
            'success': False,
            'message': f'Fehler beim Löschen: {str(e)}'
        }), 500

@bp.route('/workers/<barcode>/delete-permanent', methods=['DELETE'])
@mitarbeiter_required
def delete_worker_permanent(barcode):
    try:
        # Barcode URL-dekodieren
        decoded_barcode = unquote(barcode)
        
        # Prüfe ob der Mitarbeiter existiert (auch gelöschte)
        worker = mongodb.find_one('workers', {'barcode': decoded_barcode})
        if not worker:
            return jsonify({'success': False, 'message': 'Mitarbeiter nicht gefunden'}), 404
            
        # Prüfe ob der Mitarbeiter aktive Ausleihen hat
        active_lendings_count = mongodb.count_documents('lendings', {
            'worker_barcode': decoded_barcode, 
            'returned_at': None
        })
        if active_lendings_count > 0:
            return jsonify({'success': False, 'message': 'Mitarbeiter hat noch aktive Ausleihen'}), 400
            
        # Führe das permanente Löschen durch
        mongodb.delete_one('workers', {'barcode': decoded_barcode})
        return jsonify({'success': True, 'message': 'Mitarbeiter permanent gelöscht'})
        
    except Exception as e:
        logger.error(f"Fehler beim permanenten Löschen des Mitarbeiters: {e}")
        return jsonify({'success': False, 'message': 'Interner Serverfehler'}), 500

@bp.route('/tickets/<ticket_id>')
@login_required
@mitarbeiter_required
def ticket_detail(ticket_id):
    """Zeigt die Details eines Tickets für Administratoren."""
    ticket = mongodb.find_one('tickets', {'_id': convert_id_for_query(ticket_id)})
    
    if not ticket:
        return render_template('404.html'), 404
    
    # Füge id-Feld hinzu (für Template-Kompatibilität)
    ticket['id'] = str(ticket['_id'])
        
    # Konvertiere alle Datumsfelder zu datetime-Objekten
    date_fields = ['created_at', 'updated_at', 'resolved_at', 'due_date']
    for field in date_fields:
        if ticket.get(field):
            try:
                ticket[field] = datetime.strptime(ticket[field], '%Y-%m-%d %H:%M:%S')
            except (ValueError, TypeError):
                ticket[field] = None
        
    # Hole die Notizen für das Ticket
    notes = mongodb.find('ticket_notes', {'ticket_id': convert_id_for_query(ticket_id)})

    # Hole die Nachrichten für das Ticket
    messages = mongodb.find('ticket_messages', {'ticket_id': convert_id_for_query(ticket_id)})
    # Formatiere die Nachrichten für das Template
    formatted_messages = []
    for msg in messages:
        formatted_msg = dict(msg)
        # Konvertiere created_at zu String-Format
        if isinstance(formatted_msg.get('created_at'), datetime):
            formatted_msg['created_at'] = formatted_msg['created_at'].strftime('%d.%m.%Y %H:%M')
        # Setze is_admin Flag basierend auf dem Sender
        formatted_msg['is_admin'] = formatted_msg.get('sender') == current_user.username
        formatted_messages.append(formatted_msg)

    # Hole die Auftragsdetails
    auftrag_details = mongodb.find_one('auftrag_details', {'ticket_id': convert_id_for_query(ticket_id)})
    logging.info(f"DEBUG: auftrag_details für Ticket {ticket_id}: {auftrag_details}")
    
    # Hole die Materialliste
    material_list = mongodb.find('auftrag_material', {'ticket_id': convert_id_for_query(ticket_id)})

    # Hole alle Benutzer aus der Hauptdatenbank und wandle sie in Dicts um
    users = mongodb.find('users', {'is_active': True})
    users = [dict(user) for user in users]

    # Hole alle zugewiesenen Nutzer (Mehrfachzuweisung)
    assigned_users = mongodb.find('ticket_assignments', {'ticket_id': convert_id_for_query(ticket_id)})

    # Hole alle Kategorien aus der settings Collection
    categories = get_ticket_categories_from_settings()

    # Hole Arbeitszeiten
    arbeit_list = list(mongodb.find('auftrag_arbeit', {'ticket_id': convert_id_for_query(ticket_id)}))
    
    # Hole Notizen
    notes = list(mongodb.find('ticket_notes', {'ticket_id': convert_id_for_query(ticket_id)}, sort=[('created_at', -1)]))
    
    # Hole Nachrichten
    messages = list(mongodb.find('ticket_messages', {'ticket_id': convert_id_for_query(ticket_id)}, sort=[('created_at', -1)]))
    
    # Hole alle Mitarbeiter für Zuweisung
    workers = list(mongodb.find('workers', {'deleted': {'$ne': True}}, sort=[('lastname', 1)]))
    
    # Hole alle Abteilungen
    departments = get_departments_from_settings()
    
    # Hole alle Standorte
    locations = get_locations_from_settings()

    return render_template('admin/ticket_detail.html', 
                         ticket=ticket, 
                         notes=notes,
                         messages=formatted_messages,
                         users=users,
                         workers=users,  # Template erwartet 'workers'
                         assigned_users=assigned_users,
                         auftrag_details=auftrag_details,
                         material_list=material_list,
                         categories=categories,
                         now=datetime.now(),
                         arbeit_list=arbeit_list)

@bp.route('/tickets/<ticket_id>/message', methods=['POST'])
@login_required
@admin_required
def add_ticket_message(ticket_id):
    """Fügt eine neue Nachricht zu einem Ticket hinzu."""
    try:
        # Hole die Nachricht aus dem Request
        if not request.is_json:
            return jsonify({
                'success': False,
                'message': 'Ungültiges Anfrageformat. JSON erwartet.'
            }), 400

        data = request.get_json()
        message = data.get('message', '').strip()
        
        if not message:
            return jsonify({
                'success': False,
                'message': 'Nachricht darf nicht leer sein'
            }), 400

        # Verwende den AdminTicketService
        success, result_message = AdminTicketService.add_ticket_message(ticket_id, message, 'message')
        
        if success:
            return jsonify({
                'success': True,
                'message': {
                    'sender': current_user.username,
                    'text': message,
                    'created_at': datetime.now().strftime('%d.%m.%Y %H:%M')
                }
            })
        else:
            return jsonify({
                'success': False,
                'message': result_message
            }), 400

    except Exception as e:
        logger.error(f"Fehler beim Hinzufügen der Nachricht: {str(e)}", exc_info=True)
        return jsonify({
            'success': False,
            'message': f'Fehler beim Speichern der Nachricht: {str(e)}'
        }), 500

@bp.route('/tickets/<ticket_id>/note', methods=['POST'])
@login_required
@admin_required
def add_ticket_note(ticket_id):
    """Fügt eine neue Notiz zu einem Ticket hinzu"""
    try:
        if not request.is_json:
            return jsonify({'success': False, 'message': 'Ungültiges Anfrageformat'}), 400

        data = request.get_json()
        note_text = data.get('note', '').strip()
        
        if not note_text:
            return jsonify({'success': False, 'message': 'Notiz darf nicht leer sein'}), 400

        # Erstelle die Notiz
        note_data = {
            'ticket_id': convert_id_for_query(ticket_id),
            'note': note_text,
            'created_by': current_user.username,
            'created_at': datetime.now()
        }
        
        result = mongodb.insert_one('ticket_notes', note_data)
        
        if not result:
            return jsonify({'success': False, 'message': 'Fehler beim Speichern der Notiz'}), 500

        return jsonify({
            'success': True,
            'message': 'Notiz erfolgreich hinzugefügt',
            'note': {
                'id': str(result),
                'text': note_text,
                'created_by': current_user.username,
                'created_at': datetime.now().strftime('%d.%m.%Y %H:%M')
            }
        })

    except Exception as e:
        logger.error(f"Fehler beim Hinzufügen der Notiz: {str(e)}")
        return jsonify({'success': False, 'message': str(e)}), 500

@bp.route('/tickets/<ticket_id>/update', methods=['POST'])
@login_required
@admin_required
def update_ticket(ticket_id):
    """Aktualisiert ein Ticket"""
    try:
        data = request.get_json()
        logging.info(f"Empfangene Daten für Ticket {ticket_id}: {data}")
        
        # Verarbeite ausgeführte Arbeiten
        arbeit_list = data.get('arbeit_list', [])
        ausgefuehrte_arbeiten = '\n'.join([
            f"{arbeit['arbeit']}|{arbeit['arbeitsstunden']}|{arbeit['leistungskategorie']}"
            for arbeit in arbeit_list
        ])
        logging.info(f"Verarbeitete ausgeführte Arbeiten: {ausgefuehrte_arbeiten}")
        
        # Bereite die Auftragsdetails vor
        auftrag_details = {
            'bereich': data.get('bereich', ''),
            'auftraggeber_intern': data.get('auftraggeber_typ') == 'intern',
            'auftraggeber_extern': data.get('auftraggeber_typ') == 'extern',
            'auftraggeber_name': data.get('auftraggeber_name', ''),
            'kontakt': data.get('kontakt', ''),
            'auftragsbeschreibung': data.get('auftragsbeschreibung', ''),
            'ausgefuehrte_arbeiten': ausgefuehrte_arbeiten,
            'arbeitsstunden': data.get('arbeitsstunden', ''),
            'leistungskategorie': data.get('leistungskategorie', ''),
            'fertigstellungstermin': data.get('fertigstellungstermin', ''),
            'gesamtsumme': data.get('gesamtsumme', 0)
        }
        
        # Verwende die ursprüngliche ID direkt für das Update
        from bson import ObjectId
        try:
            # Versuche zuerst mit ObjectId
            ticket_id_for_update = ObjectId(ticket_id)
        except:
            # Falls das fehlschlägt, verwende die ursprüngliche ID als String
            ticket_id_for_update = ticket_id
        
        # Aktualisiere die Auftragsdetails
        if not mongodb.update_one('auftrag_details', {'ticket_id': ticket_id_for_update}, {'$set': auftrag_details}):
            return jsonify({'success': False, 'message': 'Fehler beim Aktualisieren der Auftragsdetails'})
        
        # Aktualisiere die Materialliste
        material_list = data.get('material_list', [])
        if not mongodb.update_many('auftrag_material', {'ticket_id': ticket_id_for_update}, {'$set': {'menge': m['menge'], 'einzelpreis': m['einzelpreis']} for m in material_list}):
            return jsonify({'success': False, 'message': 'Fehler beim Aktualisieren der Materialliste'})
        
        return jsonify({'success': True})
        
    except Exception as e:
        logging.error(f"Fehler beim Aktualisieren des Tickets {ticket_id}: {str(e)}")
        return jsonify({'success': False, 'message': str(e)})

@bp.route('/tickets/<id>/export')
@login_required
@admin_required
def export_ticket(id):
    """Exportiert das Ticket als ausgefülltes Word-Dokument."""
    try:
        # Verwende den AdminTicketService
        success, message, file_path = AdminTicketService.export_ticket_as_word(id)
        
        if success and file_path:
            # Sende das Dokument
            ticket = find_document_by_id('tickets', id)
            ticket_number = ticket.get('ticket_number', id) if ticket else id
            return send_file(file_path, as_attachment=True, download_name=f'ticket_{ticket_number}_export.docx')
        else:
            flash(message, 'error')
            return redirect(url_for('admin.ticket_detail', ticket_id=id))
        
    except Exception as e:
        logging.error(f"Fehler beim Generieren des Word-Dokuments: {str(e)}", exc_info=True)
        flash(f'Fehler beim Generieren des Dokuments: {str(e)}', 'error')
        return redirect(url_for('admin.ticket_detail', ticket_id=id))

@bp.route('/tickets/<ticket_id>/update-details', methods=['POST'])
@login_required
@admin_required
def update_ticket_details(ticket_id):
    """Aktualisiert die Details eines Tickets."""
    try:
        # Verwende die ursprüngliche ID direkt für das Update
        from bson import ObjectId
        try:
            # Versuche zuerst mit ObjectId
            ticket_id_for_update = ObjectId(ticket_id)
        except:
            # Falls das fehlschlägt, verwende die ursprüngliche ID als String
            ticket_id_for_update = ticket_id
        
        # Prüfe ob das Ticket existiert
        ticket = mongodb.find_one('tickets', {'_id': ticket_id_for_update})
        
        if not ticket:
            return jsonify({
                'success': False,
                'message': 'Ticket nicht gefunden'
            }), 404

        # Hole die Daten aus dem Request
        if not request.is_json:
            return jsonify({
                'success': False,
                'message': 'Ungültiges Anfrageformat. JSON erwartet.'
            }), 400

        data = request.get_json()
        
        # Auftragsdetails aktualisieren
        auftrag_details = {
            'ticket_id': ticket_id_for_update,
            'auftrag_an': data.get('auftrag_an', ''),
            'bereich': data.get('bereich', ''),
            'auftraggeber_intern': data.get('auftraggeber_typ') == 'intern',
            'auftraggeber_extern': data.get('auftraggeber_typ') == 'extern',
            'beschreibung': data.get('beschreibung', ''),
            'prioritaet': data.get('prioritaet', 'normal'),
            'deadline': data.get('deadline'),
            'updated_at': datetime.now()
        }
        
        # Verwende die ursprüngliche ID direkt für das Update
        from bson import ObjectId
        try:
            # Versuche zuerst mit ObjectId
            ticket_id_for_update = ObjectId(ticket_id)
        except:
            # Falls das fehlschlägt, verwende die ursprüngliche ID als String
            ticket_id_for_update = ticket_id
        
        if not mongodb.update_one('auftrag_details', {'ticket_id': ticket_id_for_update}, {'$set': auftrag_details}):
            mongodb.insert_one('auftrag_details', auftrag_details)
        
        # Materialliste aktualisieren
        material_list = data.get('material_list', [])
        if material_list:
            # Lösche alte Materialeinträge
            mongodb.delete_many('auftrag_material', {'ticket_id': ticket_id_for_update})
            
            # Füge neue Materialeinträge hinzu
            for material in material_list:
                material['ticket_id'] = ticket_id_for_update
                material['created_at'] = datetime.now()
                mongodb.insert_one('auftrag_material', material)
        
        # Ticket selbst aktualisieren
        ticket_update = {
            'title': data.get('title', ticket.get('title', '')),
            'description': data.get('description', ticket.get('description', '')),
            'priority': data.get('prioritaet', ticket.get('priority', 'normal')),
            'updated_at': datetime.now()
        }
        
        # Verarbeite estimated_time (wird in Minuten gespeichert)
        if 'estimated_time' in data:
            estimated_time = data['estimated_time']
            if estimated_time is not None and estimated_time != '':
                ticket_update['estimated_time'] = float(estimated_time)
            else:
                ticket_update['estimated_time'] = None
        
        # Verarbeite category
        if 'category' in data:
            ticket_update['category'] = data['category']
        
        # Verarbeite due_date
        if 'due_date' in data:
            due_date = data['due_date']
            if due_date:
                try:
                    # Versuche verschiedene Datumsformate zu parsen
                    if 'T' in due_date:
                        due_date = datetime.strptime(due_date, '%Y-%m-%dT%H:%M')
                    else:
                        due_date = datetime.strptime(due_date, '%Y-%m-%d')
                    ticket_update['due_date'] = due_date
                except ValueError:
                    return jsonify({'success': False, 'message': 'Ungültiges Datumsformat'}), 400
            else:
                ticket_update['due_date'] = None
        
        if not mongodb.update_one('tickets', {'_id': ticket_id_for_update}, {'$set': ticket_update}):
            return jsonify({
                'success': False,
                'message': 'Fehler beim Aktualisieren des Tickets'
            }), 500
        
        return jsonify({
            'success': True,
            'message': 'Ticket-Details erfolgreich aktualisiert'
        })
        
    except Exception as e:
        logger.error(f"Fehler beim Aktualisieren der Ticket-Details: {str(e)}", exc_info=True)
        return jsonify({
            'success': False,
            'message': f'Interner Fehler: {str(e)}'
        }), 500

def create_mongodb_backup():
    """Erstellt ein MongoDB-Backup"""
    try:
        from app.utils.backup_manager import BackupManager
        
        backup_manager = BackupManager()
        backup_filename = backup_manager.create_backup()
        
        if backup_filename:
            logger.info(f"MongoDB-Backup erfolgreich erstellt: {backup_filename}")
            return True, backup_filename
        else:
            logger.error("Fehler beim Erstellen des MongoDB-Backups")
            return False, "Fehler beim Erstellen des Backups"
            
    except Exception as e:
        logger.error(f"Fehler beim Erstellen des MongoDB-Backups: {str(e)}")
        return False, f"Fehler beim Erstellen des Backups: {str(e)}"



@bp.route('/manage_users')
@mitarbeiter_required
def manage_users():
    """Benutzerverwaltung"""
    try:
        users = AdminUserService.get_all_users()
        
        # Prüfe auf abgelaufene Konten
        today = datetime.now().date()
        expired_users = []
        
        for user in users:
            if user.get('expiry_date'):
                # Verarbeite verschiedene Datumsformate
                expiry_date = user['expiry_date']
                if isinstance(expiry_date, str):
                    try:
                        expiry_date = datetime.strptime(expiry_date, '%Y-%m-%d').date()
                    except ValueError:
                        continue
                elif isinstance(expiry_date, datetime):
                    expiry_date = expiry_date.date()
                
                # Markiere als abgelaufen wenn das Datum überschritten ist
                if expiry_date < today:
                    user['is_expired'] = True
                    expired_users.append(user)
                else:
                    user['is_expired'] = False
            else:
                user['is_expired'] = False
        
        # Warnung anzeigen wenn abgelaufene Konten existieren
        if expired_users:
            expired_count = len(expired_users)
            if expired_count == 1:
                flash(f'Warnung: 1 Benutzerkonto ist abgelaufen und sollte überprüft werden!', 'warning')
            else:
                flash(f'Warnung: {expired_count} Benutzerkonten sind abgelaufen und sollten überprüft werden!', 'warning')
        
        return render_template('admin/users.html', users=users, expired_users=expired_users)
    except Exception as e:
        logger.error(f"Fehler beim Laden der Benutzer: {e}")
        flash('Fehler beim Laden der Benutzer', 'error')
        return render_template('admin/users.html', users=[], expired_users=[])

@bp.route('/add_user', methods=['GET', 'POST'])
@mitarbeiter_required
def add_user():
    """Neuen Benutzer hinzufügen"""
    if request.method == 'POST':
        # Prüfe Berechtigung für Admin-Rolle
        role = request.form.get('role', '').strip()
        if current_user.role != 'admin' and role == 'admin':
            flash('Sie dürfen keine Admin-Benutzer anlegen.', 'error')
            return render_template('admin/user_form.html', roles=['admin', 'mitarbeiter', 'anwender', 'teilnehmer'], form_data=request.form)
        
        # Verwende Services für Validierung und E-Mail
        from app.services.validation_service import ValidationService
        from app.services.email_service import EmailService
        from app.services.utility_service import UtilityService
        
        form_data = UtilityService.get_form_data_dict(request.form)
        
        # Validierung mit ValidationService
        is_valid, errors, processed_data = ValidationService.validate_user_form(form_data, is_edit=False)
        
        if not is_valid:
            for error in errors:
                flash(error, 'error')
            return render_template('admin/user_form.html', 
                                 roles=['admin', 'mitarbeiter', 'anwender', 'teilnehmer'],
                                 form_data=request.form)
        
        # Automatische Passwort-Generierung wenn keines eingegeben wurde
        if not processed_data['password']:
            password = ValidationService.generate_secure_password()
            processed_data['password'] = password
            processed_data['password_confirm'] = password
        
        # Handlungsfelder aus Formular holen
        handlungsfelder = request.form.getlist('handlungsfelder')
        
        # Ablaufdatum verarbeiten
        expiry_date = None
        expiry_date_str = request.form.get('expiry_date', '').strip()
        if expiry_date_str:
            try:
                expiry_date = datetime.strptime(expiry_date_str, '%Y-%m-%d')
            except ValueError:
                flash('Ungültiges Ablaufdatum-Format', 'error')
                expiry_date = None
        
        # Benutzer erstellen mit AdminUserService
        user_data = {
            'username': processed_data['username'],
            'role': processed_data['role'],
            'email': processed_data['email'] if processed_data['email'] else '',
            'firstname': processed_data['firstname'],
            'lastname': processed_data['lastname'],
            'timesheet_enabled': processed_data['timesheet_enabled'],
            'canteen_plan_enabled': request.form.get('canteen_plan_enabled') == 'on',
            'is_active': request.form.get('is_active') == 'on',
            'handlungsfelder': handlungsfelder,
            'expiry_date': expiry_date
        }
        
        # Passwort nur hinzufügen falls angegeben
        if processed_data.get('password'):
            user_data['password'] = processed_data['password']
        
        success, message, user_id = AdminUserService.create_user(user_data)
        
        if success:
            # E-Mail mit Passwort versenden (falls E-Mail vorhanden)
            if processed_data['email']:
                try:
                    # Passwort aus der Antwort extrahieren (falls generiert)
                    generated_password = None
                    if 'generiert' in message.lower():
                        # Versuche das generierte Passwort aus der Nachricht zu extrahieren
                        import re
                        match = re.search(r'Passwort: ([a-zA-Z0-9]{12})', message)
                        if match:
                            generated_password = match.group(1)
                    
                    email_sent = EmailService.send_new_user_email(
                        processed_data['email'],
                        processed_data['username'],
                        generated_password or processed_data.get('password', ''),
                        processed_data['firstname']
                    )
                    
                    if email_sent:
                        flash(f'{message} Passwort wurde per E-Mail an {processed_data["email"]} gesendet.', 'success')
                    else:
                        flash(f'{message} E-Mail konnte nicht versendet werden.', 'warning')
                except Exception as e:
                    logger.error(f"Fehler beim Versenden der E-Mail: {e}")
                    flash(f'{message} E-Mail konnte nicht versendet werden.', 'warning')
            else:
                flash(f'{message} Das Passwort wurde generiert. Bitte notieren Sie es sicher oder verwenden Sie die E-Mail-Funktion für zukünftige Benutzer.', 'success')
            
            return redirect(url_for('admin.manage_users'))
        else:
            flash(message, 'error')
            return render_template('admin/user_form.html', 
                                 roles=['admin', 'mitarbeiter', 'anwender', 'teilnehmer'],
                                 form_data=request.form)
    
    # Hole alle verfügbaren Handlungsfelder (Ticket-Kategorien)
    handlungsfelder = get_ticket_categories_from_settings()
    
    # Abteilungen aus Settings
    from app.services.admin_system_settings_service import AdminSystemSettingsService
    departments = AdminSystemSettingsService.get_departments_from_settings()
    return render_template('admin/user_form.html', 
                         roles=['admin', 'mitarbeiter', 'anwender', 'teilnehmer'],
                         handlungsfelder=handlungsfelder,
                         user_handlungsfelder=[],
                         departments=departments,
                         user_allowed_departments=[],
                         user_default_department='')

@bp.route('/migrate_users_to_workers', methods=['POST'])
@admin_required
def migrate_users_to_workers():
    """Migriert bestehende Benutzer zu Mitarbeitern"""
    try:
        from app.services.admin_user_service import AdminUserService
        
        # Hole alle aktiven Benutzer
        users = mongodb.find('users', {'is_active': True})
        users = list(users)
        
        migrated_count = 0
        skipped_count = 0
        error_count = 0
        
        for user in users:
            try:
                # Prüfe ob bereits ein Mitarbeiter-Eintrag existiert
                existing_worker = mongodb.find_one('workers', {
                    'user_id': str(user['_id']),
                    'deleted': {'$ne': True}
                })
                
                if existing_worker:
                    skipped_count += 1
                    continue
                
                # Erstelle Mitarbeiter-Eintrag
                success = AdminUserService._create_worker_from_user(user, str(user['_id']))
                if success:
                    migrated_count += 1
                else:
                    error_count += 1
                    
            except Exception as e:
                logger.error(f"Fehler bei Migration von Benutzer {user.get('username', 'Unknown')}: {e}")
                error_count += 1
        
        message = f"Migration abgeschlossen: {migrated_count} Benutzer migriert, {skipped_count} übersprungen, {error_count} Fehler"
        
        if error_count == 0:
            flash(message, 'success')
        else:
            flash(message, 'warning')
            
        return redirect(url_for('admin.manage_users'))
        
    except Exception as e:
        logger.error(f"Fehler bei der Benutzer-Migration: {e}")
        flash(f'Fehler bei der Migration: {str(e)}', 'error')
        return redirect(url_for('admin.manage_users'))

@bp.route('/edit_user/<user_id>', methods=['GET', 'POST'])
@mitarbeiter_required
def edit_user(user_id):
    """Benutzer bearbeiten"""
    user = AdminUserService.get_user_by_id(user_id)
    
    if not user:
        flash('Benutzer nicht gefunden', 'error')
        return redirect(url_for('admin.manage_users'))
    
    if current_user.role != 'admin' and user.get('role') == 'admin':
        flash('Sie dürfen keine Admin-Benutzer bearbeiten.', 'error')
        return redirect(url_for('admin.manage_users'))
    
    # GET-Request: Zeige das Formular mit den aktuellen Daten
    if request.method == 'GET':
        # Hole alle verfügbaren Handlungsfelder (Ticket-Kategorien)
        handlungsfelder = get_ticket_categories_from_settings()
        user_handlungsfelder = user.get('handlungsfelder', [])
        
        from app.services.admin_system_settings_service import AdminSystemSettingsService
        departments = AdminSystemSettingsService.get_departments_from_settings()
        return render_template('admin/user_form.html', 
                             user=user,
                             roles=['admin', 'mitarbeiter', 'anwender', 'teilnehmer'],
                             handlungsfelder=handlungsfelder,
                             user_handlungsfelder=user_handlungsfelder,
                             departments=departments,
                             user_allowed_departments=user.get('allowed_departments', []),
                             user_default_department=user.get('default_department',''))
    
    # POST-Request: Verarbeite die Formulardaten
    try:
        # Verwende Services für Validierung
        from app.services.validation_service import ValidationService
        from app.services.utility_service import UtilityService
        
        form_data = UtilityService.get_form_data_dict(request.form)
        
        # Validierung mit ValidationService
        is_valid, errors, processed_data = ValidationService.validate_user_form(form_data, is_edit=True)
        
        if not is_valid:
            for error in errors:
                flash(error, 'error')
            # Hole alle verfügbaren Handlungsfelder (Ticket-Kategorien)
            handlungsfelder = get_ticket_categories_from_settings()
            user_handlungsfelder = user.get('handlungsfelder', [])
            
            return render_template('admin/user_form.html', 
                                 user=user,
                                 roles=['admin', 'mitarbeiter', 'anwender', 'teilnehmer'],
                                 handlungsfelder=handlungsfelder,
                                 user_handlungsfelder=user_handlungsfelder)
        
        # Handlungsfelder aus Formular holen
        handlungsfelder = request.form.getlist('handlungsfelder')
        
        # Ablaufdatum verarbeiten
        expiry_date = None
        expiry_date_str = request.form.get('expiry_date', '').strip()
        if expiry_date_str:
            try:
                expiry_date = datetime.strptime(expiry_date_str, '%Y-%m-%d')
            except ValueError:
                flash('Ungültiges Ablaufdatum-Format', 'error')
                expiry_date = None
        
        # Benutzer aktualisieren mit AdminUserService
        user_data = {
            'username': processed_data['username'],
            'role': processed_data['role'],
            'email': processed_data['email'] if processed_data['email'] else '',
            'firstname': processed_data['firstname'],
            'lastname': processed_data['lastname'],
            'timesheet_enabled': processed_data['timesheet_enabled'],
            'canteen_plan_enabled': request.form.get('canteen_plan_enabled') == 'on',
            'is_active': request.form.get('is_active') == 'on',
            'handlungsfelder': handlungsfelder,
            'expiry_date': expiry_date
        }
        
        # Passwort hinzufügen falls angegeben
        if processed_data['password']:
            user_data['password'] = processed_data['password']
        
        # Abteilungsrechte
        allowed_departments = request.form.getlist('allowed_departments')
        default_department = request.form.get('default_department') or None
        user_data['allowed_departments'] = allowed_departments
        user_data['default_department'] = default_department if default_department else None

        success, message = AdminUserService.update_user(user_id, user_data)
        
        if success:
            flash(message, 'success')
            return redirect(url_for('admin.manage_users'))
        else:
            flash(message, 'error')
            # Hole alle verfügbaren Handlungsfelder (Ticket-Kategorien)
            handlungsfelder = get_ticket_categories_from_settings()
            user_handlungsfelder = user.get('handlungsfelder', [])
            
            return render_template('admin/user_form.html', 
                                 user=user,
                                 roles=['admin', 'mitarbeiter', 'anwender', 'teilnehmer'],
                                 handlungsfelder=handlungsfelder,
                                 user_handlungsfelder=user_handlungsfelder)
        
    except Exception as e:
        logger.error(f"Fehler beim Aktualisieren des Benutzers: {e}")
        flash('Fehler beim Aktualisieren des Benutzers', 'error')
        # Hole alle verfügbaren Handlungsfelder (Ticket-Kategorien)
        handlungsfelder = get_ticket_categories_from_settings()
        user_handlungsfelder = user.get('handlungsfelder', [])
        
        return render_template('admin/user_form.html', 
                             user=user,
                             roles=['admin', 'mitarbeiter', 'anwender', 'teilnehmer'],
                             handlungsfelder=handlungsfelder,
                             user_handlungsfelder=user_handlungsfelder)
    
    # Hole alle verfügbaren Handlungsfelder (Ticket-Kategorien)
    handlungsfelder = get_ticket_categories_from_settings()
    user_handlungsfelder = user.get('handlungsfelder', [])
    
    return render_template('admin/user_form.html', 
                         user=user,
                         roles=['admin', 'mitarbeiter', 'anwender', 'teilnehmer'],
                         handlungsfelder=handlungsfelder,
                         user_handlungsfelder=user_handlungsfelder)

@bp.route('/notices')
@admin_required
def notices():
    """Notizen-Übersicht"""
    notices = AdminNotificationService.get_all_notices()
    return render_template('admin/notices.html', notices=notices)

@bp.route('/create_notice', methods=['GET', 'POST'])
@admin_required
def create_notice():
    """Neue Notiz erstellen"""
    if request.method == 'POST':
        title = request.form.get('title')
        content = request.form.get('content')
        notice_type = request.form.get('type', 'info')
        
        success, message = AdminNotificationService.create_notice(title, content, notice_type)
        
        if success:
            flash(message, 'success')
        else:
            flash(message, 'error')
        
        return redirect(url_for('admin.notices'))
    
    return render_template('admin/notice_form.html')

@bp.route('/edit_notice/<id>', methods=['GET', 'POST'])
@admin_required
def edit_notice(id):
    """Notiz bearbeiten"""
    if request.method == 'POST':
        title = request.form.get('title')
        content = request.form.get('content')
        notice_type = request.form.get('type', 'info')
        
        success, message = AdminNotificationService.update_notice(id, title, content, notice_type)
        
        if success:
            flash(message, 'success')
        else:
            flash(message, 'error')
        
        return redirect(url_for('admin.notices'))
    
    notice = AdminNotificationService.get_notice_by_id(id)
    if not notice:
        flash('Notiz nicht gefunden', 'error')
        return redirect(url_for('admin.notices'))
    
    return render_template('admin/notice_form.html', notice=notice)

@bp.route('/delete_notice/<id>', methods=['POST'])
@admin_required
def delete_notice(id):
    """Löscht einen Hinweis"""
    try:
        success, message = AdminNotificationService.delete_notice(id)
        
        if success:
            flash(message, 'success')
        else:
            flash(message, 'error')
            
        return redirect(url_for('admin.notices'))
        
    except Exception as e:
        logger.error(f"Fehler beim Löschen des Hinweises: {str(e)}")
        flash('Fehler beim Löschen des Hinweises', 'error')
        return redirect(url_for('admin.notices'))

@bp.route('/upload_logo', methods=['POST'])
@admin_required
def upload_logo():
    """Logo hochladen"""
    flash('Logo-Upload-Funktion noch nicht implementiert', 'warning')
    return redirect(url_for('admin.system'))

@bp.route('/delete-logo/<filename>', methods=['POST'])
@admin_required
def delete_logo(filename):
    """Logo löschen"""
    try:
        import os
        logo_path = os.path.join(current_app.root_path, 'static', 'uploads', 'logos', filename)
        if os.path.exists(logo_path):
            os.remove(logo_path)
            return jsonify({'success': True, 'message': 'Logo erfolgreich gelöscht'})
        else:
            return jsonify({'success': False, 'message': 'Logo nicht gefunden'}), 404
    except Exception as e:
        logger.error(f"Fehler beim Löschen des Logos: {e}")
        return jsonify({'success': False, 'message': 'Fehler beim Löschen des Logos'}), 500

@bp.route('/add_ticket_category', methods=['POST'])
@admin_required
def add_ticket_category():
    """Fügt eine neue Ticket-Kategorie hinzu"""
    try:
        name = request.form.get('category')
        if not name:
            flash('Bitte geben Sie einen Namen ein.', 'error')
            return redirect(url_for('tickets.create'))

        # Prüfen ob Ticket-Kategorie bereits existiert
        settings = mongodb.db.settings.find_one({'key': 'ticket_categories'})
        if settings and name in settings.get('value', []):
            flash('Diese Ticket-Kategorie existiert bereits.', 'error')
            return redirect(url_for('tickets.create'))

        # Ticket-Kategorie zur Liste hinzufügen
        if settings:
            mongodb.update_one_array(
                'settings',
                {'key': 'ticket_categories'},
                {'$push': {'value': name}}
            )
        else:
            mongodb.insert_one('settings', {
                'key': 'ticket_categories',
                'value': [name]
            })

        flash('Ticket-Kategorie erfolgreich hinzugefügt.', 'success')
        return redirect(url_for('tickets.create'))
    except Exception as e:
        logger.error(f"Fehler beim Hinzufügen der Ticket-Kategorie: {str(e)}")
        flash('Ein Fehler ist aufgetreten.', 'error')
        return redirect(url_for('tickets.create'))

@bp.route('/delete_ticket_category/<category>', methods=['POST'])
@admin_required
def delete_ticket_category(category):
    """Löscht eine Ticket-Kategorie"""
    try:
        # Überprüfen, ob die Ticket-Kategorie in Verwendung ist
        tickets_with_category = mongodb.db.tickets.find_one({'category': category})
        if tickets_with_category:
            flash('Die Ticket-Kategorie kann nicht gelöscht werden, da sie noch von Tickets verwendet wird.', 'error')
            return redirect(url_for('tickets.create'))

        # Ticket-Kategorie aus der Liste entfernen
        mongodb.update_one_array(
            'settings',
            {'key': 'ticket_categories'},
            {'$pull': {'value': category}}
        )
        
        flash('Ticket-Kategorie erfolgreich gelöscht.', 'success')
        return redirect(url_for('tickets.create'))
    except Exception as e:
        logger.error(f"Fehler beim Löschen der Ticket-Kategorie: {str(e)}")
        flash('Ein Fehler ist aufgetreten.', 'error')
        return redirect(url_for('tickets.create'))

@bp.route('/system', methods=['GET', 'POST'])
@admin_required
def system():
    """System-Einstellungen"""
    try:
        if request.method == 'POST':
            # Begriffe & Icons verarbeiten
            app_labels = {
                'tools': {
                    'name': request.form.get('label_tools_name', 'Werkzeuge'),
                    'icon': request.form.get('label_tools_icon', 'fas fa-tools')
                },
                'consumables': {
                    'name': request.form.get('label_consumables_name', 'Verbrauchsmaterial'),
                    'icon': request.form.get('label_consumables_icon', 'fas fa-box')
                },
                'tickets': {
                    'name': request.form.get('label_tickets_name', 'Tickets'),
                    'icon': request.form.get('label_tickets_icon', 'fas fa-ticket-alt')
                }
            }
            
            success, message = AdminSystemService.save_app_labels(app_labels)
            
            if success:
                flash(message, 'success')
            else:
                flash(message, 'error')
            
            # Software-Presets verarbeiten
            software_presets_text = request.form.get('software_presets', '')
            if software_presets_text:
                software_presets = [line.strip() for line in software_presets_text.split('\n') if line.strip()]
                mongodb.update_one('settings', 
                                 {'key': 'software_presets'}, 
                                 {'$set': {'value': software_presets}}, 
                                 upsert=True)
            
            # Nutzergruppen verarbeiten
            user_groups_text = request.form.get('user_groups', '')
            if user_groups_text:
                user_groups = [line.strip() for line in user_groups_text.split('\n') if line.strip()]
                mongodb.update_one('settings', 
                                 {'key': 'user_groups'}, 
                                 {'$set': {'value': user_groups}}, 
                                 upsert=True)
            
            return redirect(url_for('admin.system'))
        
        # Hole alle verfügbaren Logos
        logos = AdminSystemService.get_available_logos()
        
        # Hole aktuelle Einstellungen und App-Labels
        settings, app_labels = AdminSystemService.get_system_data()
        
        # Hole Software-Presets und Nutzergruppen
        software_presets_setting = mongodb.find_one('settings', {'key': 'software_presets'})
        software_presets = '\n'.join(software_presets_setting.get('value', [])) if software_presets_setting else ''
        
        user_groups_setting = mongodb.find_one('settings', {'key': 'user_groups'})
        user_groups = '\n'.join(user_groups_setting.get('value', [])) if user_groups_setting else ''
        
        return render_template('admin/server-settings.html', 
                             logos=logos, 
                             settings=settings,
                             app_labels=app_labels,
                             software_presets=software_presets,
                             user_groups=user_groups)
    except Exception as e:
        logger.error(f"Fehler beim Laden der Systemeinstellungen: {str(e)}")
        flash('Fehler beim Laden der Systemeinstellungen', 'error')
        return redirect(url_for('admin.index'))

# Software-Verwaltung
@bp.route('/software')
@mitarbeiter_required
def get_software():
    """Gibt alle Software-Pakete zurück"""
    try:
        software_list = mongodb.find('software', {}, sort=[('name', 1)])
        return jsonify({
            'success': True,
            'software': list(software_list)
        })
    except Exception as e:
        logger.error(f"Fehler beim Abrufen der Software: {str(e)}")
        return jsonify({
            'success': False,
            'message': 'Fehler beim Laden der Software'
        })

@bp.route('/software/add', methods=['POST'])
@mitarbeiter_required
def add_software():
    """Fügt ein neues Software-Paket hinzu"""
    try:
        name = request.form.get('name', '').strip()
        category = request.form.get('category', '').strip()
        description = request.form.get('description', '').strip()
        
        if not name:
            return jsonify({
                'success': False,
                'message': 'Bitte geben Sie einen Namen ein.'
            })

        # Prüfe ob Software bereits existiert
        existing = mongodb.find_one('software', {'name': name})
        if existing:
            return jsonify({
                'success': False,
                'message': 'Diese Software existiert bereits.'
            })

        software_data = {
            'name': name,
            'category': category,
            'description': description,
            'created_at': datetime.now(),
            'updated_at': datetime.now()
        }
        
        mongodb.insert_one('software', software_data)
        
        return jsonify({
            'success': True,
            'message': f'Software "{name}" erfolgreich hinzugefügt'
        })
        
    except Exception as e:
        logger.error(f"Fehler beim Hinzufügen der Software: {str(e)}")
        return jsonify({
            'success': False,
            'message': 'Ein Fehler ist aufgetreten.'
        })

@bp.route('/software/delete/<software_id>', methods=['POST'])
@mitarbeiter_required
def delete_software(software_id):
    """Löscht ein Software-Paket"""
    try:
        # Prüfe ob Software in Nutzergruppen verwendet wird
        groups_using_software = mongodb.find('user_groups', {'software': {'$in': [software_id]}})
        if list(groups_using_software):
            return jsonify({
                'success': False,
                'message': 'Diese Software wird noch in Nutzergruppen verwendet und kann nicht gelöscht werden.'
            })

        result = mongodb.delete_one('software', {'_id': ObjectId(software_id)})
        
        if result.deleted_count > 0:
            return jsonify({
                'success': True,
                'message': 'Software erfolgreich gelöscht'
            })
        else:
            return jsonify({
                'success': False,
                'message': 'Software nicht gefunden'
            })
            
    except Exception as e:
        logger.error(f"Fehler beim Löschen der Software: {str(e)}")
        return jsonify({
            'success': False,
            'message': 'Ein Fehler ist aufgetreten.'
        })

# Nutzergruppen-Verwaltung
@bp.route('/user_groups')
@mitarbeiter_required
def get_user_groups_admin():
    """Gibt alle Nutzergruppen zurück"""
    try:
        groups = mongodb.find('user_groups', {}, sort=[('name', 1)])
        return jsonify({
            'success': True,
            'groups': list(groups)
        })
    except Exception as e:
        logger.error(f"Fehler beim Abrufen der Nutzergruppen: {str(e)}")
        return jsonify({
            'success': False,
            'message': 'Fehler beim Laden der Nutzergruppen'
        })

@bp.route('/user_groups/add', methods=['POST'])
@mitarbeiter_required
def add_user_group():
    """Fügt eine neue Nutzergruppe hinzu"""
    try:
        name = request.form.get('name', '').strip()
        description = request.form.get('description', '').strip()
        software_ids = request.form.getlist('software')
        
        if not name:
            return jsonify({
                'success': False,
                'message': 'Bitte geben Sie einen Namen ein.'
            })

        # Prüfe ob Gruppe bereits existiert
        existing = mongodb.find_one('user_groups', {'name': name})
        if existing:
            return jsonify({
                'success': False,
                'message': 'Diese Nutzergruppe existiert bereits.'
            })

        group_data = {
            'name': name,
            'description': description,
            'software': software_ids,
            'created_at': datetime.now(),
            'updated_at': datetime.now()
        }
        
        mongodb.insert_one('user_groups', group_data)
        
        return jsonify({
            'success': True,
            'message': f'Nutzergruppe "{name}" erfolgreich hinzugefügt'
        })
        
    except Exception as e:
        logger.error(f"Fehler beim Hinzufügen der Nutzergruppe: {str(e)}")
        return jsonify({
            'success': False,
            'message': 'Ein Fehler ist aufgetreten.'
        })

@bp.route('/user_groups/delete/<group_id>', methods=['POST'])
@mitarbeiter_required
def delete_user_group(group_id):
    """Löscht eine Nutzergruppe"""
    try:
        # Prüfe ob Gruppe in Werkzeugen verwendet wird
        tools_using_group = mongodb.find('tools', {'user_groups': {'$in': [group_id]}})
        if list(tools_using_group):
            return jsonify({
                'success': False,
                'message': 'Diese Nutzergruppe wird noch in Werkzeugen verwendet und kann nicht gelöscht werden.'
            })

        result = mongodb.delete_one('user_groups', {'_id': ObjectId(group_id)})
        
        if result.deleted_count > 0:
            return jsonify({
                'success': True,
                'message': 'Nutzergruppe erfolgreich gelöscht'
            })
        else:
            return jsonify({
                'success': False,
                'message': 'Nutzergruppe nicht gefunden'
            })
            
    except Exception as e:
        logger.error(f"Fehler beim Löschen der Nutzergruppe: {str(e)}")
        return jsonify({
            'success': False,
            'message': 'Ein Fehler ist aufgetreten.'
        })

@bp.route('/user_groups/<group_id>/edit', methods=['POST'])
@mitarbeiter_required
def edit_user_group(group_id):
    """Bearbeitet eine Nutzergruppe"""
    try:
        name = request.form.get('name', '').strip()
        description = request.form.get('description', '').strip()
        software_ids = request.form.getlist('software')
        
        if not name:
            return jsonify({
                'success': False,
                'message': 'Bitte geben Sie einen Namen ein.'
            })

        # Prüfe ob Name bereits existiert (außer bei der aktuellen Gruppe)
        existing = mongodb.find_one('user_groups', {'name': name, '_id': {'$ne': ObjectId(group_id)}})
        if existing:
            return jsonify({
                'success': False,
                'message': 'Diese Nutzergruppe existiert bereits.'
            })

        update_data = {
            'name': name,
            'description': description,
            'software': software_ids,
            'updated_at': datetime.now()
        }
        
        success = mongodb.update_one('user_groups', {'_id': ObjectId(group_id)}, {'$set': update_data})
        
        if success:
            return jsonify({
                'success': True,
                'message': f'Nutzergruppe "{name}" erfolgreich aktualisiert'
            })
        else:
            return jsonify({
                'success': False,
                'message': 'Nutzergruppe nicht gefunden oder keine Änderungen'
            })
            
    except Exception as e:
        logger.error(f"Fehler beim Bearbeiten der Nutzergruppe: {str(e)}")
        return jsonify({
            'success': False,
            'message': 'Ein Fehler ist aufgetreten.'
        })

@bp.route('/software_management')
@mitarbeiter_required
def software_management():
    """Software- und Nutzergruppen-Verwaltung"""
    try:
        # Prüfe ob Software-Management aktiviert ist
        if not is_feature_enabled('software_management'):
            flash('Software-Management ist deaktiviert', 'error')
            return redirect(url_for('admin.dashboard'))
        
        # Hole alle Software-Pakete
        software_list = list(mongodb.find('software', {}, sort=[('name', 1)]))
        
        # Hole alle Nutzergruppen
        groups_list = list(mongodb.find('user_groups', {}, sort=[('name', 1)]))
        
        return render_template('admin/software_management.html',
                             software_list=software_list,
                             groups_list=groups_list)
                             
    except Exception as e:
        logger.error(f"Fehler beim Laden der Software-Verwaltung: {str(e)}")
        flash('Fehler beim Laden der Software-Verwaltung', 'error')
        return redirect(url_for('admin.dashboard'))

@bp.route('/feature_settings', methods=['GET', 'POST'])
@admin_required
def feature_settings():
    """Feature-Einstellungen verwalten"""
    try:
        if request.method == 'POST':
            # Feature-Einstellungen verarbeiten
            features = {
                'tools': request.form.get('feature_tools') == 'on',
                'consumables': request.form.get('feature_consumables') == 'on',
                'workers': request.form.get('feature_workers') == 'on',
                'lending_system': request.form.get('feature_lending_system') == 'on',
                'ticket_system': request.form.get('feature_ticket_system') == 'on',
                'timesheet': request.form.get('feature_timesheet') == 'on',
                'media_management': request.form.get('feature_media_management') == 'on',
                'software_management': request.form.get('feature_software_management') == 'on',
                'job_board': request.form.get('feature_job_board') == 'on',
                'weekly_reports': request.form.get('feature_weekly_reports') == 'on',
                'canteen_plan': request.form.get('feature_canteen_plan') == 'on',
                # Werkzeug-Felder
                'tool_field_serial_number': request.form.get('tool_field_serial_number') == 'on',
                'tool_field_invoice_number': request.form.get('tool_field_invoice_number') == 'on',
                'tool_field_mac_address': request.form.get('tool_field_mac_address') == 'on',
                'tool_field_mac_address_wlan': request.form.get('tool_field_mac_address_wlan') == 'on',
                'tool_field_user_groups': request.form.get('tool_field_user_groups') == 'on',
                'tool_field_software': request.form.get('tool_field_software') == 'on'
            }
            
            # Einstellungen speichern
            for feature_name, enabled in features.items():
                set_feature_setting(feature_name, enabled)
            
            flash('Feature-Einstellungen erfolgreich gespeichert', 'success')
            return redirect(url_for('admin.feature_settings'))
        
        # Aktuelle Feature-Einstellungen laden
        feature_settings = get_feature_settings()
        
        return render_template('admin/feature_settings.html',
                             feature_settings=feature_settings)
                             
    except Exception as e:
        logger.error(f"Fehler beim Laden der Feature-Einstellungen: {str(e)}")
        flash('Fehler beim Laden der Feature-Einstellungen', 'error')
        return redirect(url_for('admin.dashboard'))


@bp.route('/role_permissions', methods=['GET', 'POST'])
@admin_required
def role_permissions():
    """Rollen- und Berechtigungs-Matrix verwalten."""
    try:
        if request.method == 'POST':
            # Erwartet JSON im Formularfeld 'permissions' oder einzelne Checkboxen
            if request.is_json:
                payload = request.get_json()
                permissions = payload.get('permissions', {})
            else:
                # Aus HTML-Form-Checkboxen zusammensetzen: name="perm[role][area][action]"
                permissions = get_role_permissions()
                # Flaches Formular in Matrix zurückschreiben
                for key, value in request.form.items():
                    if not key.startswith('perm[') or value != 'on':
                        continue
                    try:
                        # perm[role][area][action]
                        parts = key.split('[')
                        role = parts[1][:-1]
                        area = parts[2][:-1]
                        action = parts[3][:-1]
                        permissions.setdefault(role, {}).setdefault(area, [])
                        if action not in permissions[role][area]:
                            permissions[role][area].append(action)
                    except Exception:
                        continue

            # Guardrail: Admin immer alles
            from app.utils.permissions import DEFAULT_ROLE_PERMISSIONS
            permissions['admin'] = DEFAULT_ROLE_PERMISSIONS['admin']

            if set_role_permissions(permissions):
                flash('Berechtigungen erfolgreich gespeichert', 'success')
            else:
                flash('Fehler beim Speichern der Berechtigungen', 'error')
            return redirect(url_for('admin.role_permissions'))

        # GET: aktuelle Matrix anzeigen
        permissions = get_role_permissions()
        areas = sorted({a for r in permissions.values() for a in r.keys()})
        actions = sorted({act for r in permissions.values() for acts in r.values() for act in acts})
        roles = sorted(permissions.keys())
        return render_template('admin/role_permissions.html',
                               permissions=permissions,
                               roles=roles,
                               areas=areas,
                               actions=actions)
    except Exception as e:
        logger.error(f"Fehler beim Laden der Rollenrechte: {str(e)}")
        flash('Fehler beim Laden der Rollenrechte', 'error')
        return redirect(url_for('admin.dashboard'))

# =============================================================================
# FELDVERWALTUNG - Standard- und benutzerdefinierte Felder
# =============================================================================
@bp.route('/custom_fields', methods=['GET', 'POST'])
@admin_required
def custom_fields():
    """
    Zentrale Feld-Verwaltung für Standard- und benutzerdefinierte Felder
    
    GET: Zeigt die Feldverwaltung-Seite mit Standard-Feldern und benutzerdefinierten Feldern
    POST: Verarbeitet Updates für Standard-Felder (action=update_default_fields)
    
    Integriert:
    - Standard-Werkzeugfelder (Seriennummer, MAC-Adressen, etc.)
    - Benutzerdefinierte Felder für Werkzeuge und Verbrauchsgüter
    """
    try:
        if request.method == 'POST':
            action = request.form.get('action')
            
            if action == 'update_default_fields':
                # Standard-Felder aktualisieren
                field_mappings = {
                    'tool_field_serial_number': 'tool_field_serial_number',
                    'tool_field_invoice_number': 'tool_field_invoice_number', 
                    'tool_field_mac_address': 'tool_field_mac_address',
                    'tool_field_mac_address_wlan': 'tool_field_mac_address_wlan',
                    'tool_field_user_groups': 'tool_field_user_groups',
                    'tool_field_software': 'tool_field_software'
                }
                
                for form_field, setting_key in field_mappings.items():
                    set_feature_setting(setting_key, form_field in request.form)
                
                flash('Standard-Felder erfolgreich aktualisiert', 'success')
                return redirect(url_for('admin.custom_fields'))
        
        # Alle benutzerdefinierten Felder laden
        custom_fields = CustomFieldsService.get_all_custom_fields()
        
        # Feature-Einstellungen laden
        feature_settings = get_feature_settings()
        
        return render_template('admin/custom_fields.html',
                             custom_fields=custom_fields,
                             field_types=CustomFieldsService.FIELD_TYPES,
                             target_types=CustomFieldsService.TARGET_TYPES,
                             feature_settings=feature_settings)
                             
    except Exception as e:
        logger.error(f"Fehler beim Laden der Feld-Verwaltung: {str(e)}")
        flash('Fehler beim Laden der Feld-Verwaltung', 'error')
        return redirect(url_for('admin.dashboard'))

@bp.route('/custom_fields/add', methods=['POST'])
@admin_required
def add_custom_field():
    """Neues benutzerdefiniertes Feld hinzufügen"""
    try:
        
        # Daten aus dem Formular extrahieren
        field_data = {
            'name': request.form.get('name', ''),
            'field_type': request.form.get('field_type', ''),
            'target_type': request.form.get('target_type', ''),
            'description': request.form.get('description', ''),
            'required': request.form.get('required') == 'on',
            'default_value': request.form.get('default_value', ''),
            'sort_order': int(request.form.get('sort_order', 999)),
            'select_options': request.form.get('select_options', '')
        }
        
        success, message = CustomFieldsService.create_custom_field(field_data)
        
        if success:
            flash(message, 'success')
        else:
            flash(message, 'error')
            
        return redirect(url_for('admin.custom_fields'))
        
    except Exception as e:
        logger.error(f"Fehler beim Erstellen des benutzerdefinierten Feldes: {str(e)}")
        flash('Fehler beim Erstellen des Feldes', 'error')
        return redirect(url_for('admin.custom_fields'))

@bp.route('/custom_fields/<field_id>/edit', methods=['POST'])
@admin_required  
def edit_custom_field(field_id):
    """Benutzerdefiniertes Feld bearbeiten"""
    try:
        # Daten aus dem Formular extrahieren
        field_data = {
            'name': request.form.get('name', ''),
            'field_type': request.form.get('field_type', ''),
            'target_type': request.form.get('target_type', ''),
            'description': request.form.get('description', ''),
            'required': request.form.get('required') == 'on',
            'default_value': request.form.get('default_value', ''),
            'sort_order': int(request.form.get('sort_order', 999)),
            'select_options': request.form.get('select_options', '')
        }
        
        success, message = CustomFieldsService.update_custom_field(field_id, field_data)
        
        if success:
            flash(message, 'success')
        else:
            flash(message, 'error')
            
        return redirect(url_for('admin.custom_fields'))
        
    except Exception as e:
        logger.error(f"Fehler beim Bearbeiten des benutzerdefinierten Feldes: {str(e)}")
        flash('Fehler beim Bearbeiten des Feldes', 'error')
        return redirect(url_for('admin.custom_fields'))

@bp.route('/custom_fields/<field_id>/delete', methods=['POST'])
@admin_required
def delete_custom_field(field_id):
    """Benutzerdefiniertes Feld löschen"""
    try:
        success, message = CustomFieldsService.delete_custom_field(field_id)
        
        if success:
            flash(message, 'success')
        else:
            flash(message, 'error')
            
        return redirect(url_for('admin.custom_fields'))
        
    except Exception as e:
        logger.error(f"Fehler beim Löschen des benutzerdefinierten Feldes: {str(e)}")
        flash('Fehler beim Löschen des Feldes', 'error')
        return redirect(url_for('admin.custom_fields'))

# Abteilungsverwaltung
@bp.route('/departments')
@mitarbeiter_required
def get_departments():
    """Gibt alle Abteilungen zurück"""
    try:
        # Verwende den AdminSystemSettingsService
        departments = AdminSystemSettingsService.get_departments_from_settings()
        return jsonify({
            'success': True,
            'departments': [{'name': dept} for dept in departments]
        })
    except Exception as e:
        logger.error(f"Fehler beim Abrufen der Abteilungen: {str(e)}")
        return jsonify({
            'success': False,
            'message': 'Fehler beim Laden der Abteilungen'
        })

@bp.route('/departments/add', methods=['POST'])
@mitarbeiter_required
def add_department():
    """Fügt eine neue Abteilung hinzu"""
    try:
        # Unterstütze beide Feldnamen für Kompatibilität
        name = request.form.get('name', '').strip() or request.form.get('department', '').strip()
        if not name:
            return jsonify({
                'success': False,
                'message': 'Bitte geben Sie einen Namen ein.'
            })

        # Verwende den AdminSystemSettingsService
        success, message = AdminSystemSettingsService.add_department(name)
        
        if success:
            return jsonify({
                'success': True,
                'message': message
            })
        else:
            return jsonify({
                'success': False,
                'message': message
        })
    except Exception as e:
        logger.error(f"Fehler beim Hinzufügen der Abteilung: {str(e)}")
        return jsonify({
            'success': False,
            'message': 'Ein Fehler ist aufgetreten.'
        })

@bp.route('/departments/delete/<name>', methods=['POST'])
@mitarbeiter_required
def delete_department(name):
    try:
        # URL-dekodieren
        from urllib.parse import unquote
        decoded_name = unquote(name)
        
        # Verwende den AdminSystemSettingsService
        success, message = AdminSystemSettingsService.delete_department(decoded_name)
        
        if success:
            return jsonify({
                'success': True,
                'message': message
            })
        else:
            return jsonify({
                'success': False,
                'message': message
        })
    except Exception as e:
        logger.error(f"Fehler beim Löschen der Abteilung: {str(e)}")
        return jsonify({
            'success': False,
            'message': 'Ein Fehler ist aufgetreten.'
        })

# Kategorienverwaltung
@bp.route('/categories')
@mitarbeiter_required
def get_categories_admin():
    """Gibt alle Kategorien zurück"""
    try:
        # Verwende den AdminSystemSettingsService
        categories = AdminSystemSettingsService.get_categories_from_settings()
        return jsonify({
            'success': True,
            'categories': [{'name': cat} for cat in categories]
        })
    except Exception as e:
        logger.error(f"Fehler beim Abrufen der Kategorien: {str(e)}")
        return jsonify({
            'success': False,
            'message': 'Fehler beim Laden der Kategorien'
        })

@bp.route('/categories/add', methods=['POST'])
@mitarbeiter_required
def add_category():
    """Fügt eine neue Kategorie hinzu"""
    try:
        # Unterstütze beide Feldnamen für Kompatibilität
        name = request.form.get('name', '').strip() or request.form.get('category', '').strip()
        if not name:
            return jsonify({
                'success': False,
                'message': 'Bitte geben Sie einen Namen ein.'
            })

        # Lade aktuelle Kategorien
        settings = mongodb.db.settings.find_one({'key': 'categories'})
        current_categories = []
        if settings and 'value' in settings:
            value = settings['value']
            if isinstance(value, str):
                current_categories = [cat.strip() for cat in value.split(',') if cat.strip()]
            elif isinstance(value, list):
                current_categories = value
        
        # Prüfe auf Duplikate (case-insensitive)
        if any(cat.lower() == name.lower() for cat in current_categories):
            return jsonify({
                'success': False,
                'message': 'Diese Kategorie existiert bereits.'
            })

        # Kategorie zur Liste hinzufügen
        current_categories.append(name)
        
        # Speichere als Array mit robuster Methode
        try:
            mongodb.update_one(
                'settings',
                {'key': 'categories'},
                {'value': current_categories},
                upsert=True
            )
        except Exception as db_error:
            logger.error(f"Datenbankfehler beim Speichern der Kategorie: {str(db_error)}")
            # Fallback: Versuche es mit update_one_array
            mongodb.update_one_array(
                'settings',
                {'key': 'categories'},
                {'$set': {'value': current_categories}},
                upsert=True
            )

        return jsonify({
            'success': True,
            'message': 'Kategorie erfolgreich hinzugefügt.'
        })
    except Exception as e:
        logger.error(f"Fehler beim Hinzufügen der Kategorie: {str(e)}")
        return jsonify({
            'success': False,
            'message': 'Ein Fehler ist aufgetreten.'
        })

@bp.route('/categories/delete/<name>', methods=['POST'])
@mitarbeiter_required
def delete_category(name):
    try:
        # URL-dekodieren
        decoded_name = unquote(name)
        
        # Überprüfen, ob die Kategorie in Verwendung ist
        tools_with_category = mongodb.db.tools.find_one({'category': decoded_name})
        if tools_with_category:
            return jsonify({
                'success': False,
                'message': 'Die Kategorie kann nicht gelöscht werden, da sie noch von Werkzeugen verwendet wird.'
            })

        # Lade aktuelle Kategorien
        settings = mongodb.db.settings.find_one({'key': 'categories'})
        current_categories = []
        if settings and 'value' in settings:
            value = settings['value']
            if isinstance(value, str):
                current_categories = [cat.strip() for cat in value.split(',') if cat.strip()]
            elif isinstance(value, list):
                current_categories = value
        
        # Kategorie aus der Liste entfernen
        if decoded_name in current_categories:
            current_categories.remove(decoded_name)
            try:
                mongodb.update_one(
                    'settings',
                    {'key': 'categories'},
                    {'value': current_categories},
                    upsert=True
                )
            except Exception as db_error:
                logger.error(f"Datenbankfehler beim Löschen der Kategorie: {str(db_error)}")
                # Fallback: Versuche es mit update_one_array
                mongodb.update_one_array(
                    'settings',
                    {'key': 'categories'},
                    {'$set': {'value': current_categories}},
                    upsert=True
                )
        
        return jsonify({
            'success': True,
            'message': 'Kategorie erfolgreich gelöscht.'
        })
    except Exception as e:
        logger.error(f"Fehler beim Löschen der Kategorie: {str(e)}")
        return jsonify({
            'success': False,
            'message': 'Ein Fehler ist aufgetreten.'
        })

# Standortverwaltung
@bp.route('/locations')
@mitarbeiter_required
def get_locations():
    """Gibt alle Standorte zurück"""
    try:
        # Verwende den AdminSystemSettingsService
        locations = AdminSystemSettingsService.get_locations_from_settings()
        return jsonify({
            'success': True,
            'locations': [{'name': loc} for loc in locations]
        })
    except Exception as e:
        logger.error(f"Fehler beim Abrufen der Standorte: {str(e)}")
        return jsonify({
            'success': False,
            'message': 'Fehler beim Laden der Standorte'
        })

@bp.route('/locations/add', methods=['POST'])
@mitarbeiter_required
def add_location():
    """Fügt einen neuen Standort hinzu"""
    try:
        # Unterstütze beide Feldnamen für Kompatibilität
        name = request.form.get('name', '').strip() or request.form.get('location', '').strip()
        if not name:
            return jsonify({
                'success': False,
                'message': 'Bitte geben Sie einen Namen ein.'
            })

        # Lade aktuelle Standorte
        settings = mongodb.db.settings.find_one({'key': 'locations'})
        current_locations = []
        if settings and 'value' in settings:
            value = settings['value']
            if isinstance(value, str):
                current_locations = [loc.strip() for loc in value.split(',') if loc.strip()]
            elif isinstance(value, list):
                current_locations = value
        
        # Prüfe auf Duplikate (case-insensitive)
        if any(loc.lower() == name.lower() for loc in current_locations):
            return jsonify({
                'success': False,
                'message': 'Dieser Standort existiert bereits.'
            })

        # Standort zur Liste hinzufügen
        current_locations.append(name)
        
        # Speichere als Array mit robuster Methode
        try:
            mongodb.update_one(
                'settings',
                {'key': 'locations'},
                {'value': current_locations},
                upsert=True
            )
        except Exception as db_error:
            logger.error(f"Datenbankfehler beim Speichern des Standorts: {str(db_error)}")
            # Fallback: Versuche es mit update_one_array
            mongodb.update_one_array(
                'settings',
                {'key': 'locations'},
                {'$set': {'value': current_locations}},
                upsert=True
            )

        return jsonify({
            'success': True,
            'message': 'Standort erfolgreich hinzugefügt.'
        })
    except Exception as e:
        logger.error(f"Fehler beim Hinzufügen des Standorts: {str(e)}")
        return jsonify({
            'success': False,
            'message': 'Ein Fehler ist aufgetreten.'
        })

@bp.route('/locations/delete/<name>', methods=['POST'])
@mitarbeiter_required
def delete_location(name):
    try:
        # URL-dekodieren
        decoded_name = unquote(name)
        
        # Überprüfen, ob der Standort in Verwendung ist
        tools_with_location = mongodb.db.tools.find_one({'location': decoded_name})
        if tools_with_location:
            return jsonify({
                'success': False,
                'message': 'Der Standort kann nicht gelöscht werden, da er noch von Werkzeugen verwendet wird.'
            })

        # Lade aktuelle Standorte
        settings = mongodb.db.settings.find_one({'key': 'locations'})
        current_locations = []
        if settings and 'value' in settings:
            value = settings['value']
            if isinstance(value, str):
                current_locations = [loc.strip() for loc in value.split(',') if loc.strip()]
            elif isinstance(value, list):
                current_locations = value
        
        # Standort aus der Liste entfernen
        if decoded_name in current_locations:
            current_locations.remove(decoded_name)
            try:
                mongodb.update_one(
                    'settings',
                    {'key': 'locations'},
                    {'value': current_locations},
                    upsert=True
                )
            except Exception as db_error:
                logger.error(f"Datenbankfehler beim Löschen des Standorts: {str(db_error)}")
                # Fallback: Versuche es mit update_one_array
                mongodb.update_one_array(
                    'settings',
                    {'key': 'locations'},
                    {'$set': {'value': current_locations}},
                    upsert=True
                )
        
        return jsonify({
            'success': True,
            'message': 'Standort erfolgreich gelöscht.'
        })
    except Exception as e:
        logger.error(f"Fehler beim Löschen des Standorts: {str(e)}")
        return jsonify({
            'success': False,
            'message': 'Ein Fehler ist aufgetreten.'
        })

# Ticket-Kategorienverwaltung (JSON-API)
@bp.route('/ticket_categories')
@mitarbeiter_required
def get_ticket_categories():
    """Gibt alle Ticket-Kategorien zurück"""
    try:
        settings = mongodb.db.settings.find_one({'key': 'ticket_categories'})
        categories = []
        if settings and 'value' in settings:
            value = settings['value']
            if isinstance(value, str):
                categories = [cat.strip() for cat in value.split(',') if cat.strip()]
            elif isinstance(value, list):
                categories = value
        return jsonify({
            'success': True,
            'categories': [{'name': cat} for cat in categories]
        })
    except Exception as e:
        logger.error(f"Fehler beim Abrufen der Ticket-Kategorien: {str(e)}")
        return jsonify({
            'success': False,
            'message': 'Fehler beim Laden der Ticket-Kategorien'
        })

@bp.route('/ticket_categories/add', methods=['POST'])
@admin_required
def add_ticket_category_json():
    """Fügt eine neue Ticket-Kategorie hinzu (JSON-API)"""
    try:
        # Unterstütze beide Feldnamen für Kompatibilität
        name = request.form.get('name', '').strip() or request.form.get('category', '').strip()
        if not name:
            return jsonify({
                'success': False,
                'message': 'Bitte geben Sie einen Namen ein.'
            })

        # Prüfen ob Ticket-Kategorie bereits existiert
        settings = mongodb.db.settings.find_one({'key': 'ticket_categories'})
        current_categories = []
        if settings and 'value' in settings:
            value = settings['value']
            if isinstance(value, str):
                current_categories = [cat.strip() for cat in value.split(',') if cat.strip()]
            elif isinstance(value, list):
                current_categories = value
        
        # Prüfe auf Duplikate (case-insensitive)
        if any(cat.lower() == name.lower() for cat in current_categories):
            return jsonify({
                'success': False,
                'message': 'Diese Ticket-Kategorie existiert bereits.'
            })

        # Ticket-Kategorie zur Liste hinzufügen
        current_categories.append(name)
        
        # Speichere als Array mit robuster Methode
        try:
            mongodb.update_one(
                'settings',
                {'key': 'ticket_categories'},
                {'value': current_categories},
                upsert=True
            )
        except Exception as db_error:
            logger.error(f"Datenbankfehler beim Speichern der Ticket-Kategorie: {str(db_error)}")
            # Fallback: Versuche es mit update_one_array
            mongodb.update_one_array(
                'settings',
                {'key': 'ticket_categories'},
                {'$set': {'value': current_categories}},
                upsert=True
            )

        return jsonify({
            'success': True,
            'message': 'Ticket-Kategorie erfolgreich hinzugefügt.'
        })
    except Exception as e:
        logger.error(f"Fehler beim Hinzufügen der Ticket-Kategorie: {str(e)}")
        return jsonify({
            'success': False,
            'message': 'Ein Fehler ist aufgetreten.'
        })

@bp.route('/ticket_categories/delete/<name>', methods=['POST'])
@admin_required
def delete_ticket_category_json(name):
    """Löscht eine Ticket-Kategorie (JSON-API)"""
    try:
        # URL-dekodieren
        decoded_name = unquote(name)
        
        # Überprüfen, ob die Ticket-Kategorie in Verwendung ist
        tickets_with_category = mongodb.db.tickets.find_one({'category': decoded_name})
        if tickets_with_category:
            return jsonify({
                'success': False,
                'message': 'Die Ticket-Kategorie kann nicht gelöscht werden, da sie noch von Tickets verwendet wird.'
            })

        # Ticket-Kategorie aus der Liste entfernen
        try:
            mongodb.update_one_array(
                'settings',
                {'key': 'ticket_categories'},
                {'$pull': {'value': decoded_name}}
            )
        except Exception as db_error:
            logger.error(f"Datenbankfehler beim Löschen der Ticket-Kategorie: {str(db_error)}")
            # Fallback: Lade alle Kategorien und entferne die gewünschte
            settings = mongodb.db.settings.find_one({'key': 'ticket_categories'})
            if settings and 'value' in settings:
                current_categories = settings['value']
                if isinstance(current_categories, list) and decoded_name in current_categories:
                    current_categories.remove(decoded_name)
                    mongodb.update_one(
                        'settings',
                        {'key': 'ticket_categories'},
                        {'value': current_categories},
                        upsert=True
                    )
        
        return jsonify({
            'success': True,
            'message': 'Ticket-Kategorie erfolgreich gelöscht.'
        })
    except Exception as e:
        logger.error(f"Fehler beim Löschen der Ticket-Kategorie: {str(e)}")
        return jsonify({
            'success': False,
            'message': 'Ein Fehler ist aufgetreten.'
        })

# Legacy-Routen für Ticket-Kategorien (für Kompatibilität)
@bp.route('/add_ticket_category', methods=['POST'])
@admin_required
def add_ticket_category_legacy():
    """Fügt eine neue Ticket-Kategorie hinzu (Legacy-Route)"""
    try:
        name = request.form.get('category')
        if not name:
            flash('Bitte geben Sie einen Namen ein.', 'error')
            return redirect(url_for('tickets.create'))

        # Prüfen ob Ticket-Kategorie bereits existiert
        settings = mongodb.db.settings.find_one({'key': 'ticket_categories'})
        if settings and name in settings.get('value', []):
            flash('Diese Ticket-Kategorie existiert bereits.', 'error')
            return redirect(url_for('tickets.create'))

        # Ticket-Kategorie zur Liste hinzufügen
        if settings:
            mongodb.update_one_array(
                'settings',
                {'key': 'ticket_categories'},
                {'$push': {'value': name}}
            )
        else:
            mongodb.insert_one('settings', {
                'key': 'ticket_categories',
                'value': [name]
            })

        flash('Ticket-Kategorie erfolgreich hinzugefügt.', 'success')
        return redirect(url_for('tickets.create'))
    except Exception as e:
        logger.error(f"Fehler beim Hinzufügen der Ticket-Kategorie: {str(e)}")
        flash('Ein Fehler ist aufgetreten.', 'error')
        return redirect(url_for('tickets.create'))

@bp.route('/delete_ticket_category/<category>', methods=['POST'])
@admin_required
def delete_ticket_category_legacy(category):
    """Löscht eine Ticket-Kategorie (Legacy-Route)"""
    try:
        # Überprüfen, ob die Ticket-Kategorie in Verwendung ist
        tickets_with_category = mongodb.db.tickets.find_one({'category': category})
        if tickets_with_category:
            flash('Die Ticket-Kategorie kann nicht gelöscht werden, da sie noch von Tickets verwendet wird.', 'error')
            return redirect(url_for('tickets.create'))

        # Ticket-Kategorie aus der Liste entfernen
        mongodb.update_one_array(
            'settings',
            {'key': 'ticket_categories'},
            {'$pull': {'value': category}}
        )
        
        flash('Ticket-Kategorie erfolgreich gelöscht.', 'success')
        return redirect(url_for('tickets.create'))
    except Exception as e:
        logger.error(f"Fehler beim Löschen der Ticket-Kategorie: {str(e)}")
        flash('Ein Fehler ist aufgetreten.', 'error')
        return redirect(url_for('tickets.create'))

@bp.route('/backup/list')
@mitarbeiter_required
def backup_list():
    """Gibt eine Liste der verfügbaren Backups zurück (alle Formate)"""
    try:
        from app.utils.unified_backup_manager import unified_backup_manager
        from app.services.admin_backup_service import AdminBackupService
        
        # Hole neue ZIP-Backups
        zip_backups = unified_backup_manager.list_backups()
        
        # Hole alte JSON-Backups
        json_backups = AdminBackupService.get_backup_list()
        
        # Hole native Backups
        native_backups = AdminBackupService.get_native_backup_list()
        
        # Konvertiere ZIP-Backups in das erwartete Format
        converted_zip_backups = []
        for backup in zip_backups:
            try:
                # Parse created_at String zu Timestamp
                from datetime import datetime
                created_at = backup.get('created_at', '')
                if created_at and created_at != 'Unbekannt':
                    dt = datetime.fromisoformat(created_at.replace('Z', '+00:00'))
                    created_timestamp = dt.timestamp()
                else:
                    created_timestamp = 0
                
                # Hole tatsächliche Dateigröße
                backup_path = backup_manager.backup_dir / backup['filename']
                file_size = backup_path.stat().st_size if backup_path.exists() else 0
                
                converted_zip_backups.append({
                    'name': backup['filename'],
                    'size': file_size,  # Numerische Größe in Bytes
                    'created': created_timestamp,
                    'type': 'zip',
                    'includes_media': backup.get('includes_media', False),
                    'version': backup.get('version', '2.0')
                })
            except Exception as e:
                logger.error(f"Fehler beim Konvertieren von ZIP-Backup {backup}: {e}")
                continue
        
        # Kombiniere alle Backups
        all_backups = converted_zip_backups + json_backups + native_backups
        
        # Sortiere nach Erstellungsdatum (neueste zuerst)
        all_backups.sort(key=lambda x: x.get('created', 0), reverse=True)
        
        return jsonify({
            'status': 'success',
            'backups': all_backups,
            'zip_count': len(converted_zip_backups),
            'json_count': len(json_backups),
            'native_count': len(native_backups),
            'total_count': len(all_backups)
        })
    except Exception as e:
        logger.error(f"Fehler beim Laden der Backups: {str(e)}", exc_info=True)
        return jsonify({
            'status': 'error',
            'message': f'Fehler beim Laden der Backups: {str(e)}'
        }), 500

@bp.route('/backup/create', methods=['POST'])
@admin_required
def create_backup():
    """Erstellt ein neues Backup der aktuellen Datenbank"""
    try:
        success, message, backup_filename = AdminBackupService.create_backup()
        
        if success:
            # E-Mail-Versand (optional)
            email_recipient = request.form.get('email_recipient', '').strip()
            if email_recipient and backup_filename:
                try:
                    from app.utils.email_utils import send_backup_mail
                    from app.utils.backup_manager import backup_manager
                    backup_path = backup_manager.get_backup_path(backup_filename)
                    send_backup_mail(email_recipient, str(backup_path))
                    return jsonify({
                        'status': 'success',
                        'message': f'{message} und an {email_recipient} gesendet',
                        'filename': backup_filename
                    })
                except Exception as e:
                    logger.error(f"Fehler beim E-Mail-Versand: {str(e)}")
                    return jsonify({
                        'status': 'success',
                        'message': f'{message}, aber E-Mail-Versand fehlgeschlagen',
                        'filename': backup_filename
                    })
            else:
                return jsonify({
                    'status': 'success',
                    'message': message,
                    'filename': backup_filename
                })
        else:
            return jsonify({
                'status': 'error',
                'message': message
            }), 500
            
    except Exception as e:
        logger.error(f"Fehler beim Erstellen des Backups: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'Fehler beim Erstellen des Backups: {str(e)}'
        }), 500

@bp.route('/backup/upload', methods=['POST'])
@admin_required
def upload_backup():
    """Lädt ein Backup hoch und stellt es wieder her"""
    try:
        if 'backup_file' not in request.files:
            logger.warning("Backup-Upload: Keine Datei in request.files gefunden")
            return jsonify({
                'status': 'error',
                'message': 'Keine Datei ausgewählt'
            }), 400
            
        file = request.files['backup_file']
        if file.filename == '':
            logger.warning("Backup-Upload: Leerer Dateiname")
            return jsonify({
                'status': 'error',
                'message': 'Keine Datei ausgewählt'
            }), 400
            
        # Prüfe Dateityp - unterstütze ZIP und JSON Backups
        file_extension = Path(file.filename).suffix.lower()
        
        if file_extension not in ['.zip', '.json']:
            logger.warning(f"Backup-Upload: Ungültiger Dateityp: {file.filename}")
            return jsonify({
                'status': 'error',
                'message': 'Nur ZIP- und JSON-Backups sind erlaubt'
            }), 400
        
        # Prüfe Dateigröße
        file.seek(0, 2)  # Gehe zum Ende der Datei
        file_size = file.tell()
        file.seek(0)  # Zurück zum Anfang
        
        logger.info(f"Backup-Upload: Datei {file.filename}, Größe: {file_size} bytes")
        
        if file_size == 0:
            logger.warning("Backup-Upload: Datei ist leer")
            return jsonify({
                'status': 'error',
                'message': 'Die hochgeladene Datei ist leer. Bitte wählen Sie eine gültige Backup-Datei aus.'
            }), 400
            
        # Direkt ins Backup-Verzeichnis speichern
        backup_filename = secure_filename(file.filename)
        backup_path = Path('backups') / backup_filename
        
        # Backup-Verzeichnis erstellen falls nicht vorhanden
        backup_path.parent.mkdir(exist_ok=True)
        
        # Datei direkt ins Backup-Verzeichnis speichern
        file.save(backup_path)
        
        try:
            if file_extension == '.zip':
                # ZIP-Backup über vereinheitlichtes System
                from app.utils.unified_backup_manager import unified_backup_manager
                success = unified_backup_manager.restore_backup(backup_filename)
                
                if success:
                    logger.info("Backup-Upload: ZIP-Backup erfolgreich wiederhergestellt")
                    return jsonify({
                        'status': 'success',
                        'message': f'ZIP-Backup erfolgreich hochgeladen und wiederhergestellt: {file.filename}'
                    })
                else:
                    logger.error("Backup-Upload: Fehler beim Wiederherstellen des ZIP-Backups")
                    return jsonify({
                        'status': 'error',
                        'message': f'Fehler beim Wiederherstellen des ZIP-Backups: {file.filename}'
                    }), 500
            elif file_extension == '.json':
                # JSON-Backup über BackupService
                from app.services.backup_service import BackupService
                backup_service = BackupService()
                success, message = backup_service._restore_from_json(str(backup_path))
                
                if success:
                    logger.info("Backup-Upload: JSON-Backup erfolgreich wiederhergestellt")
                    return jsonify({
                        'status': 'success',
                        'message': f'JSON-Backup erfolgreich hochgeladen und wiederhergestellt: {file.filename}'
                    })
                else:
                    logger.error("Backup-Upload: Fehler beim Wiederherstellen des JSON-Backups")
                    return jsonify({
                        'status': 'error',
                        'message': f'Fehler beim Wiederherstellen des JSON-Backups: {message}'
                    }), 500
                
        except Exception as e:
            # Bei Fehler das fehlerhafte Backup löschen
            if backup_path.exists():
                backup_path.unlink()
            raise e
            
    except Exception as e:
        logger.error(f"Fehler beim Hochladen des Backups: {str(e)}", exc_info=True)
        return jsonify({
            'status': 'error',
            'message': f'Fehler beim Hochladen des Backups: {str(e)}'
        }), 500

@bp.route('/backup/restore/<filename>', methods=['POST'])
@admin_required
def restore_backup(filename):
    """Stellt ein Backup wieder her (automatische Erkennung JSON/Native)"""
    try:
        from app.services.admin_backup_service import AdminBackupService
        from app.utils.backup_manager import backup_manager
        
        # Prüfe Backup-Typ
        backup_path = backup_manager.backup_dir / filename
        is_native = backup_path.is_dir() and filename.startswith('scandy_native_backup_')
        is_zip = backup_path.is_file() and filename.endswith('.zip')
        
        if is_zip:
            # Verwende vereinheitlichtes ZIP-Restore
            from app.utils.unified_backup_manager import unified_backup_manager
            success = unified_backup_manager.restore_backup(filename)
            if success:
                return jsonify({
                    'status': 'success',
                    'message': f'ZIP-Backup erfolgreich wiederhergestellt: {filename}',
                    'backup_type': 'zip'
                })
            else:
                return jsonify({
                    'status': 'error',
                    'message': f'Fehler beim Wiederherstellen des ZIP-Backups: {filename}',
                    'backup_type': 'zip'
                }), 500
        elif is_native:
            # Verwende native Restore
            success, message = AdminBackupService.restore_native_backup(filename)
            if success:
                return jsonify({
                    'status': 'success',
                    'message': f'Native Backup wiederhergestellt: {message}',
                    'backup_type': 'native'
                })
            else:
                return jsonify({
                    'status': 'error',
                    'message': f'Fehler beim Wiederherstellen des nativen Backups: {message}',
                    'backup_type': 'native'
                }), 500
        else:
            # Verwende JSON Restore
            success, message, validation_info = AdminBackupService.restore_backup(filename)
            if success:
                return jsonify({
                    'status': 'success',
                    'message': message,
                    'validation_info': validation_info,
                    'backup_type': 'json'
                })
            else:
                return jsonify({
                    'status': 'error',
                    'message': message,
                    'backup_type': 'json'
                }), 500
            
    except Exception as e:
        logger.error(f"Fehler beim Wiederherstellen des Backups: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'Fehler beim Wiederherstellen des Backups: {str(e)}'
        }), 500

@bp.route('/backup/download/<filename>')
@admin_required
def download_backup(filename):
    """Lädt ein Backup herunter (JSON oder Native als ZIP)"""
    try:
        from pathlib import Path
        from app.utils.backup_manager import backup_manager
        import zipfile
        import tempfile
        import os
        
        backup_path = backup_manager.backup_dir / filename
        
        if not backup_path.exists():
            return jsonify({
                'status': 'error',
                'message': 'Backup nicht gefunden'
            }), 404
        
        # Prüfe Backup-Typ
        is_native = backup_path.is_dir() and filename.startswith('scandy_native_backup_')
        is_zip = backup_path.is_file() and filename.endswith('.zip')
        
        if is_zip:
            # ZIP Backup - direkt senden
            return send_file(
                backup_path,
                as_attachment=True,
                download_name=filename,
                mimetype='application/zip'
            )
        elif is_native:
            # Native Backup (Verzeichnis) - erstelle ZIP
            try:
                # Erstelle temporäres ZIP
                temp_zip = tempfile.NamedTemporaryFile(delete=False, suffix='.zip')
                temp_zip.close()
                
                with zipfile.ZipFile(temp_zip.name, 'w', zipfile.ZIP_DEFLATED) as zipf:
                    # Füge alle Dateien im Backup-Verzeichnis zum ZIP hinzu
                    for root, dirs, files in os.walk(backup_path):
                        for file in files:
                            file_path = os.path.join(root, file)
                            arcname = os.path.relpath(file_path, backup_path)
                            zipf.write(file_path, arcname)
                
                # Sende ZIP-Datei
                return send_file(
                    temp_zip.name,
                    as_attachment=True,
                    download_name=f"{filename}.zip",
                    mimetype='application/zip'
                )
                
            except Exception as e:
                # Lösche temporäre Datei bei Fehler
                if os.path.exists(temp_zip.name):
                    os.unlink(temp_zip.name)
                raise e
                
        else:
            # JSON Backup (Datei) - direkt senden
            return send_file(
                backup_path,
                as_attachment=True,
                download_name=filename,
                mimetype='application/json'
            )
        
    except Exception as e:
        logger.error(f"Fehler beim Herunterladen des Backups: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'Fehler beim Herunterladen des Backups: {str(e)}'
        }), 500

@bp.route('/backup/delete/<filename>', methods=['DELETE'])
@admin_required
def delete_backup(filename):
    """Löscht ein Backup (JSON oder Native)"""
    try:
        from app.utils.backup_manager import backup_manager
        import shutil
        
        backup_path = backup_manager.backup_dir / filename
        
        if not backup_path.exists():
            return jsonify({
                'status': 'error',
                'message': 'Backup nicht gefunden'
            }), 404
        
        # Prüfe Backup-Typ
        is_native = backup_path.is_dir() and filename.startswith('scandy_native_backup_')
        is_zip = backup_path.is_file() and filename.endswith('.zip')
        
        if is_zip:
            # Lösche ZIP Backup (Datei)
            backup_path.unlink()
            backup_type = 'zip'
        elif is_native:
            # Lösche natives Backup (Verzeichnis)
            shutil.rmtree(backup_path)
            backup_type = 'native'
        else:
            # Lösche JSON Backup (Datei)
            backup_path.unlink()
            backup_type = 'json'
        
        return jsonify({
            'status': 'success',
            'message': f'{backup_type.capitalize()} Backup erfolgreich gelöscht',
            'backup_type': backup_type
        })
            
    except Exception as e:
        logger.error(f"Fehler beim Löschen des Backups: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'Fehler beim Löschen des Backups: {str(e)}'
        }), 500

@bp.route('/backup/test/<filename>')
@admin_required
def test_backup(filename):
    """Testet ein Backup ohne es wiederherzustellen"""
    try:
        success, result = AdminBackupService.test_backup(filename)
        
        if success:
            return jsonify({
                'status': 'success',
                'message': 'Backup-Test erfolgreich',
                'data': result
            })
        else:
            return jsonify({
                'status': 'error',
                'message': result
            }), 400
            
    except Exception as e:
        logger.error(f"Fehler beim Testen des Backups: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'Fehler beim Testen des Backups: {str(e)}'
        }), 500

@bp.route('/backup/create-json', methods=['POST'])
@admin_required
def create_json_backup():
    """Erstellt ein JSON-Backup (für Kompatibilität)"""
    try:
        from app.services.admin_backup_service import AdminBackupService
        
        success, message, backup_filename = AdminBackupService.create_json_backup()
        
        if success:
            return jsonify({
                'status': 'success',
                'message': f'JSON-Backup erfolgreich erstellt: {backup_filename}',
                'filename': backup_filename
            })
        else:
            return jsonify({
                'status': 'error',
                'message': f'Fehler beim Erstellen des JSON-Backups: {message}'
            }), 500
            
    except Exception as e:
        logger.error(f"Fehler beim Erstellen des JSON-Backups: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'Fehler beim Erstellen des JSON-Backups: {str(e)}'
        }), 500

@bp.route('/backup/create-native', methods=['POST'])
@admin_required
def create_native_backup():
    """Erstellt ein natives MongoDB-Backup mit mongodump"""
    try:
        from app.services.admin_backup_service import AdminBackupService
        
        success, message, backup_filename = AdminBackupService.create_native_backup()
        
        if success:
            return jsonify({
                'status': 'success',
                'message': f'Native Backup erfolgreich erstellt: {backup_filename}',
                'filename': backup_filename
            })
        else:
            return jsonify({
                'status': 'error',
                'message': f'Fehler beim Erstellen des nativen Backups: {message}'
            }), 500
            
    except Exception as e:
        logger.error(f"Fehler beim Erstellen des nativen Backups: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'Fehler beim Erstellen des nativen Backups: {str(e)}'
        }), 500

@bp.route('/backup/restore-native/<backup_name>', methods=['POST'])
@admin_required
def restore_native_backup(backup_name):
    """Stellt ein natives MongoDB-Backup mit mongorestore wieder her"""
    try:
        from app.services.admin_backup_service import AdminBackupService
        
        success, message = AdminBackupService.restore_native_backup(backup_name)
        
        if success:
            return jsonify({
                'status': 'success',
                'message': f'Native Backup erfolgreich wiederhergestellt: {backup_name}'
            })
        else:
            return jsonify({
                'status': 'error',
                'message': f'Fehler beim Wiederherstellen des nativen Backups: {message}'
            }), 500
            
    except Exception as e:
        logger.error(f"Fehler beim Wiederherstellen des nativen Backups: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'Fehler beim Wiederherstellen des nativen Backups: {str(e)}'
        }), 500

@bp.route('/backup/create-hybrid', methods=['POST'])
@admin_required
def create_hybrid_backup():
    """Erstellt ein hybrides Backup (Native + JSON)"""
    try:
        from app.services.admin_backup_service import AdminBackupService
        
        success, message, result = AdminBackupService.create_hybrid_backup()
        
        if success:
            return jsonify({
                'status': 'success',
                'message': f'Hybrides Backup erfolgreich erstellt: {message}',
                'result': result
            })
        else:
            return jsonify({
                'status': 'error',
                'message': f'Fehler beim Erstellen des hybriden Backups: {message}'
            }), 500
            
    except Exception as e:
        logger.error(f"Fehler beim Erstellen des hybriden Backups: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'Fehler beim Erstellen des hybriden Backups: {str(e)}'
        }), 500

@bp.route('/backup/convert-old/<filename>', methods=['POST'])
@admin_required
def convert_old_backup(filename):
    """Konvertiert ein altes Backup in das neue Format"""
    try:
        from app.utils.backup_manager import backup_manager
        converted_filename = backup_manager.convert_old_backup(filename)
        
        if converted_filename:
            return jsonify({
                'status': 'success',
                'message': f'Backup erfolgreich konvertiert: {converted_filename}',
                'converted_filename': converted_filename
            })
        else:
            return jsonify({
                'status': 'error',
                'message': 'Fehler beim Konvertieren des Backups'
            }), 500
            
    except Exception as e:
        logger.error(f"Fehler beim Konvertieren des Backups: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'Fehler beim Konvertieren des Backups: {str(e)}'
        }), 500

@bp.route('/backup/convert-all-old', methods=['POST'])
@admin_required
def convert_all_old_backups():
    """Konvertiert alle alten Backups automatisch"""
    try:
        from app.utils.backup_manager import backup_manager
        converted_backups = backup_manager.convert_all_old_backups()
        
        return jsonify({
            'status': 'success',
            'message': f'{len(converted_backups)} Backups erfolgreich konvertiert',
            'converted_backups': converted_backups
        })
            
    except Exception as e:
        logger.error(f"Fehler bei der Massenkonvertierung: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'Fehler bei der Massenkonvertierung: {str(e)}'
        }), 500

@bp.route('/backup/list-old', methods=['GET'])
@admin_required
def list_old_backups():
    """Listet alle Backups auf, die noch im alten Format sind"""
    try:
        from app.utils.backup_manager import backup_manager
        old_backups = backup_manager.list_old_backups()
        
        return jsonify({
            'status': 'success',
            'old_backups': old_backups,
            'count': len(old_backups)
        })
            
    except Exception as e:
        logger.error(f"Fehler beim Auflisten alter Backups: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'Fehler beim Auflisten alter Backups: {str(e)}'
        }), 500

@bp.route('/debug/session')
def debug_session():
    """Debug-Route für Session-Informationen"""
    try:
        # Verwende den AdminDebugService
        session_info = AdminDebugService.debug_session_info()
        return jsonify(session_info)
        
    except Exception as e:
        return jsonify({
            'error': str(e)
        }), 500

@bp.route('/debug/backup-info')
@login_required
def debug_backup_info():
    """Debug-Route für Backup-Informationen"""
    try:
        # Verwende den AdminDebugService
        backup_info = AdminDebugService.debug_backup_info()
        return jsonify(backup_info)
    except Exception as e:
        return jsonify({
            'error': str(e),
            'traceback': str(e.__traceback__)
        }), 500

@bp.route('/debug/clear-session')
def clear_session():
    """Löscht die aktuelle Session"""
    try:
        # Verwende den AdminDebugService
        success, message = AdminDebugService.clear_session()
        
        if success:
            return jsonify({
                'status': 'success',
                'message': message
            })
        else:
            return jsonify({
                'status': 'error',
                'message': message
            }), 500
        
    except Exception as e:
        return jsonify({
            'status': 'error',
            'message': str(e)
        }), 500

@bp.route('/debug/fix-session/<username>')
def fix_session(username):
    """Repariert die Session für einen bestimmten Benutzer"""
    try:
        # Verwende den AdminDebugService
        success, message = AdminDebugService.fix_session_for_user(username)
        
        if success:
            return jsonify({
                'status': 'success',
                'message': message
            })
        else:
            return jsonify({
                'status': 'error',
                'message': message
            }), 400
        
    except Exception as e:
        return jsonify({
            'status': 'error',
            'message': str(e)
        }), 500

@bp.route('/debug/normalize-user-ids')
def normalize_user_ids():
    """Normalisiert alle User-IDs in der Datenbank zu Strings"""
    try:
        # Verwende den AdminDebugService
        success, message, stats = AdminDebugService.normalize_user_ids()
        
        if success:
            return jsonify({
                'status': 'success',
                'message': message,
                'statistics': stats
            })
        else:
            return jsonify({
                'status': 'error',
                'message': message
            }), 500
        
    except Exception as e:
        return jsonify({
            'status': 'error',
            'message': str(e)
        }), 500

@bp.route('/debug/normalize-all-ids')
def normalize_all_ids():
    """Normalisiert alle IDs in allen Collections zu Strings"""
    try:
        from app.models.mongodb_database import mongodb
        from bson import ObjectId
        
        collections_to_normalize = [
            'users', 'tickets', 'tools', 'workers', 'consumables', 
            'lendings', 'ticket_messages', 'ticket_notes', 'auftrag_details',
            'auftrag_material', 'auftrag_arbeit', 'timesheets'
        ]
        
        total_updated = 0
        collection_results = {}
        
        for collection in collections_to_normalize:
            try:
                all_docs = mongodb.find(collection, {})
                updated_count = 0
                
                for doc in all_docs:
                    doc_id = doc.get('_id')
                    
                    # Falls die ID ein ObjectId ist, konvertiere sie zu String
                    if isinstance(doc_id, ObjectId):
                        string_id = str(doc_id)
                        
                        # Erstelle ein neues Dokument mit String-ID
                        new_doc = doc.copy()
                        new_doc['_id'] = string_id
                        
                        # Lösche das alte Dokument und füge das neue ein
                        mongodb.delete_one(collection, {'_id': doc_id})
                        mongodb.insert_one(collection, new_doc)
                        
                        updated_count += 1
                
                collection_results[collection] = updated_count
                total_updated += updated_count
                print(f"Collection {collection}: {updated_count} IDs normalisiert")
                
            except Exception as e:
                collection_results[collection] = f"Fehler: {str(e)}"
                print(f"Fehler bei Collection {collection}: {e}")
        
        return jsonify({
            'status': 'success',
            'message': f'{total_updated} IDs in allen Collections normalisiert',
            'total_updated': total_updated,
            'collection_results': collection_results
        })
        
    except Exception as e:
        return jsonify({
            'status': 'error',
            'message': str(e)
        }), 500

@bp.route('/debug/user-management')
@login_required
@admin_required
def debug_user_management():
    """Debug-Route für User-Management-Probleme"""
    try:
        from app.models.mongodb_models import MongoDBUser
        from app.models.mongodb_database import mongodb
        
        # Hole alle User mit detaillierten Informationen
        all_users = mongodb.find('users', {})
        user_list = []
        
        for user in all_users:
            user_id = user.get('_id')
            user_list.append({
                'id': str(user_id),
                'id_type': type(user_id).__name__,
                'username': user.get('username'),
                'role': user.get('role'),
                'is_active': user.get('is_active', True),
                'email': user.get('email', 'N/A'),
                'firstname': user.get('firstname', 'N/A'),
                'lastname': user.get('lastname', 'N/A'),
                'created_at': str(user.get('created_at', 'N/A')),
                'updated_at': str(user.get('updated_at', 'N/A'))
            })
        
        # Teste verschiedene Suchmethoden
        test_results = []
        if user_list:
            test_user = user_list[0]
            test_id = test_user['id']
            
            # Test 1: Direkte String-Suche
            try:
                direct_result = mongodb.find_one('users', {'_id': test_id})
                test_results.append({
                    'method': 'Direkte String-Suche',
                    'success': direct_result is not None,
                    'result': str(direct_result.get('username') if direct_result else 'Nicht gefunden')
                })
            except Exception as e:
                test_results.append({
                    'method': 'Direkte String-Suche',
                    'success': False,
                    'result': f'Fehler: {str(e)}'
                })
            
            # Test 2: ObjectId-Suche
            try:
                from bson import ObjectId
                obj_id = ObjectId(test_id)
                obj_result = mongodb.find_one('users', {'_id': obj_id})
                test_results.append({
                    'method': 'ObjectId-Suche',
                    'success': obj_result is not None,
                    'result': str(obj_result.get('username') if obj_result else 'Nicht gefunden')
                })
            except Exception as e:
                test_results.append({
                    'method': 'ObjectId-Suche',
                    'success': False,
                    'result': f'Fehler: {str(e)}'
                })
            
            # Test 3: find_user_by_id
            try:
                user_result = find_user_by_id(test_id)
                test_results.append({
                    'method': 'find_user_by_id',
                    'success': user_result is not None,
                    'result': str(user_result.get('username') if user_result else 'Nicht gefunden')
                })
            except Exception as e:
                test_results.append({
                    'method': 'find_user_by_id',
                    'success': False,
                    'result': f'Fehler: {str(e)}'
                })
            
            # Test 4: MongoDBUser.get_by_id
            try:
                mongo_user_result = MongoDBUser.get_by_id(test_id)
                test_results.append({
                    'method': 'MongoDBUser.get_by_id',
                    'success': mongo_user_result is not None,
                    'result': str(mongo_user_result.get('username') if mongo_user_result else 'Nicht gefunden')
                })
            except Exception as e:
                test_results.append({
                    'method': 'MongoDBUser.get_by_id',
                    'success': False,
                    'result': f'Fehler: {str(e)}'
                })
        
        return jsonify({
            'status': 'success',
            'total_users': len(user_list),
            'users': user_list,
            'test_results': test_results,
            'test_id_used': test_user['id'] if user_list else 'Keine User vorhanden'
        })
        
    except Exception as e:
        return jsonify({
            'status': 'error',
            'message': str(e)
        }), 500

@bp.route('/debug/test-user-id/<user_id>')
@login_required
@admin_required
def debug_test_user_id(user_id):
    """Testet eine spezifische User-ID"""
    try:
        # Verwende den AdminDebugService
        test_results = AdminDebugService.test_user_id(user_id)
        return jsonify(test_results)
        
    except Exception as e:
        logger.error(f"Fehler beim Testen der User-ID {user_id}: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        })

@bp.route('/available-logos')
@mitarbeiter_required
def available_logos():
    """Gibt eine Liste der verfügbaren Logos zurück"""
    try:
        # Verwende den AdminDebugService
        logos = AdminDebugService.get_available_logos()
        return jsonify({
            'success': True,
            'logos': logos
        })
    except Exception as e:
        logger.error(f"Fehler beim Laden der Logos: {str(e)}")
        return jsonify({
            'success': False,
            'message': 'Fehler beim Laden der Logos'
        }), 500



@bp.route('/tickets/<ticket_id>/update-assignment', methods=['POST'])
@login_required
@admin_required
def update_ticket_assignment(ticket_id):
    """Aktualisiert die Zuweisung eines Tickets"""
    try:
        if not request.is_json:
            return jsonify({'success': False, 'message': 'Ungültiges Anfrageformat'}), 400

        data = request.get_json()
        assigned_to = data.get('assigned_to')
        
        # Wenn assigned_to leer ist, setze es auf None
        if not assigned_to or assigned_to == "":
            assigned_to = None

        # Verwende die ursprüngliche ID direkt für das Update
        from bson import ObjectId
        try:
            # Versuche zuerst mit ObjectId
            ticket_id_for_update = ObjectId(ticket_id)
        except:
            # Falls das fehlschlägt, verwende die ursprüngliche ID als String
            ticket_id_for_update = ticket_id

        # Aktualisiere die Zuweisung direkt im Ticket
        if not mongodb.update_one('tickets', {'_id': ticket_id_for_update}, {'$set': {'assigned_to': assigned_to, 'updated_at': datetime.now()}}):
            return jsonify({'success': False, 'message': 'Fehler beim Aktualisieren der Zuweisung'})

        return jsonify({'success': True, 'message': 'Zuweisung erfolgreich aktualisiert'})

    except Exception as e:
        logger.error(f"Fehler beim Aktualisieren der Zuweisung: {str(e)}")
        return jsonify({'success': False, 'message': str(e)}), 500

@bp.route('/tickets/<ticket_id>/update-status', methods=['POST'])
@login_required
@admin_required
def update_ticket_status(ticket_id):
    """Ticket-Status aktualisieren"""
    try:
        data = request.get_json()
        if not data or 'status' not in data:
            return jsonify({'success': False, 'message': 'Status nicht angegeben'}), 400
            
        new_status = data['status']
        
        # Verwende die ursprüngliche ID direkt für das Update
        from bson import ObjectId
        try:
            # Versuche zuerst mit ObjectId
            ticket_id_for_update = ObjectId(ticket_id)
        except:
            # Falls das fehlschlägt, verwende die ursprüngliche ID als String
            ticket_id_for_update = ticket_id
        
        # Prüfe ob das Ticket existiert
        ticket = mongodb.find_one('tickets', {'_id': ticket_id_for_update})
        if not ticket:
            return jsonify({'success': False, 'message': 'Ticket nicht gefunden'}), 404
            
        # Aktualisiere den Status
        mongodb.update_one('tickets', 
                          {'_id': ticket_id_for_update}, 
                          {
                              '$set': {
                                  'status': new_status,
                                  'updated_at': datetime.now().strftime('%d.%m.%Y %H:%M:%S')
                              }
                          })
        
        return jsonify({
            'success': True,
            'message': f'Status wurde auf "{new_status}" geändert'
        })

    except Exception as e:
        logger.error(f"Fehler beim Aktualisieren des Status: {str(e)}")
        return jsonify({'success': False, 'message': str(e)}), 500

@bp.route('/tickets/<ticket_id>/delete', methods=['POST'])
@login_required
@admin_required
def delete_ticket(ticket_id):
    """Ticket soft löschen (markieren als gelöscht)"""
    try:
        # Verwende den AdminTicketService
        success, message = AdminTicketService.delete_ticket(ticket_id, permanent=False)
        
        if success:
            return jsonify({
                'success': True,
                'message': message
            })
        else:
            return jsonify({
                'success': False,
                'message': message
            }), 400
        
    except Exception as e:
        logger.error(f"Fehler beim Löschen des Tickets: {str(e)}", exc_info=True)
        return jsonify({
            'success': False,
            'message': f'Fehler beim Löschen: {str(e)}'
        }), 500

@bp.route('/tickets/<ticket_id>/delete-permanent', methods=['DELETE'])
@login_required
@admin_required
def delete_ticket_permanent(ticket_id):
    """Ticket endgültig löschen"""
    try:
        # Verwende die ursprüngliche ID direkt für das Update
        from bson import ObjectId
        try:
            # Versuche zuerst mit ObjectId
            ticket_id_for_update = ObjectId(ticket_id)
        except:
            # Falls das fehlschlägt, verwende die ursprüngliche ID als String
            ticket_id_for_update = ticket_id
        
        # Prüfe ob das Ticket existiert und gelöscht ist
        ticket = mongodb.find_one('tickets', {'_id': ticket_id_for_update, 'deleted': True})
        
        if not ticket:
            return jsonify({
                'success': False,
                'message': 'Ticket nicht gefunden oder nicht gelöscht'
            }), 404
            
        # Lösche das Ticket und alle zugehörigen Daten endgültig
        mongodb.delete_one('tickets', {'_id': ticket_id_for_update})
        mongodb.delete_many('ticket_notes', {'ticket_id': ticket_id_for_update})
        mongodb.delete_many('ticket_messages', {'ticket_id': ticket_id_for_update})
        mongodb.delete_many('ticket_assignments', {'ticket_id': ticket_id_for_update})
        mongodb.delete_one('auftrag_details', {'ticket_id': ticket_id_for_update})
        mongodb.delete_many('auftrag_material', {'ticket_id': ticket_id_for_update})
        mongodb.delete_many('auftrag_arbeit', {'ticket_id': ticket_id_for_update})
        
        return jsonify({
            'success': True,
            'message': 'Ticket wurde endgültig gelöscht'
        })
        
    except Exception as e:
        logger.error(f"Fehler beim endgültigen Löschen des Tickets: {str(e)}", exc_info=True)
        return jsonify({
            'success': False,
            'message': f'Fehler beim Löschen: {str(e)}'
        }), 500











@bp.route('/delete_user/<user_id>', methods=['POST'])
@mitarbeiter_required
def delete_user(user_id):
    """Benutzer löschen"""
    try:
        user = AdminUserService.get_user_by_id(user_id)
        if not user:
            flash('Benutzer nicht gefunden', 'error')
            return redirect(url_for('admin.manage_users'))
        
        # Verhindere das Löschen des eigenen Accounts
        if user['username'] == current_user.username:
            flash('Sie können Ihren eigenen Account nicht löschen', 'error')
            return redirect(url_for('admin.manage_users'))
        
        # Benutzer löschen mit AdminUserService
        success, message = AdminUserService.delete_user(user_id)
        
        if success:
            flash(message, 'success')
        else:
            flash(message, 'error')
        
        return redirect(url_for('admin.manage_users'))
        
    except Exception as e:
        logger.error(f"Fehler beim Löschen des Benutzers: {e}")
        flash('Fehler beim Löschen des Benutzers', 'error')
        return redirect(url_for('admin.manage_users'))









@bp.route('/user_form')
@mitarbeiter_required
def user_form():
    """Benutzer-Formular (für neue Benutzer)"""
    return render_template('admin/user_form.html', roles=['admin', 'mitarbeiter', 'anwender', 'teilnehmer'])





@bp.route('/export_all_data')
@admin_required
def export_all_data():
    """Exportiert alle Daten als Excel-Datei"""
    try:
        # Hole alle Daten aus der Datenbank
        tools = list(mongodb.find('tools', {'deleted': {'$ne': True}}))
        workers = list(mongodb.find('workers', {'deleted': {'$ne': True}}))
        consumables = list(mongodb.find('consumables', {'deleted': {'$ne': True}}))
        lendings = list(mongodb.find('lendings', {}))
        settings = list(mongodb.find('settings', {}))
        
        # Erstelle Excel-Datei mit mehreren Arbeitsblättern
        data_dict = {
            'Werkzeuge': tools,
            'Mitarbeiter': workers,
            'Verbrauchsmaterial': consumables,
            'Ausleihverlauf': lendings,
            'Settings': settings
        }
        
        # Erstelle Excel-Datei
        excel_file = create_multi_sheet_excel(data_dict)
        
        # Sende Datei als Download
        return send_file(
            excel_file,
            as_attachment=True,
            download_name=f'scandy_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        logger.error(f"Fehler beim Exportieren der Daten: {str(e)}")
        flash('Fehler beim Exportieren der Daten', 'error')
        return redirect(url_for('admin.system'))

def fix_id_for_import(doc):
    """
    Wandelt das _id-Feld in einen echten ObjectId um, wenn möglich.
    Verbesserte Version für robuste Importe.
    """
    if '_id' in doc:
        # Falls _id ein String ist und wie eine ObjectId aussieht
        if isinstance(doc['_id'], str) and len(doc['_id']) == 24:
            try:
                doc['_id'] = ObjectId(doc['_id'])
            except Exception:
                # Falls es keine gültige ObjectId ist, entferne das Feld
                # MongoDB wird automatisch eine neue ObjectId generieren
                del doc['_id']
        # Falls _id bereits eine ObjectId ist, belasse es
        elif isinstance(doc['_id'], ObjectId):
            pass
        # Falls _id ein anderer Typ ist, entferne es
        else:
            del doc['_id']
    
    # Entferne NaN-Werte und leere Strings
    cleaned_doc = {}
    for key, value in doc.items():
        if pd.isna(value) or value == '' or value == 'nan':
            continue
        cleaned_doc[key] = value
    
    return cleaned_doc

@bp.route('/import_all_data', methods=['POST'])
@admin_required
def import_all_data():
    """Importiert Daten aus einer Excel-Datei"""
    try:
        if 'file' not in request.files:
            flash('Keine Datei ausgewählt', 'error')
            return redirect(url_for('admin.system'))
            
        file = request.files['file']
        if file.filename == '':
            flash('Keine Datei ausgewählt', 'error')
            return redirect(url_for('admin.system'))
            
        if not file.filename.endswith('.xlsx'):
            flash('Nur Excel-Dateien (.xlsx) werden unterstützt', 'error')
            return redirect(url_for('admin.system'))
        
        # Speichere temporäre Datei
        import tempfile
        import os
        
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
            file.save(tmp_file.name)
            tmp_file_path = tmp_file.name
        
        try:
            # Lese Excel-Datei
            import pandas as pd
            
            # Lese alle Arbeitsblätter
            excel_file = pd.ExcelFile(tmp_file_path)
            
            imported_count = 0
            errors = []
            
            # Importiere Werkzeuge
            if 'Werkzeuge' in excel_file.sheet_names:
                try:
                    df_tools = pd.read_excel(excel_file, sheet_name='Werkzeuge')
                    for index, row in df_tools.iterrows():
                        try:
                            tool_data = row.to_dict()
                            tool_data = fix_id_for_import(tool_data)
                            
                            # Prüfe ob Barcode vorhanden ist
                            if not tool_data.get('barcode'):
                                errors.append(f"Zeile {index + 2}: Werkzeug ohne Barcode übersprungen")
                                continue
                            
                            existing = mongodb.find_one('tools', {'barcode': tool_data.get('barcode')})
                            if not existing:
                                mongodb.insert_one('tools', tool_data)
                                imported_count += 1
                            else:
                                # Update existierendes Werkzeug
                                mongodb.update_one('tools', {'barcode': tool_data.get('barcode')}, {'$set': tool_data})
                                imported_count += 1
                        except Exception as e:
                            errors.append(f"Zeile {index + 2}: Fehler bei Werkzeug: {str(e)}")
                except Exception as e:
                    errors.append(f"Fehler beim Lesen der Werkzeuge-Tabelle: {str(e)}")
            
            # Importiere Mitarbeiter
            if 'Mitarbeiter' in excel_file.sheet_names:
                try:
                    df_workers = pd.read_excel(excel_file, sheet_name='Mitarbeiter')
                    for index, row in df_workers.iterrows():
                        try:
                            worker_data = row.to_dict()
                            worker_data = fix_id_for_import(worker_data)
                            
                            # Prüfe ob Barcode vorhanden ist
                            if not worker_data.get('barcode'):
                                errors.append(f"Zeile {index + 2}: Mitarbeiter ohne Barcode übersprungen")
                                continue
                            
                            existing = mongodb.find_one('workers', {'barcode': worker_data.get('barcode')})
                            if not existing:
                                mongodb.insert_one('workers', worker_data)
                                imported_count += 1
                            else:
                                # Update existierenden Mitarbeiter
                                mongodb.update_one('workers', {'barcode': worker_data.get('barcode')}, {'$set': worker_data})
                                imported_count += 1
                        except Exception as e:
                            errors.append(f"Zeile {index + 2}: Fehler bei Mitarbeiter: {str(e)}")
                except Exception as e:
                    errors.append(f"Fehler beim Lesen der Mitarbeiter-Tabelle: {str(e)}")
            
            # Importiere Verbrauchsmaterial
            if 'Verbrauchsmaterial' in excel_file.sheet_names:
                try:
                    df_consumables = pd.read_excel(excel_file, sheet_name='Verbrauchsmaterial')
                    for index, row in df_consumables.iterrows():
                        try:
                            consumable_data = row.to_dict()
                            consumable_data = fix_id_for_import(consumable_data)
                            
                            # Prüfe ob Barcode vorhanden ist
                            if not consumable_data.get('barcode'):
                                errors.append(f"Zeile {index + 2}: Verbrauchsmaterial ohne Barcode übersprungen")
                                continue
                            
                            existing = mongodb.find_one('consumables', {'barcode': consumable_data.get('barcode')})
                            if not existing:
                                mongodb.insert_one('consumables', consumable_data)
                                imported_count += 1
                            else:
                                # Update existierendes Verbrauchsmaterial
                                mongodb.update_one('consumables', {'barcode': consumable_data.get('barcode')}, {'$set': consumable_data})
                                imported_count += 1
                        except Exception as e:
                            errors.append(f"Zeile {index + 2}: Fehler bei Verbrauchsmaterial: {str(e)}")
                except Exception as e:
                    errors.append(f"Fehler beim Lesen der Verbrauchsmaterial-Tabelle: {str(e)}")
            
            # Importiere Settings (Kategorien, Standorte, Abteilungen)
            if 'Settings' in excel_file.sheet_names:
                try:
                    df_settings = pd.read_excel(excel_file, sheet_name='Settings')
                    for index, row in df_settings.iterrows():
                        try:
                            setting_data = row.to_dict()
                            setting_data = fix_id_for_import(setting_data)
                            valid_settings = ['categories', 'locations', 'departments', 'ticket_categories', 
                                            'label_tools_name', 'label_tools_icon', 'label_consumables_name', 
                                            'label_consumables_icon', 'label_tickets_name', 'label_tickets_icon']
                            
                            if setting_data.get('key') in valid_settings:
                                existing = mongodb.find_one('settings', {'key': setting_data.get('key')})
                                if not existing:
                                    mongodb.insert_one('settings', setting_data)
                                    imported_count += 1
                                else:
                                    mongodb.update_one('settings', 
                                                     {'key': setting_data.get('key')}, 
                                                     {'$set': setting_data})
                                    imported_count += 1
                            else:
                                errors.append(f"Zeile {index + 2}: Ungültige Setting '{setting_data.get('key')}' übersprungen")
                        except Exception as e:
                            errors.append(f"Zeile {index + 2}: Fehler bei Setting: {str(e)}")
                except Exception as e:
                    errors.append(f"Fehler beim Lesen der Settings-Tabelle: {str(e)}")
            
            # Zeige Erfolgsmeldung und eventuelle Fehler
            if errors:
                error_message = f'{imported_count} Datensätze importiert. {len(errors)} Fehler aufgetreten: ' + '; '.join(errors[:5])
                if len(errors) > 5:
                    error_message += f'... und {len(errors) - 5} weitere'
                flash(error_message, 'warning')
            else:
                flash(f'{imported_count} Datensätze erfolgreich importiert', 'success')
            
        finally:
            os.unlink(tmp_file_path)
        
        return redirect(url_for('admin.system'))
        
    except Exception as e:
        logger.error(f"Fehler beim Importieren der Daten: {str(e)}")
        flash(f'Fehler beim Importieren: {str(e)}', 'error')
        return redirect(url_for('admin.system'))

@bp.route('/reset_password', methods=['GET', 'POST'])
def reset_password():
    """Passwort-Reset per E-Mail"""
    if request.method == 'POST':
        email = request.form.get('email', '').strip()
        
        if not email:
            flash('Bitte geben Sie eine E-Mail-Adresse ein.', 'error')
            return render_template('auth/reset_password.html')
        
        # Prüfe ob Benutzer mit dieser E-Mail existiert
        user = mongodb.find_one('users', {'email': email})
        if not user:
            flash('Kein Benutzer mit dieser E-Mail-Adresse gefunden.', 'error')
            return render_template('auth/reset_password.html')
        
        # Generiere sicheres neues Passwort (wie bei add_user)
        import secrets
        import string
        # Mindestens 1 von jeder Kategorie sicherstellen
        password = (
            secrets.choice(string.ascii_uppercase) +  # 1 Großbuchstabe
            secrets.choice(string.ascii_lowercase) +  # 1 Kleinbuchstabe
            secrets.choice(string.digits) +           # 1 Ziffer
            secrets.choice("!@#$%^&*") +              # 1 Sonderzeichen
            ''.join(secrets.choice(string.ascii_letters + string.digits + "!@#$%^&*") for _ in range(8))  # 8 weitere zufällige Zeichen
        )
        # Passwort mischen
        password_list = list(password)
        secrets.SystemRandom().shuffle(password_list)
        password = ''.join(password_list)
        
        # Hash das neue Passwort
        from werkzeug.security import generate_password_hash
        password_hash = generate_password_hash(password)
        
        # Aktualisiere das Passwort in der Datenbank
        try:
            result = mongodb.update_one('users', 
                                     {'_id': convert_id_for_query(user['_id'])}, 
                                     {'$set': {'password_hash': password_hash, 'updated_at': datetime.now()}})
            
            # Prüfe ob das Update erfolgreich war (result ist ein bool)
            if not result:
                logger.error(f"Fehler beim Aktualisieren des Passworts für Benutzer {user['username']} - Update fehlgeschlagen")
                flash('Fehler beim Zurücksetzen des Passworts.', 'error')
                return render_template('auth/reset_password.html')
            
            logger.info(f"Passwort erfolgreich zurückgesetzt für Benutzer {user['username']} ({email})")
            
        except Exception as e:
            logger.error(f"Fehler beim Aktualisieren des Passworts in der Datenbank: {e}")
            flash('Fehler beim Zurücksetzen des Passworts.', 'error')
            return render_template('auth/reset_password.html')
        
        # Sende E-Mail mit neuem Passwort
        try:
            from app.utils.email_utils import send_password_reset_mail
            send_result = send_password_reset_mail(email, password=password)
            if send_result:
                flash('Ein neues Passwort wurde an Ihre E-Mail-Adresse gesendet.', 'success')
                logger.info(f"Passwort-Reset-E-Mail erfolgreich an {email} gesendet")
            else:
                flash('Passwort wurde zurückgesetzt, aber E-Mail konnte nicht versendet werden.', 'warning')
                logger.warning(f"Passwort-Reset-E-Mail konnte nicht an {email} gesendet werden")
        except Exception as e:
            logger.error(f"Fehler beim Versenden der E-Mail: {e}")
            flash('Passwort wurde zurückgesetzt, aber E-Mail konnte nicht versendet werden.', 'warning')
        
        return redirect(url_for('auth.login'))
    
    return render_template('auth/reset_password.html')



@bp.route('/backup/auto/status')
@login_required
@admin_required
def auto_backup_status():
    """Gibt den Status des automatischen Backup-Systems zurück"""
    try:
        from app.utils.auto_backup import get_auto_backup_status
        status = get_auto_backup_status()
        return jsonify({
            'success': True,
            'status': status
        })
    except Exception as e:
        logger.error(f"Fehler beim Abrufen des Auto-Backup-Status: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'Fehler beim Abrufen des Status: {str(e)}'
        })

@bp.route('/backup/auto/start', methods=['POST'])
@login_required
@admin_required
def start_auto_backup():
    """Startet das automatische Backup-System"""
    try:
        from app.utils.auto_backup import start_auto_backup
        start_auto_backup()
        return jsonify({
            'success': True,
            'message': 'Automatisches Backup-System gestartet'
        })
    except Exception as e:
        logger.error(f"Fehler beim Starten des Auto-Backup-Systems: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'Fehler beim Starten: {str(e)}'
        })

@bp.route('/backup/auto/stop', methods=['POST'])
@login_required
@admin_required
def stop_auto_backup():
    """Stoppt das automatische Backup-System"""
    try:
        from app.utils.auto_backup import stop_auto_backup
        stop_auto_backup()
        return jsonify({
            'success': True,
            'message': 'Automatisches Backup-System gestoppt'
        })
    except Exception as e:
        logger.error(f"Fehler beim Stoppen des Auto-Backup-Systems: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'Fehler beim Stoppen: {str(e)}'
        })

@bp.route('/backup/auto/logs')
@login_required
@admin_required
def auto_backup_logs():
    """Gibt die Auto-Backup-Logs zurück"""
    try:
        from app.utils.auto_backup import auto_backup_scheduler
        log_file = auto_backup_scheduler.log_file
        
        if log_file.exists():
            with open(log_file, 'r', encoding='utf-8') as f:
                logs = f.readlines()
            # Letzte 50 Zeilen
            recent_logs = logs[-50:] if len(logs) > 50 else logs
            return jsonify({
                'success': True,
                'logs': recent_logs
            })
        else:
            return jsonify({
                'success': True,
                'logs': ['Keine Logs verfügbar']
            })
    except Exception as e:
        logger.error(f"Fehler beim Abrufen der Auto-Backup-Logs: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'Fehler beim Abrufen der Logs: {str(e)}'
        })

@bp.route('/auto-backup', methods=['GET', 'POST'])
@login_required
@admin_required
def auto_backup():
    """Automatisches Backup-System Verwaltungsseite"""
    try:
        from app.utils.auto_backup import auto_backup_scheduler
        
        if request.method == 'POST':
            action = request.form.get('action')
            
            if action == 'save_times':
                # Backup-Zeiten speichern
                times_input = request.form.get('backup_times', '')
                times_list = [t.strip() for t in times_input.split(',') if t.strip()]
                
                success, message = auto_backup_scheduler.save_backup_times(times_list)
                
                if success:
                    flash('Backup-Zeiten erfolgreich gespeichert.', 'success')
                else:
                    flash(f'Fehler beim Speichern der Backup-Zeiten: {message}', 'error')
            
            elif action == 'save_weekly_time':
                # Wöchentliche Backup-Zeit speichern
                weekly_time = request.form.get('weekly_backup_time', '').strip()
                
                success, message = auto_backup_scheduler.save_weekly_backup_time(weekly_time)
                
                if success:
                    flash('Wöchentliche Backup-Zeit erfolgreich gespeichert.', 'success')
                else:
                    flash(f'Fehler beim Speichern der wöchentlichen Backup-Zeit: {message}', 'error')
            
            elif action == 'save_weekly_email':
                # E-Mail für wöchentliche Backups speichern
                weekly_email = request.form.get('weekly_backup_email', '').strip()
                
                if weekly_email and '@' in weekly_email:
                    try:
                        from app.models.mongodb_database import mongodb
                        mongodb.update_one('settings', 
                                         {'key': 'weekly_backup_email'}, 
                                         {'$set': {'value': weekly_email}}, 
                                         upsert=True)
                        flash('E-Mail-Adresse für wöchentliche Backups erfolgreich gespeichert.', 'success')
                    except Exception as e:
                        logger.error(f"Fehler beim Speichern der wöchentlichen Backup-E-Mail: {e}")
                        flash(f'Fehler beim Speichern der E-Mail-Adresse: {str(e)}', 'error')
                else:
                    flash('Bitte geben Sie eine gültige E-Mail-Adresse ein.', 'error')
        
        # Lade aktuelle Einstellungen
        current_times = auto_backup_scheduler.get_backup_times()
        current_weekly_time = auto_backup_scheduler.get_weekly_backup_time()
        current_weekly_email = auto_backup_scheduler._get_weekly_backup_email()
        
        return render_template('admin/auto_backup.html', 
                             backup_times=current_times,
                             weekly_backup_time=current_weekly_time,
                             weekly_backup_email=current_weekly_email)
        
    except Exception as e:
        logger.error(f"Fehler bei Auto-Backup-Einstellungen: {e}")
        flash('Fehler beim Laden der Auto-Backup-Einstellungen.', 'error')
        return render_template('admin/auto_backup.html', 
                             backup_times=['06:00', '18:00'],
                             weekly_backup_time='17:00',
                             weekly_backup_email='')

@bp.route('/email_settings', methods=['GET', 'POST'])
@admin_required
def email_settings():
    """E-Mail-Konfiguration verwalten"""
    try:
        # Automatische E-Mail-Reparatur beim Aufruf der E-Mail-Einstellungen
        try:
            AdminDebugService.fix_email_configuration()
            logger.info("Automatische E-Mail-Reparatur durchgeführt")
        except Exception as e:
            logger.warning(f"Automatische E-Mail-Reparatur fehlgeschlagen: {e}")
        
        # Session-Persistierung vor dem Speichern
        current_user_id = session.get('user_id')
        current_username = session.get('username')
        current_role = session.get('role')
        current_authenticated = session.get('is_authenticated', False)
        
        if request.method == 'POST':
            action = request.form.get('action')
            
            if action == 'save':
                # E-Mail-Konfiguration speichern
                use_auth = request.form.get('use_auth') == 'on'
                
                if use_auth:
                    # Mit Authentifizierung
                    new_password = request.form.get('mail_password', '').strip()
                    
                    # Wenn kein neues Passwort eingegeben wurde, verwende das gespeicherte
                    if not new_password:
                        stored_config = AdminEmailService.get_email_config()
                        if stored_config and stored_config.get('mail_password'):
                            new_password = stored_config['mail_password']
                    
                    config_data = {
                        'mail_server': request.form.get('mail_server', 'smtp.gmail.com'),
                        'mail_port': int(request.form.get('mail_port', 587)),
                        'mail_use_tls': request.form.get('mail_use_tls') == 'on',
                        'mail_username': request.form.get('mail_username', ''),
                        'mail_password': new_password,
                        'test_email': request.form.get('test_email', ''),
                        'use_auth': True
                    }
                else:
                    # Ohne Authentifizierung
                    config_data = {
                        'mail_server': request.form.get('mail_server', 'smtp.gmail.com'),
                        'mail_port': int(request.form.get('mail_port', 587)),
                        'mail_use_tls': request.form.get('mail_use_tls') == 'on',
                        'mail_username': request.form.get('sender_email', ''),  # Absender-E-Mail
                        'mail_password': '',  # Kein Passwort
                        'test_email': request.form.get('test_email', ''),
                        'use_auth': False
                    }
                
                success, message = AdminEmailService.save_email_config(config_data)
                
                # Session wiederherstellen nach dem Speichern
                if current_user_id:
                    session['user_id'] = current_user_id
                if current_username:
                    session['username'] = current_username
                if current_role:
                    session['role'] = current_role
                if current_authenticated:
                    session['is_authenticated'] = current_authenticated
                
                if success:
                    flash(message, 'success')
                else:
                    flash(message, 'error')
                    
            elif action == 'test':
                # E-Mail-Konfiguration testen
                use_auth = request.form.get('use_auth') == 'on'
                
                if use_auth:
                    # Mit Authentifizierung
                    config_data = {
                        'mail_server': request.form.get('mail_server', 'smtp.gmail.com'),
                        'mail_port': int(request.form.get('mail_port', 587)),
                        'mail_use_tls': request.form.get('mail_use_tls') == 'on',
                        'mail_username': request.form.get('mail_username', ''),
                        'mail_password': request.form.get('mail_password', ''),
                        'test_email': request.form.get('test_email', '')
                    }
                    
                    # Wenn kein neues Passwort eingegeben wurde, verwende das gespeicherte
                    if not config_data['mail_password']:
                        stored_config = AdminEmailService.get_email_config()
                        if stored_config and stored_config.get('mail_password'):
                            config_data['mail_password'] = stored_config['mail_password']
                else:
                    # Ohne Authentifizierung
                    config_data = {
                        'mail_server': request.form.get('mail_server', 'smtp.gmail.com'),
                        'mail_port': int(request.form.get('mail_port', 587)),
                        'mail_use_tls': request.form.get('mail_use_tls') == 'on',
                        'mail_username': request.form.get('sender_email', ''),  # Absender-E-Mail
                        'mail_password': '',  # Kein Passwort
                        'test_email': request.form.get('test_email', '')
                    }
                
                success, message = AdminEmailService.test_email_config(config_data)
                
                # Session wiederherstellen nach dem Test
                if current_user_id:
                    session['user_id'] = current_user_id
                if current_username:
                    session['username'] = current_username
                if current_role:
                    session['role'] = current_role
                if current_authenticated:
                    session['is_authenticated'] = current_authenticated
                
                if success:
                    flash(f'E-Mail-Test erfolgreich: {message}', 'success')
                else:
                    flash(f'E-Mail-Test fehlgeschlagen: {message}', 'error')
        
        # Lade aktuelle Konfiguration
        config = AdminEmailService.get_email_config()
        
        # Entferne das Passwort aus der Konfiguration für das Template
        if config and 'mail_password' in config:
            config['mail_password'] = ''  # Leeres Feld, da Passwort verschlüsselt ist
        
        return render_template('admin/email_settings.html', config=config)
        
    except Exception as e:
        logger.error(f"Fehler bei E-Mail-Einstellungen: {e}")
        
        # Session wiederherstellen bei Fehlern
        if current_user_id:
            session['user_id'] = current_user_id
        if current_username:
            session['username'] = current_username
        if current_role:
            session['role'] = current_role
        if current_authenticated:
            session['is_authenticated'] = current_authenticated
        
        flash('Fehler beim Laden der E-Mail-Einstellungen.', 'error')
        return redirect(url_for('admin.dashboard'))

@bp.route('/switch-department/<department>')
@mitarbeiter_required
def switch_department(department):
    """Setzt das aktive Department in der Session, falls der Benutzer berechtigt ist."""
    try:
        user = mongodb.find_one('users', {'username': current_user.username})
        allowed = user.get('allowed_departments', []) if user else []
        if department in allowed:
            session['department'] = department
            flash(f'Aktives Department: {department}', 'success')
        else:
            flash('Keine Berechtigung für diese Abteilung', 'error')
    except Exception as e:
        logger = logging.getLogger(__name__)
        logger.error(f"Fehler beim Wechseln des Departments: {e}")
        flash('Fehler beim Wechseln des Departments', 'error')
    return redirect(request.referrer or url_for('main.index'))

@bp.route('/admin/email/diagnose', methods=['POST'])
@login_required
@admin_required
def diagnose_email():
    """Diagnostiziert die SMTP-Verbindung"""
    try:
        config_data = AdminEmailService.get_email_config()
        if not config_data:
            return jsonify({'success': False, 'message': 'Keine E-Mail-Konfiguration gefunden'})
        
        success, result = AdminEmailService.diagnose_smtp_connection(config_data)
        
        if success:
            return jsonify({
                'success': True,
                'message': 'SMTP-Diagnose erfolgreich',
                'diagnosis': result
            })
        else:
            return jsonify({
                'success': False,
                'message': result
            })
            
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Diagnose-Fehler: {str(e)}'
        })

@bp.route('/admin/email/test-simple', methods=['POST'])
@login_required
@admin_required
def test_email_simple():
    """Einfacher E-Mail-Test"""
    try:
        config_data = AdminEmailService.get_email_config()
        if not config_data:
            return jsonify({'success': False, 'message': 'Keine E-Mail-Konfiguration gefunden'})
        
        # Verwende die gleiche Logik wie im Debug-Tool
        from app.utils.email_utils import _decrypt_password
        
        # Erstelle Test-Konfiguration mit entschlüsseltem Passwort
        test_config = config_data.copy()
        
        # Entschlüssele Passwort falls verschlüsselt
        if test_config.get('mail_password') and test_config['mail_password'].startswith('gAAAAA'):
            try:
                decrypted_password = _decrypt_password(test_config['mail_password'])
                if decrypted_password:
                    test_config['mail_password'] = decrypted_password
                else:
                    return jsonify({
                        'success': False,
                        'message': 'Passwort konnte nicht entschlüsselt werden'
                    })
            except Exception as e:
                return jsonify({
                    'success': False,
                    'message': f'Fehler beim Entschlüsseln des Passworts: {str(e)}'
                })
        
        # Verwende die test_email_config aus email_utils direkt
        from app.utils.email_utils import test_email_config
        
        success, message = test_email_config(test_config)
        
        if success:
            return jsonify({
                'success': True,
                'message': message
            })
        else:
            return jsonify({
                'success': False,
                'message': message
            })
            
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Test-Fehler: {str(e)}'
        })

@bp.route('/backup/weekly/test', methods=['POST'])
@login_required
@admin_required
def test_weekly_backup():
    """Sendet das wöchentliche Backup-Archiv manuell"""
    try:
        from app.utils.auto_backup import auto_backup_scheduler
        
        # Führe wöchentliches Backup-Archiv manuell aus
        auto_backup_scheduler._create_weekly_backup_archive()
        
        return jsonify({
            'success': True,
            'message': 'Backup-Archiv erfolgreich erstellt, versendet und ZIP-Datei gelöscht'
        })
    except Exception as e:
        logger.error(f"Fehler beim Versenden des Backup-Archivs: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'Fehler beim Versenden: {str(e)}'
        })

@bp.route('/version/check', methods=['GET'])
@login_required
@admin_required
def check_version():
    """Prüft ob Updates verfügbar sind"""
    try:
        from app.utils.version_checker import check_version
        
        result = check_version()
        return jsonify(result)
        
    except Exception as e:
        logger.error(f"Fehler beim Versionscheck: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'Fehler beim Versionscheck: {str(e)}'
        }), 500

@bp.route('/version_check', methods=['GET'])
@login_required
def version_check():
    """Prüft ob Updates verfügbar sind (für alle Benutzer)"""
    try:
        from app.utils.version_checker import check_version
        
        result = check_version()
        
        # Vereinfachte Antwort für das Menü
        if result.get('status') == 'update_available':
            return jsonify({
                'update_available': True,
                'current_version': result.get('current_version'),
                'latest_version': result.get('latest_version')
            })
        else:
            return jsonify({
                'update_available': False,
                'current_version': result.get('current_version')
            })
        
    except Exception as e:
        logger.error(f"Fehler beim Versionscheck: {str(e)}")
        return jsonify({
            'update_available': False,
            'error': str(e)
        })

@bp.route('/version/info', methods=['GET'])
@login_required
@admin_required
def get_version_info():
    """Gibt detaillierte Versionsinformationen zurück"""
    try:
        from app.utils.version_checker import get_version_info
        
        info = get_version_info()
        return jsonify(info)
        
    except Exception as e:
        logger.error(f"Fehler beim Abrufen der Versionsinformationen: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'Fehler beim Abrufen der Versionsinformationen: {str(e)}'
        }), 500

@bp.route('/version', methods=['GET'])
@login_required
@admin_required
def version_check_page():
    """Versionscheck-Seite"""
    try:
        return render_template('admin/version_check.html')
    except Exception as e:
        logger.error(f"Fehler beim Laden der Versionscheck-Seite: {str(e)}")
        flash(f'Fehler beim Laden der Seite: {str(e)}', 'error')
        return redirect(url_for('admin.dashboard'))

@bp.route('/debug/fix-lending-inconsistencies', methods=['POST'])
@admin_required
def fix_lending_inconsistencies():
    """Behebt Inkonsistenzen in den Ausleihdaten"""
    try:
        from app.services.lending_service import LendingService
        
        success, message, statistics = LendingService.fix_lending_inconsistencies()
        
        if success:
            return jsonify({
                'success': True,
                'message': message,
                'statistics': statistics
            })
        else:
            return jsonify({
                'success': False,
                'message': message
            }), 500
        
    except Exception as e:
        logger.error(f"Fehler beim Beheben der Inkonsistenzen: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'Fehler beim Beheben der Inkonsistenzen: {str(e)}'
        }), 500

@bp.route('/export_excel_detailed')
@admin_required
def export_excel_detailed():
    """Exportiert alle Scandy-Daten in eine detaillierte Excel-Datei"""
    try:
        logger.info(f"Detaillierter Excel-Export gestartet von Benutzer: {current_user.username}")
        
        # Erstelle Excel-Export
        export_service = ExcelExportService()
        excel_file = export_service.generate_complete_export()
        
        # Generiere Dateinamen mit Zeitstempel
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'scandy_detailliert_{timestamp}.xlsx'
        
        logger.info(f"Detaillierter Excel-Export erfolgreich erstellt: {filename}")
        
        return send_file(
            excel_file,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        logger.error(f"Fehler beim detaillierten Excel-Export: {str(e)}", exc_info=True)
        flash('Fehler beim Generieren des detaillierten Excel-Exports', 'error')
        return redirect(url_for('admin.dashboard'))

@bp.route('/debug/validate-lending-consistency', methods=['GET'])
@admin_required
def validate_lending_consistency():
    """Validiert die Konsistenz der Ausleihdaten"""
    try:
        from app.services.lending_service import LendingService
        
        is_consistent, message, issues = LendingService.validate_lending_consistency()
        
        return jsonify({
            'success': True,
            'is_consistent': is_consistent,
            'message': message,
            'issues': issues
        })
        
    except Exception as e:
        logger.error(f"Fehler bei der Konsistenzprüfung: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'Fehler bei der Konsistenzprüfung: {str(e)}'
        }), 500

@bp.route('/debug/fix-missing-created-at', methods=['POST'])
@admin_required
def fix_missing_created_at():
    """Korrigiert fehlende created_at Felder in der Datenbank"""
    try:
        from app.services.admin_backup_service import AdminBackupService
        
        fixed_count = AdminBackupService._fix_missing_created_at_fields()
        
        return jsonify({
            'success': True,
            'message': f'{fixed_count} fehlende created_at Felder wurden ergänzt',
            'fixed_count': fixed_count
        })
        
    except Exception as e:
        logger.error(f"Fehler beim Korrigieren fehlender created_at Felder: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'Fehler beim Korrigieren: {str(e)}'
        }), 500

@bp.route('/debug/fix-backup-fields', methods=['GET', 'POST'])
@admin_required
def fix_backup_fields():
    """Korrigiert fehlende Felder in der Datenbank nach Backup-Restore"""
    try:
        from app.services.admin_backup_service import AdminBackupService
        
        # Führe die Korrektur aus
        fixed_count = AdminBackupService._fix_missing_created_at_fields()
        
        # Zusätzliche Korrekturen für Dashboard-Probleme
        dashboard_fixes = 0
        
        # Stelle sicher, dass alle Tools ein gültiges created_at Feld haben
        all_tools = mongodb.find('tools', {})
        for tool in all_tools:
            created_at = tool.get('created_at')
            if created_at is None:
                # Verwende updated_at oder aktuelles Datum
                fallback_date = tool.get('updated_at') or datetime.now()
                if isinstance(fallback_date, str):
                    try:
                        fallback_date = datetime.strptime(fallback_date, '%Y-%m-%d %H:%M:%S.%f')
                    except ValueError:
                        try:
                            fallback_date = datetime.strptime(fallback_date, '%Y-%m-%d %H:%M:%S')
                        except ValueError:
                            fallback_date = datetime.now()
                
                mongodb.update_one('tools', 
                                 {'_id': tool['_id']}, 
                                 {'$set': {'created_at': fallback_date}})
                dashboard_fixes += 1
        
        # Stelle sicher, dass alle Workers ein gültiges created_at Feld haben
        all_workers = mongodb.find('workers', {})
        for worker in all_workers:
            created_at = worker.get('created_at')
            if created_at is None:
                fallback_date = worker.get('updated_at') or datetime.now()
                if isinstance(fallback_date, str):
                    try:
                        fallback_date = datetime.strptime(fallback_date, '%Y-%m-%d %H:%M:%S.%f')
                    except ValueError:
                        try:
                            fallback_date = datetime.strptime(fallback_date, '%Y-%m-%d %H:%M:%S')
                        except ValueError:
                            fallback_date = datetime.now()
                
                mongodb.update_one('workers', 
                                 {'_id': worker['_id']}, 
                                 {'$set': {'created_at': fallback_date}})
                dashboard_fixes += 1
        
        total_fixes = fixed_count + dashboard_fixes
        
        return jsonify({
            'success': True,
            'message': f'{total_fixes} fehlende Felder wurden ergänzt (Backup-Service: {fixed_count}, Dashboard-Fixes: {dashboard_fixes})',
            'fixed_count': total_fixes,
            'backup_service_fixes': fixed_count,
            'dashboard_fixes': dashboard_fixes
        })
        
    except Exception as e:
        logger.error(f"Fehler beim Korrigieren der Backup-Felder: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'Fehler beim Korrigieren: {str(e)}'
        }), 500

@bp.route('/debug/test-dashboard-fix', methods=['GET'])
@admin_required
def test_dashboard_fix():
    """Testet und behebt Dashboard-Probleme nach Backup-Restore"""
    try:
        # Teste ob das Dashboard geladen werden kann
        from app.services.admin_dashboard_service import AdminDashboardService
        
        result = {
            'success': False,
            'message': '',
            'tests': {},
            'errors': []
        }
        
        # Teste alle Dashboard-Services
        services_to_test = [
            ('recent_activity', AdminDashboardService.get_recent_activity),
            ('material_usage', AdminDashboardService.get_material_usage),
            ('warnings', AdminDashboardService.get_warnings),
            ('consumables_forecast', AdminDashboardService.get_consumables_forecast),
            ('consumable_trend', AdminDashboardService.get_consumable_trend)
        ]
        
        working_services = 0
        for service_name, service_func in services_to_test:
            try:
                data = service_func()
                if data is not None:
                    if isinstance(data, list):
                        result['tests'][service_name] = len(data)
                    elif isinstance(data, dict):
                        result['tests'][service_name] = len(data.get('usage_data', [])) if 'usage_data' in data else len(data)
                    else:
                        result['tests'][service_name] = 'OK'
                    working_services += 1
                else:
                    result['tests'][service_name] = 'Keine Daten'
            except Exception as e:
                error_msg = f"{str(e)}"
                result['tests'][service_name] = f"Fehler: {error_msg}"
                result['errors'].append(f"{service_name}: {error_msg}")
        
        # Bewerte das Ergebnis
        if working_services == len(services_to_test) and not result['errors']:
            result['success'] = True
            result['message'] = 'Dashboard funktioniert einwandfrei!'
        elif working_services > 0:
            result['success'] = True
            result['message'] = f'Dashboard teilweise funktional. {working_services}/{len(services_to_test)} Services funktionieren.'
        else:
            result['message'] = 'Dashboard hat Probleme. Keine Services funktionieren.'
        
        return jsonify(result)
        
    except Exception as e:
        logger.error(f"Fehler beim Testen des Dashboard-Fixes: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'Fehler beim Testen: {str(e)}',
            'errors': [f"Unerwarteter Fehler: {str(e)}"],
            'tests': {}
        }), 500

@bp.route('/debug/fix-dashboard-complete', methods=['GET', 'POST'])
@admin_required
def fix_dashboard_complete():
    """Behebt alle Dashboard-Probleme umfassend"""
    try:
        from app.services.admin_backup_service import AdminBackupService
        from app.services.admin_dashboard_service import AdminDashboardService
        
        result = {
            'success': False,
            'message': '',
            'fixes': {
                'backup_fields': 0,
                'data_consistency': 0,
                'missing_relations': 0,
                'date_fixes': 0
            },
            'dashboard_tests': {},
            'errors': []
        }
        
        # 1. Backup-Felder korrigieren
        try:
            result['fixes']['backup_fields'] = AdminBackupService._fix_missing_created_at_fields()
        except Exception as e:
            result['errors'].append(f"Backup-Felder: {str(e)}")
        
        # 2. Datenkonsistenz prüfen und korrigieren
        collections_to_check = ['tools', 'workers', 'consumables', 'lendings', 'consumable_usages']
        
        for collection in collections_to_check:
            try:
                # Prüfe auf ungültige Datumsfelder
                date_fields = ['created_at', 'updated_at', 'lent_at', 'returned_at', 'used_at']
                for field in date_fields:
                    docs_with_invalid_dates = mongodb.find(collection, {
                        field: {'$exists': True, '$type': 'string'}
                    })
                    
                    for doc in docs_with_invalid_dates:
                        date_value = doc.get(field)
                        if isinstance(date_value, str):
                            try:
                                # Versuche verschiedene Datumsformate
                                if '.' in date_value:
                                    parsed_date = datetime.strptime(date_value, '%Y-%m-%d %H:%M:%S.%f')
                                elif 'T' in date_value:
                                    parsed_date = datetime.fromisoformat(date_value.replace('Z', '+00:00'))
                                else:
                                    parsed_date = datetime.strptime(date_value, '%Y-%m-%d %H:%M:%S')
                                
                                mongodb.update_one(collection, 
                                                 {'_id': doc['_id']}, 
                                                 {'$set': {field: parsed_date}})
                                result['fixes']['date_fixes'] += 1
                            except ValueError:
                                # Ungültiges Datum - setze auf aktuelles Datum
                                mongodb.update_one(collection, 
                                                 {'_id': doc['_id']}, 
                                                 {'$set': {field: datetime.now()}})
                                result['fixes']['date_fixes'] += 1
            except Exception as e:
                result['errors'].append(f"{collection} Datumsfelder: {str(e)}")
        
        # 3. Teste Dashboard-Services
        dashboard_services = [
            ('recent_activity', AdminDashboardService.get_recent_activity),
            ('material_usage', AdminDashboardService.get_material_usage),
            ('warnings', AdminDashboardService.get_warnings),
            ('consumables_forecast', AdminDashboardService.get_consumables_forecast),
            ('consumable_trend', AdminDashboardService.get_consumable_trend)
        ]
        
        working_services = 0
        for service_name, service_func in dashboard_services:
            try:
                data = service_func()
                if data is not None:
                    if isinstance(data, list):
                        result['dashboard_tests'][service_name] = len(data)
                    elif isinstance(data, dict):
                        result['dashboard_tests'][service_name] = len(data.get('usage_data', [])) if 'usage_data' in data else len(data)
                    else:
                        result['dashboard_tests'][service_name] = 'OK'
                    working_services += 1
                else:
                    result['dashboard_tests'][service_name] = 'Keine Daten'
            except Exception as e:
                error_msg = f"{str(e)}"
                result['dashboard_tests'][service_name] = f"Fehler: {error_msg}"
                result['errors'].append(f"{service_name}: {error_msg}")
        
        # 4. Bewerte das Ergebnis
        total_fixes = sum(result['fixes'].values())
        
        if working_services == len(dashboard_services) and not result['errors']:
            result['success'] = True
            result['message'] = f'Dashboard-Korrektur abgeschlossen: {total_fixes} Probleme behoben'
        elif working_services > 0:
            result['success'] = True
            result['message'] = f'Dashboard teilweise repariert: {total_fixes} Probleme behoben, {working_services}/{len(dashboard_services)} Services funktionieren'
        else:
            result['message'] = f'Dashboard konnte nicht repariert werden: {total_fixes} Probleme behoben, aber Services funktionieren nicht'
        
        return jsonify(result)
        
    except Exception as e:
        logger.error(f"Fehler bei der umfassenden Dashboard-Korrektur: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'Fehler bei der Dashboard-Korrektur: {str(e)}',
            'errors': [f"Unerwarteter Fehler: {str(e)}"],
            'fixes': {'backup_fields': 0, 'data_consistency': 0, 'missing_relations': 0, 'date_fixes': 0},
            'dashboard_tests': {}
        }), 500

@bp.route('/debug/dashboard-status', methods=['GET'])
@admin_required
def dashboard_status():
    """Zeigt den aktuellen Dashboard-Status und behebt Probleme automatisch"""
    try:
        from app.services.admin_dashboard_service import AdminDashboardService
        
        status = {
            'dashboard_working': False,
            'errors': [],
            'fixes_applied': 0,
            'data_counts': {}
        }
        
        # Teste Dashboard-Services
        try:
            recent_activity = AdminDashboardService.get_recent_activity()
            status['data_counts']['recent_activity'] = len(recent_activity)
        except Exception as e:
            status['errors'].append(f"Recent Activity: {str(e)}")
        
        try:
            material_usage = AdminDashboardService.get_material_usage()
            status['data_counts']['material_usage'] = len(material_usage.get('usage_data', []))
        except Exception as e:
            status['errors'].append(f"Material Usage: {str(e)}")
        
        try:
            warnings = AdminDashboardService.get_warnings()
            status['data_counts']['warnings'] = sum(len(w) for w in warnings.values())
        except Exception as e:
            status['errors'].append(f"Warnings: {str(e)}")
        
        try:
            consumables_forecast = AdminDashboardService.get_consumables_forecast()
            status['data_counts']['consumables_forecast'] = len(consumables_forecast)
        except Exception as e:
            status['errors'].append(f"Consumables Forecast: {str(e)}")
        
        try:
            consumable_trend = AdminDashboardService.get_consumable_trend()
            status['data_counts']['consumable_trend'] = len(consumable_trend.get('labels', []))
        except Exception as e:
            status['errors'].append(f"Consumable Trend: {str(e)}")
        
        # Prüfe Datenbank-Zugriff
        try:
            total_tools = mongodb.count_documents('tools', {'deleted': {'$ne': True}})
            total_consumables = mongodb.count_documents('consumables', {'deleted': {'$ne': True}})
            total_workers = mongodb.count_documents('workers', {'deleted': {'$ne': True}})
            total_lendings = mongodb.count_documents('lendings', {})
            
            status['data_counts'].update({
                'tools': total_tools,
                'consumables': total_consumables,
                'workers': total_workers,
                'lendings': total_lendings
            })
        except Exception as e:
            status['errors'].append(f"Database Access: {str(e)}")
        
        # Wenn es Fehler gibt, versuche automatische Korrektur
        if status['errors']:
            try:
                from app.services.admin_backup_service import AdminBackupService
                
                # Führe Backup-Feld-Korrektur aus
                fixed_count = AdminBackupService._fix_missing_created_at_fields()
                status['fixes_applied'] = fixed_count
                
                # Teste erneut nach der Korrektur
                retry_success = True
                for error in status['errors'][:]:  # Kopie für Iteration
                    if 'Recent Activity' in error:
                        try:
                            recent_activity = AdminDashboardService.get_recent_activity()
                            status['data_counts']['recent_activity'] = len(recent_activity)
                            status['errors'].remove(error)
                        except:
                            retry_success = False
                
                if retry_success and not status['errors']:
                    status['dashboard_working'] = True
                    status['message'] = f'Dashboard repariert! {fixed_count} Felder korrigiert'
                else:
                    status['message'] = f'Teilweise repariert. {fixed_count} Felder korrigiert, aber {len(status["errors"])} Fehler verbleiben'
            except Exception as fix_error:
                status['errors'].append(f"Auto-Fix failed: {str(fix_error)}")
                status['message'] = 'Automatische Reparatur fehlgeschlagen'
        else:
            status['dashboard_working'] = True
            status['message'] = 'Dashboard funktioniert einwandfrei'
        
        return jsonify(status)
        
    except Exception as e:
        logger.error(f"Fehler beim Prüfen des Dashboard-Status: {str(e)}")
        return jsonify({
            'dashboard_working': False,
            'errors': [f"Status check failed: {str(e)}"],
            'fixes_applied': 0,
            'data_counts': {},
            'message': 'Fehler beim Prüfen des Dashboard-Status'
        }), 500

@bp.route('/debug/dashboard-details', methods=['GET'])
@admin_required
def dashboard_details():
    """Zeigt detaillierte Informationen über Dashboard-Probleme"""
    try:
        from app.services.admin_dashboard_service import AdminDashboardService
        
        details = {
            'database_counts': {},
            'service_tests': {},
            'template_variables': {},
            'errors': []
        }
        
        # Datenbank-Zählungen
        try:
            details['database_counts'] = {
                'tools': mongodb.count_documents('tools', {'deleted': {'$ne': True}}),
                'consumables': mongodb.count_documents('consumables', {'deleted': {'$ne': True}}),
                'workers': mongodb.count_documents('workers', {'deleted': {'$ne': True}}),
                'lendings': mongodb.count_documents('lendings', {}),
                'consumable_usages': mongodb.count_documents('consumable_usages', {}),
                'tickets': mongodb.count_documents('tickets', {})
            }
        except Exception as e:
            details['errors'].append(f"Database counts: {str(e)}")
        
        # Service-Tests
        try:
            recent_activity = AdminDashboardService.get_recent_activity()
            details['service_tests']['recent_activity'] = {
                'count': len(recent_activity),
                'sample': recent_activity[:2] if recent_activity else []
            }
        except Exception as e:
            details['service_tests']['recent_activity'] = {'error': str(e)}
        
        try:
            material_usage = AdminDashboardService.get_material_usage()
            details['service_tests']['material_usage'] = {
                'count': len(material_usage.get('usage_data', [])),
                'period_days': material_usage.get('period_days', 0)
            }
        except Exception as e:
            details['service_tests']['material_usage'] = {'error': str(e)}
        
        try:
            warnings = AdminDashboardService.get_warnings()
            details['service_tests']['warnings'] = {
                'defect_tools': len(warnings.get('defect_tools', [])),
                'overdue_lendings': len(warnings.get('overdue_lendings', [])),
                'low_stock_consumables': len(warnings.get('low_stock_consumables', []))
            }
        except Exception as e:
            details['service_tests']['warnings'] = {'error': str(e)}
        
        try:
            consumables_forecast = AdminDashboardService.get_consumables_forecast()
            details['service_tests']['consumables_forecast'] = {
                'count': len(consumables_forecast)
            }
        except Exception as e:
            details['service_tests']['consumables_forecast'] = {'error': str(e)}
        
        try:
            consumable_trend = AdminDashboardService.get_consumable_trend()
            details['service_tests']['consumable_trend'] = {
                'labels_count': len(consumable_trend.get('labels', [])),
                'datasets_count': len(consumable_trend.get('datasets', []))
            }
        except Exception as e:
            details['service_tests']['consumable_trend'] = {'error': str(e)}
        
        # Template-Variablen simulieren
        try:
            # Simuliere Dashboard-Route Logik
            total_tools = mongodb.count_documents('tools', {'deleted': {'$ne': True}})
            total_consumables = mongodb.count_documents('consumables', {'deleted': {'$ne': True}})
            total_workers = mongodb.count_documents('workers', {'deleted': {'$ne': True}})
            total_tickets = mongodb.count_documents('tickets', {})
            
            # Tool-Statistiken
            tool_stats = {
                'total': total_tools,
                'available': mongodb.count_documents('tools', {'status': 'verfügbar', 'deleted': {'$ne': True}}),
                'lent': mongodb.count_documents('tools', {'status': 'ausgeliehen', 'deleted': {'$ne': True}}),
                'defect': mongodb.count_documents('tools', {'status': 'defekt', 'deleted': {'$ne': True}})
            }
            
            # Consumable-Statistiken
            consumables = list(mongodb.find('consumables', {'deleted': {'$ne': True}}))
            sufficient = 0
            warning = 0
            critical = 0
            
            for consumable in consumables:
                if consumable['quantity'] >= consumable.get('warning_threshold', 10):
                    sufficient += 1
                elif consumable['quantity'] >= consumable.get('critical_threshold', 5):
                    warning += 1
                else:
                    critical += 1
            
            consumable_stats = {
                'total': total_consumables,
                'sufficient': sufficient,
                'warning': warning,
                'critical': critical
            }
            
            # Worker-Statistiken
            workers = list(mongodb.find('workers', {'deleted': {'$ne': True}}))
            worker_stats = {
                'total': total_workers,
                'by_department': []
            }
            
            # Gruppiere nach Abteilung
            dept_counts = {}
            for worker in workers:
                dept = worker.get('department', 'Ohne Abteilung')
                dept_counts[dept] = dept_counts.get(dept, 0) + 1
            
            for dept, count in dept_counts.items():
                worker_stats['by_department'].append({
                    'name': dept,
                    'count': count
                })
            
            details['template_variables'] = {
                'total_tools': total_tools,
                'total_consumables': total_consumables,
                'total_workers': total_workers,
                'total_tickets': total_tickets,
                'tool_stats': tool_stats,
                'consumable_stats': consumable_stats,
                'worker_stats': worker_stats
            }
            
        except Exception as e:
            details['errors'].append(f"Template variables: {str(e)}")
        
        return jsonify({
            'success': True,
            'details': details,
            'message': 'Dashboard-Details erfolgreich geladen'
        })
        
    except Exception as e:
        logger.error(f"Fehler beim Laden der Dashboard-Details: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'Fehler beim Laden der Dashboard-Details: {str(e)}'
        }), 500

@bp.route('/debug/dashboard-page')
@admin_required
def dashboard_debug_page():
    """Dashboard Debug-Seite"""
    return render_template('admin/dashboard_debug.html')

@bp.route('/debug/test-backup-restore/<filename>', methods=['GET'])
@admin_required
def test_backup_restore(filename):
    """Testet die Backup-Wiederherstellung für eine spezifische Datei"""
    try:
        from app.utils.backup_manager import backup_manager
        
        # Prüfe ob die Datei existiert
        backup_path = backup_manager.backup_dir / filename
        if not backup_path.exists():
            return jsonify({
                'success': False,
                'message': f'Backup-Datei "{filename}" nicht gefunden',
                'backup_path': str(backup_path)
            })
        
        # Teste die Backup-Wiederherstellung ohne sie tatsächlich durchzuführen
        try:
            with open(backup_path, 'r', encoding='utf-8') as f:
                backup_data = json.load(f)
            
            # Validiere Backup-Daten
            is_valid, validation_message = backup_manager._validate_backup_data(backup_data)
            
            return jsonify({
                'success': True,
                'message': f'Backup "{filename}" ist gültig',
                'validation_message': validation_message,
                'collections': list(backup_data.keys()),
                'total_documents': sum(len(docs) for docs in backup_data.values()),
                'file_size': backup_path.stat().st_size
            })
            
        except Exception as e:
            return jsonify({
                'success': False,
                'message': f'Fehler beim Testen des Backups: {str(e)}',
                'backup_path': str(backup_path)
            })
            
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Unerwarteter Fehler: {str(e)}'
        }), 500

@bp.route('/debug/email-debug-page')
@admin_required
def email_debug_page():
    """E-Mail Debug-Seite"""
    return render_template('admin/email_debug.html')

@bp.route('/debug/test-email-config', methods=['GET'])
@admin_required
def test_email_config_debug():
    """Testet die E-Mail-Konfiguration mit detaillierten Informationen"""
    try:
        from app.utils.email_utils import get_email_config, test_email_config, _decrypt_password
        
        result = {
            'success': False,
            'message': '',
            'config_loaded': False,
            'config_details': {},
            'test_result': {},
            'errors': []
        }
        
        # 1. Lade E-Mail-Konfiguration
        try:
            config = get_email_config()
            if config:
                result['config_loaded'] = True
                result['config_details'] = {
                    'mail_server': config.get('mail_server'),
                    'mail_port': config.get('mail_port'),
                    'mail_use_tls': config.get('mail_use_tls'),
                    'mail_username': config.get('mail_username'),
                    'mail_password_length': len(config.get('mail_password', '')),
                    'mail_password_encrypted': config.get('mail_password', '').startswith('gAAAAA'),
                    'test_email': config.get('test_email'),
                    'use_auth': config.get('use_auth')
                }
            else:
                result['errors'].append("E-Mail-Konfiguration konnte nicht geladen werden")
        except Exception as e:
            result['errors'].append(f"Fehler beim Laden der Konfiguration: {str(e)}")
        
        # 2. Teste E-Mail-Konfiguration
        if config:
            try:
                # Erstelle Test-Konfiguration mit entschlüsseltem Passwort
                test_config = config.copy()
                
                # Entschlüssele Passwort falls verschlüsselt
                if test_config.get('mail_password') and test_config['mail_password'].startswith('gAAAAA'):
                    try:
                        decrypted_password = _decrypt_password(test_config['mail_password'])
                        if decrypted_password:
                            test_config['mail_password'] = decrypted_password
                            result['config_details']['password_decrypted'] = True
                        else:
                            result['errors'].append("Passwort konnte nicht entschlüsselt werden")
                    except Exception as e:
                        result['errors'].append(f"Fehler beim Entschlüsseln des Passworts: {str(e)}")
                
                # Teste E-Mail-Konfiguration
                success, message = test_email_config(test_config)
                
                result['test_result'] = {
                    'success': success,
                    'message': message,
                    'config_used': {
                        'mail_server': test_config.get('mail_server'),
                        'mail_port': test_config.get('mail_port'),
                        'mail_use_tls': test_config.get('mail_use_tls'),
                        'mail_username': test_config.get('mail_username'),
                        'mail_password_length': len(test_config.get('mail_password', '')),
                        'test_email': test_config.get('test_email')
                    }
                }
                
                if success:
                    result['success'] = True
                    result['message'] = "E-Mail-Konfiguration funktioniert korrekt!"
                else:
                    result['message'] = f"E-Mail-Test fehlgeschlagen: {message}"
                    
            except Exception as e:
                result['errors'].append(f"Fehler beim E-Mail-Test: {str(e)}")
                result['message'] = f"Fehler beim E-Mail-Test: {str(e)}"
        else:
            result['message'] = "Keine E-Mail-Konfiguration verfügbar"
        
        return jsonify(result)
        
    except Exception as e:
        logger.error(f"Fehler beim E-Mail-Konfigurations-Test: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'Kritischer Fehler: {str(e)}',
            'config_loaded': False,
            'config_details': {},
            'test_result': {},
            'errors': [f"Unerwarteter Fehler: {str(e)}"]
        }), 500

@bp.route('/debug/fix-email-config', methods=['GET'])
@admin_required
def fix_email_config():
    """Korrigiert E-Mail-Konfigurationsprobleme nach Backup-Restore"""
    try:
        result = {
            'success': False,
            'message': '',
            'fixes_applied': 0,
            'old_settings': {},
            'new_config': {},
            'errors': []
        }
        
        # 1. Prüfe alte E-Mail-Einstellungen in settings Collection
        try:
            old_settings = {}
            settings_docs = mongodb.find('settings', {'key': {'$regex': '^email_'}})
            
            for doc in settings_docs:
                old_settings[doc['key']] = doc.get('value', '')
            
            result['old_settings'] = old_settings
            
            # 2. Prüfe neue E-Mail-Konfiguration in email_config Collection
            try:
                new_config_doc = mongodb.find_one('email_config', {'_id': 'email_config'})
                if new_config_doc:
                    result['new_config'] = {k: v for k, v in new_config_doc.items() if k != '_id'}
            except Exception as e:
                result['errors'].append(f"Neue E-Mail-Konfiguration prüfen: {str(e)}")
            
            # 3. Migriere alte Einstellungen zu neuem Format
            if old_settings and not result['new_config']:
                try:
                    # Konvertiere alte Einstellungen zu neuem Format
                    new_config = {
                        'mail_server': old_settings.get('email_smtp_server', 'smtp.gmail.com'),
                        'mail_port': int(old_settings.get('email_smtp_port', 587)),
                        'mail_use_tls': old_settings.get('email_use_tls', 'true').lower() == 'true',
                        'mail_username': old_settings.get('email_username', ''),
                        'mail_password': old_settings.get('email_password', ''),
                        'test_email': old_settings.get('email_test_email', ''),
                        'use_auth': old_settings.get('email_use_auth', 'true').lower() == 'true'
                    }
                    
                    # Speichere neue Konfiguration
                    mongodb.update_one('email_config', 
                                     {'_id': 'email_config'}, 
                                     {'$set': new_config}, 
                                     upsert=True)
                    
                    result['new_config'] = new_config
                    result['fixes_applied'] += 1
                    result['message'] += "E-Mail-Konfiguration von altem Format migriert. "
                    
                except Exception as e:
                    result['errors'].append(f"Migration fehlgeschlagen: {str(e)}")
            
            # 4. Prüfe ob Admin-Benutzer E-Mail-Adresse hat
            try:
                admin_users = list(mongodb.find('users', {'role': 'admin'}))
                admin_without_email = 0
                
                for admin in admin_users:
                    if not admin.get('email'):
                        admin_without_email += 1
                        # Setze Standard-E-Mail für Admin ohne E-Mail
                        mongodb.update_one('users', 
                                         {'_id': admin['_id']}, 
                                         {'$set': {'email': 'admin@scandy.local'}})
                
                if admin_without_email > 0:
                    result['fixes_applied'] += admin_without_email
                    result['message'] += f"{admin_without_email} Admin-Benutzer ohne E-Mail-Adresse korrigiert. "
                    
            except Exception as e:
                result['errors'].append(f"Admin-E-Mail prüfen: {str(e)}")
            
            # 5. Bewerte das Ergebnis
            if result['fixes_applied'] > 0 and not result['errors']:
                result['success'] = True
                result['message'] += "E-Mail-Konfiguration erfolgreich repariert!"
            elif result['fixes_applied'] > 0:
                result['success'] = True
                result['message'] += "E-Mail-Konfiguration teilweise repariert."
            else:
                result['message'] = "Keine E-Mail-Konfigurationsprobleme gefunden."
            
        except Exception as e:
            result['errors'].append(f"Allgemeiner Fehler: {str(e)}")
            result['message'] = "Fehler beim Prüfen der E-Mail-Konfiguration."
        
        return jsonify(result)
        
    except Exception as e:
        logger.error(f"Fehler bei E-Mail-Konfigurations-Fix: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'Kritischer Fehler: {str(e)}',
            'fixes_applied': 0,
            'errors': [f"Unerwarteter Fehler: {str(e)}"],
            'old_settings': {},
            'new_config': {}
        }), 500

@bp.route('/debug/analyze-lendings', methods=['GET'])
@admin_required
def analyze_lendings():
    """Analysiert die Ausleihen in der Datenbank"""
    try:
        from app.models.mongodb_database import mongodb
        
        # Hole alle Ausleihen
        all_lendings = list(mongodb.find('lendings', {}))
        current_lendings = list(mongodb.find('lendings', {'returned_at': {'$exists': False}}))
        
        # Analysiere Duplikate
        lending_counts = {}
        for lending in current_lendings:
            tool_barcode = lending.get('tool_barcode')
            if tool_barcode:
                lending_counts[tool_barcode] = lending_counts.get(tool_barcode, 0) + 1
        
        duplicate_lendings = {barcode: count for barcode, count in lending_counts.items() if count > 1}
        
        # Erstelle detaillierte Analyse
        analysis = {
            'total_lendings': len(all_lendings),
            'current_lendings': len(current_lendings),
            'unique_tools_lent': len(lending_counts),
            'duplicate_lendings': duplicate_lendings,
            'lending_details': []
        }
        
        # Detaillierte Ausleihen
        for lending in current_lendings:
            tool = mongodb.find_one('tools', {'barcode': lending.get('tool_barcode')})
            worker = mongodb.find_one('workers', {'barcode': lending.get('worker_barcode')})
            
            analysis['lending_details'].append({
                'id': str(lending.get('_id')),
                'tool_barcode': lending.get('tool_barcode'),
                'tool_name': tool.get('name', 'Unbekannt') if tool else 'Unbekannt',
                'worker_barcode': lending.get('worker_barcode'),
                'worker_name': worker.get('name', 'Unbekannt') if worker else 'Unbekannt',
                'lent_at': lending.get('lent_at'),
                'returned_at': lending.get('returned_at')
            })
        
        return jsonify({
            'success': True,
            'analysis': analysis
        })
        
    except Exception as e:
        logger.error(f"Fehler bei Ausleihen-Analyse: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@bp.route('/debug/fix-dashboard-simple', methods=['GET'])
@admin_required
def fix_dashboard_simple():
    """Einfache Dashboard-Korrektur mit detaillierten Informationen"""
    try:
        import traceback
        from app.services.admin_backup_service import AdminBackupService
        from app.services.admin_dashboard_service import AdminDashboardService
        
        result = {
            'success': False,
            'message': '',
            'fixes_applied': 0,
            'errors': [],
            'tests': {},
            'database_info': {}
        }
        
        # 1. Sammle Datenbank-Informationen
        try:
            result['database_info'] = {
                'tools_count': mongodb.count_documents('tools', {'deleted': {'$ne': True}}),
                'consumables_count': mongodb.count_documents('consumables', {'deleted': {'$ne': True}}),
                'workers_count': mongodb.count_documents('workers', {'deleted': {'$ne': True}}),
                'lendings_count': mongodb.count_documents('lendings', {}),
                'tickets_count': mongodb.count_documents('tickets', {})
            }
        except Exception as e:
            result['errors'].append(f"Datenbank-Zugriff: {str(e)}")
        
        # 2. Führe Backup-Feld-Korrektur aus
        try:
            fixed_count = AdminBackupService._fix_missing_created_at_fields()
            result['fixes_applied'] = fixed_count
            result['message'] += f"{fixed_count} fehlende Felder korrigiert. "
        except Exception as e:
            result['errors'].append(f"Backup-Feld-Korrektur: {str(e)}")
        
        # 3. Teste Dashboard-Services
        dashboard_services = [
            ('recent_activity', AdminDashboardService.get_recent_activity),
            ('material_usage', AdminDashboardService.get_material_usage),
            ('warnings', AdminDashboardService.get_warnings),
            ('consumables_forecast', AdminDashboardService.get_consumables_forecast),
            ('consumable_trend', AdminDashboardService.get_consumable_trend)
        ]
        
        working_services = 0
        for service_name, service_func in dashboard_services:
            try:
                data = service_func()
                if data is not None:
                    if isinstance(data, list):
                        result['tests'][service_name] = len(data)
                    elif isinstance(data, dict):
                        result['tests'][service_name] = len(data.get('usage_data', [])) if 'usage_data' in data else len(data)
                    else:
                        result['tests'][service_name] = 'OK'
                    working_services += 1
                else:
                    result['tests'][service_name] = 'Keine Daten'
            except Exception as e:
                error_msg = f"{str(e)}"
                result['tests'][service_name] = f"Fehler: {error_msg}"
                result['errors'].append(f"{service_name}: {error_msg}")
        
        # 4. Bewerte das Ergebnis
        if working_services == len(dashboard_services) and not result['errors']:
            result['success'] = True
            result['message'] += "Dashboard funktioniert einwandfrei!"
        elif working_services > 0:
            result['success'] = True
            result['message'] += f"Dashboard teilweise repariert. {working_services}/{len(dashboard_services)} Services funktionieren."
        else:
            result['message'] += "Dashboard konnte nicht repariert werden."
        
        return jsonify(result)
        
    except Exception as e:
        logger.error(f"Fehler bei der einfachen Dashboard-Korrektur: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'Kritischer Fehler: {str(e)}',
            'fixes_applied': 0,
            'errors': [f"Unerwarteter Fehler: {str(e)}"],
            'tests': {},
            'database_info': {}
        }), 500

@bp.route('/test/json-backups')
def test_json_backups():
    """Test-Route für JSON-Backups"""
    try:
        from app.services.admin_backup_service import AdminBackupService
        
        json_backups = AdminBackupService.get_backup_list()
        
        return jsonify({
            'status': 'success',
            'json_backups': json_backups,
            'count': len(json_backups)
        })
    except Exception as e:
        return jsonify({
            'status': 'error',
            'message': str(e)
        }), 500

























































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































































