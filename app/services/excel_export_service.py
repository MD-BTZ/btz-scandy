"""
Excel Export Service für Scandy
Exportiert alle wichtigen Daten in eine strukturierte Excel-Datei
"""
import io
from datetime import datetime
from typing import Dict, List, Any, BinaryIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.worksheet import Worksheet
from app.models.mongodb_database import mongodb
import logging

logger = logging.getLogger(__name__)

class ExcelExportService:
    """Service für Excel-Export aller Scandy-Daten"""
    
    def __init__(self):
        self.workbook = None
        self.styles = self._create_styles()
    
    def _create_styles(self) -> Dict[str, Any]:
        """Erstellt Styling für Excel-Tabellen"""
        return {
            'header': {
                'font': Font(bold=True, color='FFFFFF'),
                'fill': PatternFill(start_color='366092', end_color='366092', fill_type='solid'),
                'alignment': Alignment(horizontal='center', vertical='center'),
                'border': Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
            },
            'data': {
                'border': Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                ),
                'alignment': Alignment(vertical='center')
            },
            'highlight': {
                'fill': PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
            },
            'error': {
                'fill': PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
            }
        }
    
    def generate_complete_export(self) -> BinaryIO:
        """
        Generiert eine komplette Excel-Datei mit allen Scandy-Daten
        
        Returns:
            BinaryIO: Excel-Datei als BytesIO-Stream
        """
        try:
            self.workbook = Workbook()
            
            # Entferne das Standard-Arbeitsblatt
            self.workbook.remove(self.workbook.active)
            
            # Erstelle alle Arbeitsblätter
            self._create_tools_sheet()
            self._create_consumables_sheet()
            self._create_workers_sheet()
            self._create_lendings_sheet()
            self._create_consumptions_sheet()
            self._create_tickets_sheet()
            self._create_summary_sheet()
            
            # Speichere in BytesIO
            output = io.BytesIO()
            self.workbook.save(output)
            output.seek(0)
            
            return output
            
        except Exception as e:
            logger.error(f"Fehler beim Generieren des Excel-Exports: {str(e)}")
            raise
    
    def _create_tools_sheet(self):
        """Erstellt das Werkzeuge-Arbeitsblatt"""
        try:
            ws = self.workbook.create_sheet("Werkzeuge")
            
            # Header
            headers = [
                'Barcode', 'Name', 'Kategorie', 'Standort', 'Beschreibung', 
                'Status', 'Verfügbar', 'Ausgeliehen an', 'Ausgeliehen seit', 
                'Rückgabe bis', 'Erstellt am', 'Aktualisiert am'
            ]
            
            # Schreibe Header
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                self._apply_style(cell, 'header')
            
            # Lade Werkzeuge
            tools = list(mongodb.find('tools', {'deleted': {'$ne': True}}, sort=[('name', 1)]))
            
            # Schreibe Daten
            for row, tool in enumerate(tools, 2):
                # Hole aktuelle Ausleihe
                current_lending = mongodb.find_one('lendings', {
                    'tool_barcode': tool.get('barcode'),
                    'returned_at': None
                })
                
                # Hole Mitarbeiter-Info falls ausgeliehen
                lent_to = None
                lent_since = None
                return_date = None
                
                if current_lending:
                    worker = mongodb.find_one('workers', {
                        'barcode': current_lending.get('worker_barcode'),
                        'deleted': {'$ne': True}
                    })
                    if worker:
                        lent_to = f"{worker.get('firstname', '')} {worker.get('lastname', '')}"
                    lent_since = current_lending.get('lent_at')
                    return_date = current_lending.get('expected_return_date')
                
                # Bestimme Status
                status = tool.get('status', 'verfügbar')
                if current_lending and return_date:
                    if isinstance(return_date, str):
                        try:
                            return_date_dt = datetime.strptime(return_date, '%Y-%m-%d')
                        except ValueError:
                            return_date_dt = None
                    else:
                        return_date_dt = return_date
                    
                    if return_date_dt and return_date_dt.date() < datetime.now().date():
                        status = 'überfällig'
                
                data = [
                    tool.get('barcode', ''),
                    tool.get('name', ''),
                    tool.get('category', ''),
                    tool.get('location', ''),
                    tool.get('description', ''),
                    status,
                    'Nein' if current_lending else 'Ja',
                    lent_to or '',
                    lent_since.strftime('%d.%m.%Y %H:%M') if lent_since else '',
                    return_date.strftime('%d.%m.%Y') if return_date else '',
                    tool.get('created_at', '').strftime('%d.%m.%Y') if tool.get('created_at') else '',
                    tool.get('updated_at', '').strftime('%d.%m.%Y') if tool.get('updated_at') else ''
                ]
                
                for col, value in enumerate(data, 1):
                    cell = ws.cell(row=row, column=col, value=value)
                    self._apply_style(cell, 'data')
                    
                    # Markiere überfällige Werkzeuge
                    if status == 'überfällig':
                        self._apply_style(cell, 'error')
            
            self._auto_adjust_columns(ws)
            
        except Exception as e:
            logger.error(f"Fehler beim Erstellen des Werkzeuge-Arbeitsblatts: {str(e)}")
    
    def _create_consumables_sheet(self):
        """Erstellt das Verbrauchsmaterial-Arbeitsblatt"""
        try:
            ws = self.workbook.create_sheet("Verbrauchsmaterial")
            
            headers = [
                'Barcode', 'Name', 'Kategorie', 'Standort', 'Beschreibung',
                'Bestand', 'Mindestbestand', 'Status', 'Einheit',
                'Erstellt am', 'Aktualisiert am'
            ]
            
            # Schreibe Header
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                self._apply_style(cell, 'header')
            
            # Lade Verbrauchsmaterial
            consumables = list(mongodb.find('consumables', {'deleted': {'$ne': True}}, sort=[('name', 1)]))
            
            # Schreibe Daten
            for row, consumable in enumerate(consumables, 2):
                quantity = consumable.get('quantity', 0)
                min_quantity = consumable.get('min_quantity', 0)
                
                # Bestimme Status
                if quantity <= 0:
                    status = 'Aufgebraucht'
                elif quantity <= min_quantity:
                    status = 'Niedrig'
                else:
                    status = 'Ausreichend'
                
                data = [
                    consumable.get('barcode', ''),
                    consumable.get('name', ''),
                    consumable.get('category', ''),
                    consumable.get('location', ''),
                    consumable.get('description', ''),
                    quantity,
                    min_quantity,
                    status,
                    consumable.get('unit', ''),
                    consumable.get('created_at', '').strftime('%d.%m.%Y') if consumable.get('created_at') else '',
                    consumable.get('updated_at', '').strftime('%d.%m.%Y') if consumable.get('updated_at') else ''
                ]
                
                for col, value in enumerate(data, 1):
                    cell = ws.cell(row=row, column=col, value=value)
                    self._apply_style(cell, 'data')
                    
                    # Markiere niedrige/aufgebrauchte Bestände
                    if status in ['Niedrig', 'Aufgebraucht']:
                        self._apply_style(cell, 'highlight' if status == 'Niedrig' else 'error')
            
            self._auto_adjust_columns(ws)
            
        except Exception as e:
            logger.error(f"Fehler beim Erstellen des Verbrauchsmaterial-Arbeitsblatts: {str(e)}")
    
    def _create_workers_sheet(self):
        """Erstellt das Mitarbeiter-Arbeitsblatt"""
        try:
            ws = self.workbook.create_sheet("Mitarbeiter")
            
            headers = [
                'Barcode', 'Vorname', 'Nachname', 'Abteilung', 'Position',
                'E-Mail', 'Telefon', 'Status', 'Handlungsfelder',
                'Erstellt am', 'Aktualisiert am'
            ]
            
            # Schreibe Header
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                self._apply_style(cell, 'header')
            
            # Lade Mitarbeiter
            workers = list(mongodb.find('workers', {'deleted': {'$ne': True}}, sort=[('lastname', 1), ('firstname', 1)]))
            
            # Schreibe Daten
            for row, worker in enumerate(workers, 2):
                handlungsfelder = ', '.join(worker.get('handlungsfelder', [])) if worker.get('handlungsfelder') else ''
                
                data = [
                    worker.get('barcode', ''),
                    worker.get('firstname', ''),
                    worker.get('lastname', ''),
                    worker.get('department', ''),
                    worker.get('position', ''),
                    worker.get('email', ''),
                    worker.get('phone', ''),
                    'Aktiv' if worker.get('is_active', True) else 'Inaktiv',
                    handlungsfelder,
                    worker.get('created_at', '').strftime('%d.%m.%Y') if worker.get('created_at') else '',
                    worker.get('updated_at', '').strftime('%d.%m.%Y') if worker.get('updated_at') else ''
                ]
                
                for col, value in enumerate(data, 1):
                    cell = ws.cell(row=row, column=col, value=value)
                    self._apply_style(cell, 'data')
                    
                    # Markiere inaktive Mitarbeiter
                    if not worker.get('is_active', True):
                        self._apply_style(cell, 'highlight')
            
            self._auto_adjust_columns(ws)
            
        except Exception as e:
            logger.error(f"Fehler beim Erstellen des Mitarbeiter-Arbeitsblatts: {str(e)}")
    
    def _create_lendings_sheet(self):
        """Erstellt das Ausleihen-Arbeitsblatt"""
        try:
            ws = self.workbook.create_sheet("Ausleihen")
            
            headers = [
                'Werkzeug Barcode', 'Werkzeug Name', 'Mitarbeiter Barcode',
                'Mitarbeiter Name', 'Ausgeliehen am', 'Geplante Rückgabe',
                'Zurückgegeben am', 'Status', 'Tage ausgeliehen'
            ]
            
            # Schreibe Header
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                self._apply_style(cell, 'header')
            
            # Lade Ausleihen
            lendings = list(mongodb.find('lendings', {}, sort=[('lent_at', -1)]))
            
            # Schreibe Daten
            for row, lending in enumerate(lendings, 2):
                # Hole Werkzeug-Info
                tool = mongodb.find_one('tools', {'barcode': lending.get('tool_barcode')})
                tool_name = tool.get('name', 'Unbekannt') if tool else 'Unbekannt'
                
                # Hole Mitarbeiter-Info
                worker = mongodb.find_one('workers', {'barcode': lending.get('worker_barcode')})
                worker_name = f"{worker.get('firstname', '')} {worker.get('lastname', '')}" if worker else 'Unbekannt'
                
                # Berechne Status und Tage
                lent_at = lending.get('lent_at')
                returned_at = lending.get('returned_at')
                expected_return = lending.get('expected_return_date')
                
                if returned_at:
                    status = 'Zurückgegeben'
                    if lent_at:
                        days = (returned_at - lent_at).days
                    else:
                        days = 0
                elif expected_return:
                    if isinstance(expected_return, str):
                        try:
                            expected_return_dt = datetime.strptime(expected_return, '%Y-%m-%d')
                        except ValueError:
                            expected_return_dt = None
                    else:
                        expected_return_dt = expected_return
                    
                    if expected_return_dt and expected_return_dt.date() < datetime.now().date():
                        status = 'Überfällig'
                    else:
                        status = 'Ausgeliehen'
                    
                    if lent_at:
                        days = (datetime.now() - lent_at).days
                    else:
                        days = 0
                else:
                    status = 'Ausgeliehen'
                    if lent_at:
                        days = (datetime.now() - lent_at).days
                    else:
                        days = 0
                
                data = [
                    lending.get('tool_barcode', ''),
                    tool_name,
                    lending.get('worker_barcode', ''),
                    worker_name,
                    lent_at.strftime('%d.%m.%Y %H:%M') if lent_at else '',
                    expected_return.strftime('%d.%m.%Y') if expected_return else '',
                    returned_at.strftime('%d.%m.%Y %H:%M') if returned_at else '',
                    status,
                    days
                ]
                
                for col, value in enumerate(data, 1):
                    cell = ws.cell(row=row, column=col, value=value)
                    self._apply_style(cell, 'data')
                    
                    # Markiere überfällige Ausleihen
                    if status == 'Überfällig':
                        self._apply_style(cell, 'error')
            
            self._auto_adjust_columns(ws)
            
        except Exception as e:
            logger.error(f"Fehler beim Erstellen des Ausleihen-Arbeitsblatts: {str(e)}")
    
    def _create_consumptions_sheet(self):
        """Erstellt das Ausgaben-Arbeitsblatt"""
        try:
            ws = self.workbook.create_sheet("Ausgaben")
            
            headers = [
                'Verbrauchsmaterial Barcode', 'Verbrauchsmaterial Name',
                'Mitarbeiter Barcode', 'Mitarbeiter Name', 'Menge',
                'Einheit', 'Ausgegeben am', 'Grund/Notiz'
            ]
            
            # Schreibe Header
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                self._apply_style(cell, 'header')
            
            # Lade Ausgaben
            consumptions = list(mongodb.find('consumptions', {}, sort=[('consumed_at', -1)]))
            
            # Schreibe Daten
            for row, consumption in enumerate(consumptions, 2):
                # Hole Verbrauchsmaterial-Info
                consumable = mongodb.find_one('consumables', {'barcode': consumption.get('consumable_barcode')})
                consumable_name = consumable.get('name', 'Unbekannt') if consumable else 'Unbekannt'
                unit = consumable.get('unit', '') if consumable else ''
                
                # Hole Mitarbeiter-Info
                worker = mongodb.find_one('workers', {'barcode': consumption.get('worker_barcode')})
                worker_name = f"{worker.get('firstname', '')} {worker.get('lastname', '')}" if worker else 'Unbekannt'
                
                data = [
                    consumption.get('consumable_barcode', ''),
                    consumable_name,
                    consumption.get('worker_barcode', ''),
                    worker_name,
                    consumption.get('quantity', 0),
                    unit,
                    consumption.get('consumed_at', '').strftime('%d.%m.%Y %H:%M') if consumption.get('consumed_at') else '',
                    consumption.get('reason', '')
                ]
                
                for col, value in enumerate(data, 1):
                    cell = ws.cell(row=row, column=col, value=value)
                    self._apply_style(cell, 'data')
            
            self._auto_adjust_columns(ws)
            
        except Exception as e:
            logger.error(f"Fehler beim Erstellen des Ausgaben-Arbeitsblatts: {str(e)}")
    
    def _create_tickets_sheet(self):
        """Erstellt das Tickets-Arbeitsblatt"""
        try:
            ws = self.workbook.create_sheet("Tickets")
            
            headers = [
                'Ticket Nr.', 'Titel', 'Kategorie', 'Priorität', 'Status',
                'Erstellt von', 'Zugewiesen an', 'Abteilung',
                'Erstellt am', 'Aktualisiert am', 'Beschreibung'
            ]
            
            # Schreibe Header
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                self._apply_style(cell, 'header')
            
            # Lade Tickets
            tickets = list(mongodb.find('tickets', {}, sort=[('created_at', -1)]))
            
            # Schreibe Daten
            for row, ticket in enumerate(tickets, 2):
                data = [
                    ticket.get('ticket_number', ''),
                    ticket.get('title', ''),
                    ticket.get('category', ''),
                    ticket.get('priority', ''),
                    ticket.get('status', ''),
                    ticket.get('created_by', ''),
                    ticket.get('assigned_to', ''),
                    ticket.get('department', ''),
                    ticket.get('created_at', '').strftime('%d.%m.%Y %H:%M') if ticket.get('created_at') else '',
                    ticket.get('updated_at', '').strftime('%d.%m.%Y %H:%M') if ticket.get('updated_at') else '',
                    ticket.get('description', '')
                ]
                
                for col, value in enumerate(data, 1):
                    cell = ws.cell(row=row, column=col, value=value)
                    self._apply_style(cell, 'data')
                    
                    # Markiere offene Tickets
                    if ticket.get('status') == 'offen':
                        self._apply_style(cell, 'highlight')
            
            self._auto_adjust_columns(ws)
            
        except Exception as e:
            logger.error(f"Fehler beim Erstellen des Tickets-Arbeitsblatts: {str(e)}")
    
    def _create_summary_sheet(self):
        """Erstellt das Zusammenfassung-Arbeitsblatt"""
        try:
            ws = self.workbook.create_sheet("Zusammenfassung", 0)  # Füge am Anfang hinzu
            
            # Titel
            ws.cell(row=1, column=1, value="Scandy - Datenexport Zusammenfassung")
            title_cell = ws.cell(row=1, column=1)
            title_cell.font = Font(size=16, bold=True)
            ws.merge_cells('A1:D1')
            
            # Export-Info
            ws.cell(row=2, column=1, value=f"Exportiert am: {datetime.now().strftime('%d.%m.%Y %H:%M')}")
            
            # Statistiken sammeln
            tool_count = mongodb.count_documents('tools', {'deleted': {'$ne': True}})
            consumable_count = mongodb.count_documents('consumables', {'deleted': {'$ne': True}})
            worker_count = mongodb.count_documents('workers', {'deleted': {'$ne': True}})
            lending_count = mongodb.count_documents('lendings', {})
            active_lending_count = mongodb.count_documents('lendings', {'returned_at': None})
            consumption_count = mongodb.count_documents('consumptions', {})
            ticket_count = mongodb.count_documents('tickets', {})
            
            # Statistiken anzeigen
            stats = [
                ('Werkzeuge:', tool_count),
                ('Verbrauchsmaterial:', consumable_count),
                ('Mitarbeiter:', worker_count),
                ('Ausleihen (gesamt):', lending_count),
                ('Aktive Ausleihen:', active_lending_count),
                ('Ausgaben:', consumption_count),
                ('Tickets:', ticket_count)
            ]
            
            row = 4
            for label, count in stats:
                ws.cell(row=row, column=1, value=label)
                ws.cell(row=row, column=2, value=count)
                row += 1
            
            # Arbeitsblatt-Beschreibungen
            descriptions = [
                ('Werkzeuge:', 'Alle Werkzeuge mit Status und Ausleihinformationen'),
                ('Verbrauchsmaterial:', 'Alle Verbrauchsmaterialien mit Bestandsinformationen'),
                ('Mitarbeiter:', 'Alle Mitarbeiter mit Kontaktdaten und Status'),
                ('Ausleihen:', 'Alle Werkzeug-Ausleihen mit Zeiten und Status'),
                ('Ausgaben:', 'Alle Verbrauchsmaterial-Ausgaben'),
                ('Tickets:', 'Alle Support-Tickets und Anfragen')
            ]
            
            row += 2
            ws.cell(row=row, column=1, value="Arbeitsblätter:")
            header_cell = ws.cell(row=row, column=1)
            header_cell.font = Font(bold=True)
            row += 1
            
            for label, desc in descriptions:
                ws.cell(row=row, column=1, value=label)
                ws.cell(row=row, column=2, value=desc)
                row += 1
            
            self._auto_adjust_columns(ws)
            
        except Exception as e:
            logger.error(f"Fehler beim Erstellen des Zusammenfassungs-Arbeitsblatts: {str(e)}")
    
    def _apply_style(self, cell, style_name: str):
        """Wendet einen Stil auf eine Zelle an"""
        if style_name in self.styles:
            style = self.styles[style_name]
            if 'font' in style:
                cell.font = style['font']
            if 'fill' in style:
                cell.fill = style['fill']
            if 'alignment' in style:
                cell.alignment = style['alignment']
            if 'border' in style:
                cell.border = style['border']
    
    def _auto_adjust_columns(self, worksheet: Worksheet):
        """Passt die Spaltenbreite automatisch an"""
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)  # Maximal 50 Zeichen
            worksheet.column_dimensions[column_letter].width = adjusted_width